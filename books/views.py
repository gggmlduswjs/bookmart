import json
import math
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import Workbook, load_workbook

from accounts.decorators import role_required
from .forms import PublisherForm, BookForm
from .models import Publisher, Book


# ── 교재 목록 ──────────────────────────────────────────────────────────────────

@role_required('admin')
def book_list(request):
    from accounts.models import User
    publishers = Publisher.objects.filter(is_active=True).order_by('name')
    all_books = (Book.objects.select_related('publisher')
                 .prefetch_related('agencies')
                 .order_by('series', 'month', 'grade', 'sort_order', 'name'))
    series_list = sorted(set(b.series for b in all_books if b.series))
    agency_list = User.objects.filter(role='agency', is_active=True).order_by('name')

    for book in all_books:
        book.agency_ids_str = ','.join(str(a.pk) for a in book.agencies.all())

    return render(request, 'books/book_list.html', {
        'publishers': publishers,
        'all_books': all_books,
        'series_list': series_list,
        'months': range(1, 13),
        'agency_list': agency_list,
        'total_count': all_books.count(),
    })


# ── 출판사 CRUD ────────────────────────────────────────────────────────────────

@role_required('admin')
def publisher_list(request):
    publishers = Publisher.objects.all()
    return render(request, 'books/publisher_list.html', {'publishers': publishers})


@role_required('admin')
def publisher_create(request):
    if request.method == 'POST':
        form = PublisherForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '출판사가 등록되었습니다.')
            return redirect('publisher_list')
    else:
        form = PublisherForm()
    return render(request, 'books/publisher_form.html', {'form': form, 'title': '출판사 등록'})


@role_required('admin')
def publisher_edit(request, pk):
    publisher = get_object_or_404(Publisher, pk=pk)
    if request.method == 'POST':
        form = PublisherForm(request.POST, instance=publisher)
        if form.is_valid():
            form.save()
            messages.success(request, '출판사 정보가 수정되었습니다.')
            return redirect('publisher_list')
    else:
        form = PublisherForm(instance=publisher)
    return render(request, 'books/publisher_form.html', {'form': form, 'title': '출판사 수정'})


# ── 교재 CRUD ──────────────────────────────────────────────────────────────────

@role_required('admin')
def book_create(request):
    if request.method == 'POST':
        form = BookForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '교재가 등록되었습니다.')
            return redirect('book_list')
    else:
        form = BookForm()
    return render(request, 'books/book_form.html', {'form': form, 'title': '교재 등록'})


@role_required('admin')
def book_edit(request, pk):
    book = get_object_or_404(Book, pk=pk)
    if request.method == 'POST':
        form = BookForm(request.POST, instance=book)
        if form.is_valid():
            form.save()
            messages.success(request, '교재 정보가 수정되었습니다.')
            return redirect('book_list')
    else:
        form = BookForm(instance=book)
    return render(request, 'books/book_form.html', {'form': form, 'title': '교재 수정'})


@role_required('admin')
def book_toggle(request, pk):
    book = get_object_or_404(Book, pk=pk)
    book.is_active = not book.is_active
    book.save(update_fields=['is_active'])
    action = '활성화' if book.is_active else '비활성화'
    messages.success(request, f'[{book.name}] {action}되었습니다.')
    return redirect('book_list')


@role_required('admin')
def book_delete(request, pk):
    book = get_object_or_404(Book, pk=pk)
    if request.method == 'POST':
        name = book.name
        try:
            book.delete()
            messages.success(request, f'[{name}] 교재를 삭제했습니다.')
        except Exception:
            messages.error(request, f'[{name}] 주문에 사용된 교재는 삭제할 수 없습니다. 비활성화를 이용하세요.')
    return redirect('book_list')


@role_required('admin')
def book_bulk_agencies(request):
    """벌크 취급 업체 지정 API"""
    import json as _json
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    try:
        data = _json.loads(request.body)
        book_ids = data.get('book_ids', [])
        agency_ids = data.get('agency_ids', [])
        books = Book.objects.filter(pk__in=book_ids)
        for book in books:
            book.agencies.set(agency_ids)
        return JsonResponse({'ok': True, 'count': books.count()})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@role_required('admin')
def book_bulk_delete(request):
    if request.method != 'POST':
        return redirect('book_list')
    ids = request.POST.getlist('ids')
    if ids:
        deleted = 0
        skipped = 0
        for book in Book.objects.filter(pk__in=ids):
            try:
                book.delete()
                deleted += 1
            except Exception:
                skipped += 1
        if deleted:
            messages.success(request, f'{deleted}건의 교재를 삭제했습니다.')
        if skipped:
            messages.warning(request, f'{skipped}건은 주문에 사용되어 삭제할 수 없습니다.')
    return redirect('book_list')


# ── 주문 폼용 드롭다운 (로그인만 있으면 됨) ────────────────────────────────────

@login_required
def book_options(request):
    series = request.GET.get('series', '')
    row = request.GET.get('row', '0')
    books = (Book.objects.filter(series=series, is_active=True).select_related('publisher')
             if series else Book.objects.none())
    return render(request, 'books/partials/book_options.html', {'books': books, 'row': row})


# ── 교재 엑셀 일괄 등록 ───────────────────────────────────────────────────────

@role_required('admin')
def book_import(request):
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file or not file.name.endswith('.xlsx'):
            messages.error(request, '.xlsx 파일만 업로드할 수 있습니다.')
            return redirect('book_import')

        try:
            wb = load_workbook(file, read_only=True)
            ws = wb.active

            created, updated, skipped = 0, 0, 0

            # 헤더 자동 감지: 'NO.' 또는 '교재명' 포함된 행 찾기
            header_row = 0
            col_map = {}  # 컬럼 매핑
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                cells = [str(c or '').strip() for c in row]
                # 표준 형식: 출판사 | 시리즈 | 교재명 | 정가 | 반품가능 | 월 | 학년
                if '출판사' in cells and '교재명' in cells:
                    col_map = {
                        'publisher': cells.index('출판사'),
                        'series': cells.index('시리즈') if '시리즈' in cells else None,
                        'name': cells.index('교재명'),
                        'price': cells.index('정가'),
                        'returnable': cells.index('반품가능(O/X)') if '반품가능(O/X)' in cells else None,
                        'month': cells.index('월') if '월' in cells else None,
                        'grade': cells.index('학년') if '학년' in cells else None,
                    }
                    header_row = i + 1
                    break
                # 교재리스트 형식: NO. | 시리즈명 | 교재명 | 출판사 | 정가
                if 'NO.' in cells and ('교 재 명' in cells or '교재명' in cells):
                    name_label = '교 재 명' if '교 재 명' in cells else '교재명'
                    series_label = '시리즈명' if '시리즈명' in cells else None
                    col_map = {
                        'publisher': cells.index('출판사'),
                        'series': cells.index(series_label) if series_label else None,
                        'name': cells.index(name_label),
                        'price': cells.index('정가'),
                        'returnable': None,
                        'month': None,
                        'grade': None,
                    }
                    header_row = i + 1
                    break

            if not col_map:
                messages.error(request, '엑셀 형식을 인식할 수 없습니다. 헤더에 "출판사", "교재명", "정가" 컬럼이 필요합니다.')
                return redirect('book_import')

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < header_row:
                    continue

                cells = list(row)
                pub_name = str(cells[col_map['publisher']] or '').strip() if col_map['publisher'] < len(cells) else ''
                book_name = str(cells[col_map['name']] or '').strip() if col_map['name'] < len(cells) else ''
                series = str(cells[col_map['series']] or '').strip() if col_map.get('series') is not None and col_map['series'] < len(cells) else ''
                price_raw = cells[col_map['price']] if col_map['price'] < len(cells) else None

                if not pub_name or not book_name:
                    continue

                try:
                    list_price = int(price_raw)
                except (TypeError, ValueError):
                    skipped += 1
                    continue

                is_returnable = True
                if col_map.get('returnable') is not None and col_map['returnable'] < len(cells):
                    returnable_raw = str(cells[col_map['returnable']] or '').strip().upper()
                    is_returnable = returnable_raw != 'X'

                publisher, _ = Publisher.objects.get_or_create(
                    name=pub_name,
                    defaults={'supply_rate': Decimal('40.00')},
                )

                # 월, 학년 파싱
                month_val = None
                if col_map.get('month') is not None and col_map['month'] < len(cells):
                    try:
                        month_val = int(cells[col_map['month']])
                        if month_val < 1 or month_val > 12:
                            month_val = None
                    except (TypeError, ValueError):
                        month_val = None
                grade_val = ''
                if col_map.get('grade') is not None and col_map['grade'] < len(cells):
                    grade_raw = str(cells[col_map['grade']] or '').strip()
                    if grade_raw in ('1', '1학년'):
                        grade_val = '1'
                    elif grade_raw in ('2', '2학년'):
                        grade_val = '2'

                book, is_new = Book.objects.get_or_create(
                    publisher=publisher,
                    name=book_name,
                    defaults={
                        'series': series,
                        'list_price': list_price,
                        'is_returnable': is_returnable,
                        'month': month_val,
                        'grade': grade_val,
                    },
                )

                if is_new:
                    created += 1
                else:
                    book.series = series
                    book.list_price = list_price
                    book.is_returnable = is_returnable
                    if month_val is not None:
                        book.month = month_val
                    if grade_val:
                        book.grade = grade_val
                    book.save(update_fields=['series', 'list_price', 'is_returnable', 'month', 'grade'])
                    updated += 1

            wb.close()
            msg = f'엑셀 등록 완료: 신규 {created}건, 업데이트 {updated}건'
            if skipped:
                msg += f', 건너뜀 {skipped}건'
            messages.success(request, msg)
            return redirect('book_list')

        except Exception as e:
            messages.error(request, f'파일 처리 중 오류가 발생했습니다: {e}')
            return redirect('book_import')

    return render(request, 'books/book_import.html')


@role_required('admin')
def book_import_sample(request):
    wb = Workbook()
    ws = wb.active
    ws.title = '교재 일괄등록'

    headers = ['출판사', '시리즈', '교재명', '정가', '반품가능(O/X)', '월', '학년']
    ws.append(headers)

    ws.append(['오은라이프사이언스', '마음이야기', '마음이 뭘까, 사회성이 뭘까', 8000, 'O', 3, '통합'])
    ws.append(['오은라이프사이언스', '연산수학', '더하기 6, 7', 6000, 'O', 3, '1학년'])
    ws.append(['오은라이프사이언스', '연산수학', '두 자리 수의 덧셈', 6000, 'O', 3, '2학년'])

    for col in range(1, 8):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="book_import_sample.xlsx"'
    wb.save(response)
    return response
