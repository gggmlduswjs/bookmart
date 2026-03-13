import io
import json
from collections import defaultdict
from datetime import date

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render

from accounts.decorators import role_required
from accounts.models import User
from books.models import Book
from orders.models import (
    Order, OrderItem, Payment, Return, ReturnItem,
)
from orders.services import get_order_queryset
from orders.services.import_legacy import import_geukdong_data, parse_geukdong_excel


# ── 거래내역 (V03) ─────────────────────────────────────────────────────────────

@login_required
def ledger(request):
    today = date.today()
    if request.user.role == 'admin':
        agencies = User.objects.filter(role='agency', is_active=True).order_by('name')
        category_filter = request.GET.get('category', '')
        agency_categories = (
            User.objects.filter(role='agency', is_active=True, agency_category__gt='')
            .values_list('agency_category', flat=True).distinct().order_by('agency_category')
        )
        if category_filter:
            agencies = agencies.filter(agency_category=category_filter)
        agency_id = request.GET.get('agency', '')

        # 업체 미선택 시 → 전체 업체 잔액 요약 모드
        if not agency_id:
            from django.db.models import Sum as DSum
            agency_summaries = []
            for agency in agencies:
                sales = OrderItem.objects.filter(
                    order__agency=agency,
                    order__status__in=['shipping', 'delivered'],
                ).aggregate(total=DSum('amount'))['total'] or 0
                returns = ReturnItem.objects.filter(
                    ret__agency=agency,
                    ret__status='confirmed',
                ).aggregate(total=DSum('confirmed_amount'))['total'] or 0
                paid = Payment.objects.filter(agency=agency).aggregate(
                    total=DSum('amount'))['total'] or 0
                bal = sales - returns - paid
                if sales > 0 or bal != 0:
                    agency_summaries.append({
                        'pk': agency.pk, 'name': agency.name,
                        'category': agency.agency_category,
                        'sales': sales, 'returns': returns,
                        'paid': paid, 'balance': bal,
                    })
            agency_summaries.sort(key=lambda x: -x['balance'])
            year = int(request.GET.get('year', today.year))
            month = int(request.GET.get('month', today.month))
            return render(request, 'orders/ledger.html', {
                'mode': 'summary',
                'agency_summaries': agency_summaries,
                'total_balance': sum(a['balance'] for a in agency_summaries),
                'agencies': agencies,
                'agency_categories': agency_categories,
                'category_filter': category_filter,
                'agency_id': '',
                'year': year, 'month': month,
                'years': range(today.year - 2, today.year + 1),
                'months': range(1, 13),
            })

        selected_agency = get_object_or_404(User, pk=agency_id, role='agency')
    else:
        agencies = None
        agency_categories = []
        category_filter = ''
        selected_agency = request.user
        agency_id = str(request.user.pk)

    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))

    rows = []
    total_sales = total_returns = total_paid = 0

    if selected_agency:
        order_items = (
            OrderItem.objects
            .filter(
                order__agency=selected_agency,
                order__status__in=[Order.Status.SHIPPING, Order.Status.DELIVERED],
                order__ordered_at__year=year,
                order__ordered_at__month=month,
            )
            .select_related('order', 'order__delivery', 'book', 'book__publisher')
            .order_by('order__ordered_at')
        )
        for oi in order_items:
            rows.append({
                'date': oi.order.ordered_at.date(),
                'type': '매출',
                'delivery': oi.order.delivery.name,
                'publisher': oi.display_publisher,
                'book': oi.display_name,
                'qty': oi.quantity,
                'list_price': oi.list_price,
                'supply_rate': oi.supply_rate,
                'amount': oi.amount,
            })
            total_sales += oi.amount

        return_items = (
            ReturnItem.objects
            .filter(
                ret__agency=selected_agency,
                ret__status=Return.Status.CONFIRMED,
                ret__confirmed_at__year=year,
                ret__confirmed_at__month=month,
            )
            .select_related('ret', 'ret__delivery', 'book', 'book__publisher')
            .order_by('ret__confirmed_at')
        )
        for ri in return_items:
            confirmed_amount = ri.confirmed_amount or 0
            rows.append({
                'date': ri.ret.confirmed_at.date(),
                'type': '반품',
                'delivery': ri.ret.delivery.name,
                'publisher': ri.book.publisher.name if ri.book else '',
                'book': ri.book.name if ri.book else '',
                'qty': -(ri.confirmed_qty or 0),
                'list_price': ri.list_price,
                'supply_rate': ri.supply_rate,
                'amount': -confirmed_amount,
                'return_pk': ri.ret.pk,
            })
            total_returns += confirmed_amount

        payments = Payment.objects.filter(
            agency=selected_agency,
            paid_at__year=year,
            paid_at__month=month,
        ).order_by('paid_at')
        for p in payments:
            total_paid += p.amount

        rows.sort(key=lambda r: (r['delivery'], r['date']))

    balance = total_sales - total_returns - total_paid

    # 학교별 그룹핑
    from itertools import groupby
    grouped_rows = []
    for delivery_name, items in groupby(rows, key=lambda r: r['delivery']):
        items_list = list(items)
        subtotal = sum(r['amount'] for r in items_list)
        grouped_rows.append({
            'delivery': delivery_name,
            'items': items_list,
            'subtotal': subtotal,
        })

    return render(request, 'orders/ledger.html', {
        'mode': 'detail',
        'agencies': agencies,
        'selected_agency': selected_agency,
        'agency_id': agency_id,
        'year': year,
        'month': month,
        'years': range(today.year - 2, today.year + 1),
        'months': range(1, 13),
        'rows': rows,
        'grouped_rows': grouped_rows,
        'total_sales': total_sales,
        'total_returns': total_returns,
        'total_paid': total_paid,
        'balance': balance,
        'payments': payments if selected_agency else [],
        'agency_categories': agency_categories,
        'category_filter': category_filter,
    })


# ── 판매현황 (V04) ─────────────────────────────────────────────────────────────

@login_required
def sales_report(request):
    today = date.today()
    date_from = request.GET.get('date_from', today.strftime('%Y-%m-01'))
    date_to = request.GET.get('date_to', today.strftime('%Y-%m-%d'))

    qs = OrderItem.objects.filter(
        order__status__in=[Order.Status.SHIPPING, Order.Status.DELIVERED],
        order__ordered_at__date__gte=date_from,
        order__ordered_at__date__lte=date_to,
    ).select_related('order', 'order__teacher', 'order__delivery', 'order__agency',
                     'book', 'book__publisher').order_by('order__ordered_at')

    agency_filter = request.GET.get('agency', '')
    category_filter = request.GET.get('category', '')
    agencies = None
    agency_categories = []

    if request.user.role == 'admin':
        agencies = User.objects.filter(role='agency', is_active=True).order_by('name')
        agency_categories = (
            User.objects.filter(role='agency', is_active=True, agency_category__gt='')
            .values_list('agency_category', flat=True).distinct().order_by('agency_category')
        )
        if agency_filter:
            qs = qs.filter(order__agency_id=agency_filter)
        if category_filter:
            qs = qs.filter(order__agency__agency_category=category_filter)
    elif request.user.role == 'agency':
        qs = qs.filter(order__agency=request.user)
    elif request.user.role == 'teacher':
        qs = qs.filter(order__teacher=request.user)

    total_amount = sum(i.amount for i in qs)

    return render(request, 'orders/sales_report.html', {
        'items': qs,
        'date_from': date_from,
        'date_to': date_to,
        'total_amount': total_amount,
        'agencies': agencies,
        'agency_categories': agency_categories,
        'agency_filter': agency_filter,
        'category_filter': category_filter,
    })


# ── 발주 집계 (V05, 총판 전용) ─────────────────────────────────────────────────

@role_required('admin')
def purchase_order(request):
    today = date.today()
    date_from = request.GET.get('date_from', today.strftime('%Y-%m-01'))
    date_to = request.GET.get('date_to', today.strftime('%Y-%m-%d'))

    items = (
        OrderItem.objects
        .filter(
            order__status__in=[Order.Status.PENDING, Order.Status.SHIPPING, Order.Status.DELIVERED],
            order__ordered_at__date__gte=date_from,
            order__ordered_at__date__lte=date_to,
        )
        .select_related('book', 'book__publisher')
        .order_by('book__publisher__name', 'book__series', 'book__name')
    )

    publishers = defaultdict(lambda: {'books': defaultdict(lambda: {'name': '', 'series': '', 'qty': 0, 'amount': 0})})
    for item in items:
        pub = item.display_publisher or '기타'
        book_id = item.book.pk if item.book else f'custom_{item.pk}'
        publishers[pub]['books'][book_id]['name'] = item.display_name
        publishers[pub]['books'][book_id]['series'] = item.display_series
        publishers[pub]['books'][book_id]['qty'] += item.quantity
        publishers[pub]['books'][book_id]['amount'] += item.amount

    pub_list = []
    for pub_name, data in sorted(publishers.items()):
        book_rows = sorted(data['books'].values(), key=lambda b: (b['series'], b['name']))
        pub_total_qty = sum(b['qty'] for b in book_rows)
        pub_total_amount = sum(b['amount'] for b in book_rows)
        pub_list.append({
            'name': pub_name,
            'books': book_rows,
            'total_qty': pub_total_qty,
            'total_amount': pub_total_amount,
        })

    return render(request, 'orders/purchase_order.html', {
        'publishers': pub_list,
        'date_from': date_from,
        'date_to': date_to,
        'grand_total': sum(p['total_amount'] for p in pub_list),
    })


# ── 입금 등록 (총판 전용) ──────────────────────────────────────────────────────

@role_required('admin')
def payment_create(request):
    agencies = User.objects.filter(role='agency', is_active=True).order_by('name')
    if request.method == 'POST':
        agency_id = request.POST.get('agency')
        amount_str = request.POST.get('amount', '0').replace(',', '')
        paid_at = request.POST.get('paid_at', '')
        memo = request.POST.get('memo', '')
        try:
            agency = User.objects.get(pk=agency_id, role='agency')
            amount = int(amount_str)
            Payment.objects.create(agency=agency, amount=amount, paid_at=paid_at, memo=memo)
            messages.success(request, f'{agency.name} 입금 {amount:,}원 등록 완료.')
            return redirect('ledger')
        except (User.DoesNotExist, ValueError) as e:
            messages.error(request, '입력값을 확인해주세요.')
    return render(request, 'orders/payment_form.html', {'agencies': agencies})


@role_required('admin')
def payment_create_inline(request):
    """거래원장에서 인라인 입금 등록 (AJAX)"""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'ok': False, 'error': '잘못된 요청'}, status=400)

    agency_id = data.get('agency_id')
    amount_str = str(data.get('amount', '0')).replace(',', '')
    paid_at = data.get('paid_at', '')
    memo = data.get('memo', '')

    try:
        agency = User.objects.get(pk=agency_id, role='agency')
        amount = int(amount_str)
        if amount <= 0:
            return JsonResponse({'ok': False, 'error': '금액을 입력해주세요.'})
        Payment.objects.create(agency=agency, amount=amount, paid_at=paid_at, memo=memo)
        return JsonResponse({'ok': True, 'message': f'{agency.name} 입금 {amount:,}원 등록 완료'})
    except User.DoesNotExist:
        return JsonResponse({'ok': False, 'error': '업체를 찾을 수 없습니다.'}, status=400)
    except (ValueError, TypeError):
        return JsonResponse({'ok': False, 'error': '금액을 확인해주세요.'}, status=400)


# ── 엑셀 Export ────────────────────────────────────────────────────────────────

def _make_workbook():
    try:
        import openpyxl
        return openpyxl, openpyxl.Workbook()
    except ImportError:
        return None, None


@login_required
def export_ledger(request):
    openpyxl, wb = _make_workbook()
    if not wb:
        messages.error(request, 'openpyxl 패키지가 필요합니다: uv add openpyxl')
        return redirect('ledger')

    today = date.today()
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))
    agency_id = request.GET.get('agency', '')

    if request.user.role == 'admin' and agency_id:
        selected_agency = get_object_or_404(User, pk=agency_id, role='agency')
    elif request.user.role == 'agency':
        selected_agency = request.user
    else:
        return redirect('ledger')

    ws = wb.active
    ws.title = '거래내역'
    ws.append(['날짜', '구분', '배송지', '출판사', '교재명', '수량', '정가', '공급률', '금액'])

    order_items = OrderItem.objects.filter(
        order__agency=selected_agency,
        order__status__in=[Order.Status.SHIPPING, Order.Status.DELIVERED],
        order__ordered_at__year=year, order__ordered_at__month=month,
    ).select_related('order', 'order__delivery', 'book', 'book__publisher').order_by('order__ordered_at')

    for oi in order_items:
        ws.append([
            oi.order.ordered_at.strftime('%Y-%m-%d'), '매출',
            oi.order.delivery.name, oi.display_publisher, oi.display_name,
            oi.quantity, oi.list_price, float(oi.supply_rate), oi.amount,
        ])

    return_items = ReturnItem.objects.filter(
        ret__agency=selected_agency, ret__status=Return.Status.CONFIRMED,
        ret__confirmed_at__year=year, ret__confirmed_at__month=month,
    ).select_related('ret', 'ret__delivery', 'book', 'book__publisher').order_by('ret__confirmed_at')

    for ri in return_items:
        ws.append([
            ri.ret.confirmed_at.strftime('%Y-%m-%d'), '반품',
            ri.ret.delivery.name, ri.book.publisher.name if ri.book else '', ri.book.name if ri.book else '',
            -(ri.confirmed_qty or 0), ri.list_price, float(ri.supply_rate),
            -(ri.confirmed_amount or 0),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'거래내역_{selected_agency.name}_{year}년{month}월.xlsx'
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
    return resp


@login_required
def export_sales(request):
    openpyxl, wb = _make_workbook()
    if not wb:
        messages.error(request, 'openpyxl 패키지가 필요합니다: uv add openpyxl')
        return redirect('sales_report')

    today = date.today()
    date_from = request.GET.get('date_from', today.strftime('%Y-%m-01'))
    date_to = request.GET.get('date_to', today.strftime('%Y-%m-%d'))

    qs = OrderItem.objects.filter(
        order__status__in=[Order.Status.SHIPPING, Order.Status.DELIVERED],
        order__ordered_at__date__gte=date_from,
        order__ordered_at__date__lte=date_to,
    ).select_related('order', 'order__teacher', 'order__delivery', 'order__agency',
                     'book', 'book__publisher').order_by('order__ordered_at')

    if request.user.role == 'agency':
        qs = qs.filter(order__agency=request.user)

    ws = wb.active
    ws.title = '판매현황'
    ws.append(['출고일', '업체', '배송지', '선생님', '출판사', '시리즈', '교재명', '수량', '정가', '공급률', '금액'])
    for oi in qs:
        ws.append([
            oi.order.ordered_at.strftime('%Y-%m-%d'),
            oi.order.agency.name, oi.order.delivery.name, oi.order.teacher.name,
            oi.display_publisher, oi.display_series, oi.display_name,
            oi.quantity, oi.list_price, float(oi.supply_rate), oi.amount,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'판매현황_{date_from}_{date_to}.xlsx'
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
    return resp


@login_required
def export_orders(request):
    """주문 목록 엑셀 내보내기"""
    openpyxl, wb = _make_workbook()
    if not wb:
        messages.error(request, 'openpyxl 패키지가 필요합니다.')
        return redirect('order_list')

    qs = get_order_queryset(request.user).filter(is_deleted=False)
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if status:
        qs = qs.filter(status=status)
    if date_from:
        qs = qs.filter(ordered_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(ordered_at__date__lte=date_to)

    qs = qs.order_by('-ordered_at').prefetch_related('items__book__publisher')

    ws = wb.active
    ws.title = '주문목록'
    ws.append(['주문번호', '업체', '선생님', '배송지', '상태', '주문일시', '교재명', '수량', '단가', '금액', '메모'])
    for order in qs:
        for item in order.items.all():
            ws.append([
                order.order_no, order.agency.name, order.teacher.name,
                order.delivery.name, order.get_status_display(),
                order.ordered_at.strftime('%Y-%m-%d %H:%M'),
                item.display_name, item.quantity, item.unit_price, item.amount,
                order.memo,
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'주문목록_{date.today()}.xlsx'
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{filename}"
    return resp


@role_required('admin')
def export_purchase(request):
    openpyxl, wb = _make_workbook()
    if not wb:
        messages.error(request, 'openpyxl 패키지가 필요합니다: uv add openpyxl')
        return redirect('purchase_order')

    today = date.today()
    date_from = request.GET.get('date_from', today.strftime('%Y-%m-01'))
    date_to = request.GET.get('date_to', today.strftime('%Y-%m-%d'))

    items = OrderItem.objects.filter(
        order__status__in=[Order.Status.PENDING, Order.Status.SHIPPING, Order.Status.DELIVERED],
        order__ordered_at__date__gte=date_from,
        order__ordered_at__date__lte=date_to,
    ).select_related('book', 'book__publisher').order_by('book__publisher__name', 'book__series', 'book__name')

    publishers = defaultdict(lambda: defaultdict(lambda: {'name': '', 'series': '', 'qty': 0, 'amount': 0}))
    for item in items:
        pub = item.display_publisher or '기타'
        bid = item.book.pk if item.book else f'custom_{item.pk}'
        publishers[pub][bid]['name'] = item.display_name
        publishers[pub][bid]['series'] = item.display_series
        publishers[pub][bid]['qty'] += item.quantity
        publishers[pub][bid]['amount'] += item.amount

    for pub_name in sorted(publishers.keys()):
        ws = wb.create_sheet(title=pub_name[:31])
        ws.append(['시리즈', '교재명', '수량', '금액'])
        pub_data = publishers[pub_name]
        for book_data in sorted(pub_data.values(), key=lambda b: (b['series'], b['name'])):
            ws.append([book_data['series'], book_data['name'], book_data['qty'], book_data['amount']])

    if not wb.sheetnames:
        wb.create_sheet('발주집계')
    else:
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'발주집계_{date_from}_{date_to}.xlsx'
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
    return resp


# ── 극동 데이터 임포트 ────────────────────────────────────────────────────────

@role_required('admin')
def import_legacy(request):
    agencies = User.objects.filter(role='agency', is_active=True).order_by('name')

    # 기존 임포트 현황
    import_stats = []
    for a in agencies:
        o_cnt = Order.objects.filter(agency=a, source=Order.Source.IMPORT, is_deleted=False).count()
        r_cnt = Return.objects.filter(agency=a, memo__startswith='극동임포트').count()
        if o_cnt or r_cnt:
            import_stats.append({'agency': a, 'orders': o_cnt, 'returns': r_cnt})

    context = {'agencies': agencies, 'import_stats': import_stats}

    if request.method == 'POST':
        action = request.POST.get('action', '')
        agency_id = request.POST.get('agency', '')
        quarter_label = request.POST.get('quarter_label', '')

        if not agency_id:
            messages.error(request, '업체를 선택해주세요.')
            return render(request, 'orders/import_legacy.html', context)

        agency = get_object_or_404(User, pk=agency_id, role='agency')

        # Step 1: 미리보기
        if action == 'preview':
            file = request.FILES.get('file')
            if not file:
                messages.error(request, '엑셀 파일을 선택해주세요.')
                return render(request, 'orders/import_legacy.html', context)

            try:
                schools = parse_geukdong_excel(file)
            except Exception as e:
                messages.error(request, f'엑셀 파싱 오류: {e}')
                return render(request, 'orders/import_legacy.html', context)

            if not schools:
                messages.warning(request, '파싱된 데이터가 없습니다. 극동 엑셀 형식이 맞는지 확인해주세요.')
                return render(request, 'orders/import_legacy.html', context)

            total_items = sum(len(s['items']) for s in schools)
            total_amount = sum(s['subtotal'] for s in schools)

            serialized = json.dumps(schools, default=str)
            request.session['import_preview'] = serialized
            request.session['import_agency_id'] = agency.pk
            request.session['import_quarter_label'] = quarter_label

            context.update({
                'preview': True,
                'schools': schools,
                'total_schools': len(schools),
                'total_items': total_items,
                'total_amount': total_amount,
                'selected_agency': agency,
                'quarter_label': quarter_label,
            })
            return render(request, 'orders/import_legacy.html', context)

        # Step 2: 확정 임포트
        if action == 'confirm':
            serialized = request.session.get('import_preview')
            saved_agency_id = request.session.get('import_agency_id')
            saved_label = request.session.get('import_quarter_label', '')

            if not serialized or str(saved_agency_id) != str(agency_id):
                messages.error(request, '미리보기 데이터가 없습니다. 다시 업로드해주세요.')
                return render(request, 'orders/import_legacy.html', context)

            schools = json.loads(serialized)
            for school in schools:
                for item in school['items']:
                    item['date'] = date.fromisoformat(item['date'])

            stats = import_geukdong_data(agency, schools, saved_label)

            for key in ['import_preview', 'import_agency_id', 'import_quarter_label']:
                request.session.pop(key, None)

            messages.success(
                request,
                f'임포트 완료! 주문 {stats["orders"]}건, '
                f'반품 {stats["returns"]}건, '
                f'항목 {stats["items"] + stats["return_items"]}개'
            )
            return redirect('import_legacy')

    return render(request, 'orders/import_legacy.html', context)


@role_required('admin')
def import_legacy_delete(request):
    """극동 임포트 데이터 일괄 삭제."""
    if request.method != 'POST':
        return redirect('import_legacy')

    agency_id = request.POST.get('agency', '')
    if not agency_id:
        messages.error(request, '업체를 선택해주세요.')
        return redirect('import_legacy')

    agency = get_object_or_404(User, pk=agency_id, role='agency')

    orders = Order.objects.filter(agency=agency, source=Order.Source.IMPORT, is_deleted=False)
    returns = Return.objects.filter(agency=agency, memo__startswith='극동임포트')
    o_count = orders.count()
    r_count = returns.count()

    orders.delete()
    returns.delete()

    messages.success(request, f'{agency.name} 임포트 데이터 삭제: 주문 {o_count}건, 반품 {r_count}건')
    return redirect('import_legacy')
