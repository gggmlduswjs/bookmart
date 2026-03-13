import json
import logging
import math
import re as _re
import threading
import uuid
from collections import OrderedDict

from django.contrib import messages
from django.core.files.base import ContentFile
from django.db.models import Count, Q, Max, Subquery, OuterRef
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.views.decorators.csrf import csrf_exempt

from accounts.decorators import role_required
from accounts.models import User
from books.models import Book
from orders.models import (
    AuditLog, CallRecording, DeliveryAddress, InboxAttachment, InboxMessage,
    Order, OrderItem, OrderStatusLog,
)
from orders.sms import send_sms

from ._helpers import _audit, get_books_json, get_agencies_json, get_teachers_json, get_series_list

logger = logging.getLogger(__name__)


def _extract_phone_digits(sender: str) -> str:
    """sender 문자열에서 전화번호 숫자만 추출 (예: '이쁜둘째딸' → '', '010-5429-6196' → '01054296196')"""
    m = _re.search(r'(\d[\d\-]{8,})', sender or '')
    if m:
        return m.group(1).replace('-', '')
    return ''


def _extract_reply_phone(msg) -> str:
    """수신 SMS에서 답장용 전화번호 추출 (sender → subject → content 순)"""
    for text in [msg.sender, msg.subject, (msg.content or '')[:200]]:
        if not text:
            continue
        m = _re.search(r'(\d[\d\-]{8,})', text)
        if m:
            return m.group(1).replace('-', '')
    return ''


def _build_sms_conversations(sms_qs, hide_done):
    """SMS 메시지를 전화번호 기준으로 대화별로 그룹핑.
    phone 필드가 있으면 우선 사용, 없으면 sender에서 추출."""
    conversations = OrderedDict()  # phone_key -> conv dict

    msgs = list(sms_qs)
    for msg in msgs:
        is_sent = msg.subject == '[발신]'
        # phone 필드 우선, 없으면 sender에서 추출
        phone = getattr(msg, 'phone', '') or ''
        if not phone:
            phone = _extract_phone_digits(msg.sender)
        if not phone:
            phone = msg.sender.strip()

        if phone not in conversations:
            conversations[phone] = {
                'phone': phone,
                'sender_name': '' if is_sent else msg.sender,
                'latest_msg': msg,
                'latest_time': msg.received_at,
                'total_count': 0,
                'unread_count': 0,
                'unprocessed_count': 0,
                'latest_pk': msg.pk,
                'preview': msg.content[:80] if msg.content else '',
                'has_order': msg.order_id is not None,
            }
        else:
            conv = conversations[phone]
            if msg.received_at > conv['latest_time']:
                conv['latest_msg'] = msg
                conv['latest_time'] = msg.received_at
                conv['latest_pk'] = msg.pk
                conv['preview'] = msg.content[:80] if msg.content else ''
            if not conv['sender_name'] and not is_sent:
                conv['sender_name'] = msg.sender

        conv = conversations[phone]
        if not is_sent:
            conv['total_count'] += 1
        if not msg.is_read and not is_sent:
            conv['unread_count'] += 1
        if not msg.is_processed and not is_sent:
            conv['unprocessed_count'] += 1
        if msg.order_id:
            conv['has_order'] = True

    result = list(conversations.values())
    if hide_done:
        result = [c for c in result if c['total_count'] > 0]
    return result


# ── 통합 수신함 ────────────────────────────────────────────────────────────────

@role_required('admin')
def inbox_list(request):
    hide_done = request.GET.get('hide_done', '')
    search = request.GET.get('q', '').strip()
    tab = request.GET.get('tab', 'email')
    from pathlib import Path
    from django.conf import settings as conf
    from django.core.paginator import Paginator

    # 공통 카운트 (탭 배지용) — 가볍게
    unread_email = InboxMessage.objects.filter(is_processed=False, source='email').count()
    unread_sms = InboxMessage.objects.filter(is_processed=False, source='sms').exclude(subject='[발신]').count()
    call_counts = dict(
        CallRecording.objects.values_list('status').annotate(c=Count('id')).values_list('status', 'c')
    )
    pending_calls = call_counts.get('pending', 0) + call_counts.get('parsed', 0)
    call_counts['unprocessed'] = pending_calls
    call_counts['total'] = sum(call_counts.get(s, 0) for s in ['pending', 'parsed', 'ordered', 'failed', 'skipped', 'processing'])
    token_path = Path(conf.BASE_DIR) / 'gdrive_token.json'

    # 활성 탭 데이터만 로드 (다른 탭은 빈 값)
    email_page = None
    sms_conversations = []
    sms_page = None
    call_page = None
    all_items = None

    base_qs = InboxMessage.objects.select_related('order')
    if hide_done:
        base_qs = base_qs.filter(is_processed=False)
    if search:
        base_qs = base_qs.filter(
            Q(sender__icontains=search) |
            Q(subject__icontains=search) |
            Q(content__icontains=search)
        )

    if tab == 'email':
        email_qs = base_qs.filter(source='email').exclude(
            subject__startswith='[발신]'
        ).annotate(
            attachment_count=Count('attachments')
        ).prefetch_related('attachments').order_by('-received_at')
        paginator = Paginator(email_qs, 50)
        email_page = paginator.get_page(request.GET.get('page'))
    elif tab == 'sms':
        # DB-level grouping by phone
        sms_group_qs = (InboxMessage.objects.filter(source='sms')
            .exclude(subject='[발신]')
            .exclude(phone=''))
        if hide_done:
            sms_group_qs = sms_group_qs.filter(is_processed=False)
        if search:
            sms_group_qs = sms_group_qs.filter(
                Q(sender__icontains=search) |
                Q(content__icontains=search) |
                Q(phone__icontains=search)
            )
        sms_groups = (sms_group_qs
            .values('phone')
            .annotate(
                latest_time=Max('received_at'),
                total_count=Count('id'),
                unread_count=Count('id', filter=Q(is_read=False)),
                has_order=Count('order', distinct=True),
            )
            .order_by('-latest_time'))

        sms_paginator = Paginator(sms_groups, 50)
        sms_page = sms_paginator.get_page(request.GET.get('sms_page'))

        # Fetch latest message per phone for preview
        phone_list = [g['phone'] for g in sms_page]
        latest_msgs = {}
        if phone_list:
            latest_pks = (InboxMessage.objects.filter(source='sms', phone__in=phone_list)
                .exclude(subject='[발신]')
                .values('phone')
                .annotate(latest_pk=Max('id'))
                .values_list('latest_pk', flat=True))
            for msg in InboxMessage.objects.filter(pk__in=latest_pks):
                latest_msgs[msg.phone] = msg

        sms_conversations = []
        for g in sms_page:
            msg = latest_msgs.get(g['phone'])
            sms_conversations.append({
                'phone': g['phone'],
                'sender_name': msg.sender if msg else '',
                'latest_time': g['latest_time'],
                'total_count': g['total_count'],
                'unread_count': g['unread_count'],
                'latest_pk': msg.pk if msg else 0,
                'preview': msg.content[:80] if msg and msg.content else '',
                'has_order': g['has_order'] > 0,
            })
    elif tab == 'call':
        call_qs = CallRecording.objects.all()
        call_status_filter = request.GET.get('call_status', '')
        if call_status_filter == 'unprocessed' or not call_status_filter:
            call_qs = call_qs.filter(status__in=['pending', 'parsed'])
        elif call_status_filter == 'failed':
            call_qs = call_qs.filter(status='failed')
        elif call_status_filter == 'all':
            pass  # no filter
        elif call_status_filter:
            call_qs = call_qs.filter(status=call_status_filter)
        call_paginator = Paginator(call_qs.order_by('-created_at'), 30)
        call_page = call_paginator.get_page(request.GET.get('call_page'))
    elif tab == 'all':
        # 통합 타임라인: 미처리 이메일 + SMS + 대기/파싱완료 통화
        unified = []
        emails = base_qs.filter(source='email', is_processed=False).annotate(
            attachment_count=Count('attachments')
        ).order_by('-received_at')[:100]
        for e in emails:
            unified.append({
                'type': 'email', 'pk': e.pk, 'sender': e.sender,
                'preview': e.subject or (e.content[:80] if e.content else ''),
                'timestamp': e.received_at, 'status': 'has_order' if e.order_id else 'unprocessed',
                'attachment_count': e.attachment_count,
            })
        # SMS: 최근 미처리 수신 메시지 (발신 제외)
        sms_msgs = base_qs.filter(source='sms', is_processed=False).exclude(
            subject='[발신]'
        ).order_by('-received_at')[:100]
        for s in sms_msgs:
            unified.append({
                'type': 'sms', 'pk': s.pk, 'sender': s.sender or s.phone,
                'preview': (s.content[:80] if s.content else ''),
                'timestamp': s.received_at, 'status': 'has_order' if s.order_id else 'unprocessed',
                'phone': s.phone,
            })
        # 통화: 대기/파싱완료
        calls = CallRecording.objects.filter(
            status__in=['pending', 'parsed']
        ).order_by('-created_at')[:50]
        for c in calls:
            unified.append({
                'type': 'call', 'pk': c.pk,
                'sender': c.caller_phone or c.file_name,
                'preview': c.summary or '통화 녹음',
                'timestamp': c.created_at,
                'status': c.status,
                'duration': c.duration_sec,
            })
        unified.sort(key=lambda x: x['timestamp'], reverse=True)
        all_paginator = Paginator(unified, 50)
        all_items = all_paginator.get_page(request.GET.get('page'))

    call_status_val = request.GET.get('call_status', '')

    # Cross-tab search counts (Phase 3.2)
    search_counts = None
    if search:
        search_filter = Q(sender__icontains=search) | Q(subject__icontains=search) | Q(content__icontains=search)
        search_counts = {
            'email': InboxMessage.objects.filter(source='email').filter(search_filter).count(),
            'sms': InboxMessage.objects.filter(source='sms').exclude(subject='[발신]').filter(search_filter).count(),
            'call': CallRecording.objects.filter(
                Q(caller_phone__icontains=search) | Q(summary__icontains=search) | Q(transcript__icontains=search) | Q(file_name__icontains=search)
            ).count(),
        }

    total_unprocessed = unread_email + unread_sms + pending_calls

    return render(request, 'orders/inbox_list.html', {
        'email_page': email_page,
        'email_messages': email_page,  # 템플릿 호환
        'sms_conversations': sms_conversations,
        'sms_page': sms_page,
        'tab': tab,
        'hide_done': hide_done,
        'search': search,
        'search_counts': search_counts,
        'unread_email': unread_email,
        'unread_sms': unread_sms,
        'unread_count': unread_email + unread_sms,
        'call_page': call_page,
        'call_status': call_status_val,
        'call_counts': call_counts,
        'pending_calls': pending_calls,
        'total_unprocessed': total_unprocessed,
        'all_items': all_items,
        'gdrive_connected': token_path.exists(),
    })


@role_required('admin')
def inbox_single_skip(request, pk):
    """단건 처리완료 (is_processed=True)"""
    if request.method != 'POST':
        return redirect('inbox_list')
    msg = get_object_or_404(InboxMessage, pk=pk)
    tab = 'sms' if msg.source == 'sms' else 'email'
    if not msg.is_processed:
        msg.is_processed = True
        msg.save(update_fields=['is_processed'])
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'pk': pk})
        messages.success(request, '처리완료했습니다.')
    return redirect(f'/inbox/?tab={tab}')


def _get_imap_account(label):
    """imap_key의 account_label로 계정 정보 반환"""
    from django.conf import settings as conf
    account_map = {}
    if hasattr(conf, 'NAVER_EMAIL_1_ID'):
        account_map['007bm'] = (conf.NAVER_EMAIL_1_ID, conf.NAVER_EMAIL_1_PW)
    if hasattr(conf, 'NAVER_EMAIL_2_ID'):
        account_map['002bm'] = (conf.NAVER_EMAIL_2_ID, conf.NAVER_EMAIL_2_PW)
    return account_map.get(label)


def _delete_imap_by_key(imap_key):
    from orders.email_utils import delete_email_imap
    parts = imap_key.split(':', 1)
    if len(parts) != 2:
        return
    label, uid_str = parts
    creds = _get_imap_account(label)
    if creds:
        delete_email_imap(creds[0], creds[1], uid_str)


def _delete_imap_bulk(imap_keys):
    from orders.email_utils import delete_emails_imap
    by_account = {}
    for key in imap_keys:
        parts = key.split(':', 1)
        if len(parts) != 2:
            continue
        label, uid_str = parts
        by_account.setdefault(label, []).append(uid_str)
    for label, uids in by_account.items():
        creds = _get_imap_account(label)
        if creds:
            delete_emails_imap(creds[0], creds[1], uids)


@role_required('admin')
def inbox_delete(request, pk):
    """수신 메시지 단건 삭제"""
    if request.method != 'POST':
        return redirect('inbox_list')
    msg = get_object_or_404(InboxMessage, pk=pk)
    tab = 'sms' if msg.source == 'sms' else 'email'
    if msg.source == 'email' and msg.imap_key:
        _delete_imap_by_key(msg.imap_key)
    msg.attachments.all().delete()
    msg.delete()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': True, 'pk': pk})
    messages.success(request, '삭제했습니다.')
    return redirect(f'/inbox/?tab={tab}')


@role_required('admin')
def inbox_bulk_delete(request):
    """선택한 수신 메시지 일괄 삭제"""
    if request.method != 'POST':
        return redirect('inbox_list')
    tab = request.POST.get('tab', 'email')
    msg_ids = request.POST.getlist('msg_ids')
    if msg_ids:
        qs = InboxMessage.objects.filter(pk__in=msg_ids)
        count = qs.count()
        email_msgs = qs.filter(source='email').exclude(imap_key='').values_list('imap_key', flat=True)
        if email_msgs:
            _delete_imap_bulk(list(email_msgs))
        qs.delete()
        messages.success(request, f'{count}건을 삭제했습니다.')
    else:
        messages.warning(request, '선택된 메시지가 없습니다.')
    return redirect(f'/inbox/?tab={tab}')


@role_required('admin')
def inbox_bulk_skip(request):
    """선택한 수신 메시지를 일괄 건너뛰기 (is_processed=True)"""
    if request.method != 'POST':
        return redirect('inbox_list')

    action = request.POST.get('action', '')
    tab = request.POST.get('tab', 'email')
    if action == 'skip_all':
        updated = InboxMessage.objects.filter(is_processed=False, source='email').update(is_processed=True)
        messages.success(request, f'미처리 메일 {updated}건을 모두 건너뛰었습니다.')
    elif action == 'skip_all_sms':
        updated = InboxMessage.objects.filter(is_processed=False, source='sms').update(is_processed=True)
        messages.success(request, f'미처리 문자 {updated}건을 모두 건너뛰었습니다.')
    else:
        msg_ids = request.POST.getlist('msg_ids')
        if msg_ids:
            updated = InboxMessage.objects.filter(
                pk__in=msg_ids, is_processed=False
            ).update(is_processed=True)
            messages.success(request, f'{updated}건을 건너뛰었습니다.')
        else:
            messages.warning(request, '선택된 메시지가 없습니다.')

    return redirect(f'/inbox/?tab={tab}')


# 비동기 이메일 가져오기 상태 관리
_fetch_tasks = {}  # task_id -> {status, count, sync, error}
_fetch_lock = threading.Lock()


def _do_fetch_emails(task_id):
    """백그라운드 스레드에서 이메일 가져오기 실행"""
    import django
    django.setup()

    from django.conf import settings as conf
    from orders.email_utils import fetch_naver_emails

    try:
        accounts = [
            (conf.NAVER_EMAIL_1_ID, conf.NAVER_EMAIL_1_PW, '007bm'),
            (conf.NAVER_EMAIL_2_ID, conf.NAVER_EMAIL_2_PW, '002bm'),
        ]

        all_existing_keys = set(
            InboxMessage.objects.values_list('imap_key', flat=True)
        )

        new_count = 0
        sync_count = 0
        for acc_id, acc_pw, label in accounts:
            if not acc_id or not acc_pw:
                continue
            emails, read_sync = fetch_naver_emails(acc_id, acc_pw, label,
                                                   existing_keys=all_existing_keys)

            if read_sync:
                for imap_key, is_seen in read_sync.items():
                    updated = InboxMessage.objects.filter(
                        imap_key=imap_key, is_read=not is_seen
                    ).update(is_read=is_seen)
                    sync_count += updated

            if not emails:
                continue

            for e in emails:
                is_sent = e.get('is_sent', False)
                msg_obj = InboxMessage.objects.create(
                    source=InboxMessage.Source.EMAIL,
                    account_label=e['account_label'],
                    sender=e['sender'],
                    subject=e['subject'],
                    content=e['content'],
                    received_at=e['received_at'],
                    imap_key=e['imap_key'],
                    is_processed=is_sent,  # 발신 메일은 처리완료 상태
                    is_read=True if is_sent else e.get('is_seen', False),
                    message_id=e.get('message_id', ''),
                    phone=e.get('to_email', ''),  # 발신 메일: 수신자 이메일 (스레드 그룹핑용)
                )
                if not is_sent:
                    for att in e.get('attachments', []):
                        file_obj = ContentFile(att['data'], name=att['filename'])
                        InboxAttachment.objects.create(
                            message=msg_obj,
                            file=file_obj,
                            filename=att['filename'],
                            content_type=att['content_type'],
                            size=len(att['data']),
                        )
                new_count += 1

        with _fetch_lock:
            _fetch_tasks[task_id] = {'status': 'done', 'count': new_count, 'sync': sync_count, 'error': ''}

    except Exception as exc:
        logger.exception('fetch_emails background error')
        with _fetch_lock:
            _fetch_tasks[task_id] = {'status': 'error', 'count': 0, 'sync': 0, 'error': str(exc)}


@role_required('admin')
def fetch_emails(request):
    """네이버 IMAP 메일 가져오기 (비동기)"""
    if request.method != 'POST':
        return redirect('inbox_list')

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    # 이미 실행 중인 작업이 있으면 중복 실행 방지
    with _fetch_lock:
        for tid, task in list(_fetch_tasks.items()):
            if task['status'] == 'running':
                if is_ajax:
                    return JsonResponse({'ok': True, 'task_id': tid, 'status': 'running'})
                messages.info(request, '이미 메일을 가져오고 있습니다.')
                return redirect('inbox_list')

    task_id = uuid.uuid4().hex[:12]
    with _fetch_lock:
        _fetch_tasks[task_id] = {'status': 'running', 'count': 0, 'sync': 0, 'error': ''}

    t = threading.Thread(target=_do_fetch_emails, args=(task_id,), daemon=True)
    t.start()

    if is_ajax:
        return JsonResponse({'ok': True, 'task_id': task_id, 'status': 'running'})

    # 비-AJAX 폴백: 스레드 완료 대기 (최대 90초)
    t.join(timeout=90)
    with _fetch_lock:
        result = _fetch_tasks.pop(task_id, {})
    new_count = result.get('count', 0)
    sync_count = result.get('sync', 0)
    sync_msg = f' (읽음 상태 {sync_count}건 동기화)' if sync_count else ''
    messages.success(request, f'새 메일 {new_count}건을 가져왔습니다.{sync_msg}')
    return redirect('inbox_list')


@role_required('admin')
def fetch_emails_status(request):
    """비동기 이메일 가져오기 상태 확인 (polling)"""
    task_id = request.GET.get('task_id', '')
    with _fetch_lock:
        task = _fetch_tasks.get(task_id)
    if not task:
        return JsonResponse({'status': 'not_found'})
    if task['status'] == 'done':
        with _fetch_lock:
            _fetch_tasks.pop(task_id, None)
    return JsonResponse({
        'status': task['status'],
        'count': task.get('count', 0),
        'sync': task.get('sync', 0),
        'error': task.get('error', ''),
    })


@role_required('admin')
def inbox_process(request, pk):
    """수신 메시지를 보면서 주문 등록"""
    inbox_msg = get_object_or_404(InboxMessage, pk=pk)

    # 열 때 읽음 처리
    if not inbox_msg.is_read:
        inbox_msg.is_read = True
        inbox_msg.save(update_fields=['is_read'])
        if inbox_msg.source == 'email' and inbox_msg.imap_key:
            from django.conf import settings as conf
            from orders.email_utils import mark_as_read_imap
            parts = inbox_msg.imap_key.split(':', 1)
            if len(parts) == 2:
                label, uid_str = parts
                account_map = {}
                if hasattr(conf, 'NAVER_EMAIL_1_ID'):
                    account_map['007bm'] = (conf.NAVER_EMAIL_1_ID, conf.NAVER_EMAIL_1_PW)
                if hasattr(conf, 'NAVER_EMAIL_2_ID'):
                    account_map['002bm'] = (conf.NAVER_EMAIL_2_ID, conf.NAVER_EMAIL_2_PW)
                creds = account_map.get(label)
                if creds:
                    mark_as_read_imap(creds[0], creds[1], uid_str)

    next_unprocessed = (
        InboxMessage.objects.filter(is_processed=False, source=inbox_msg.source)
        .exclude(pk=pk).order_by('-received_at').first()
    )

    # 처리 완료 (주문 없이)
    if request.method == 'POST' and 'skip' in request.POST:
        inbox_msg.is_processed = True
        inbox_msg.save(update_fields=['is_processed'])
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'skipped': True, 'inbox_pk': inbox_msg.pk})
        messages.info(request, '처리 완료되었습니다.')
        nxt = InboxMessage.objects.filter(is_processed=False, source=inbox_msg.source).order_by('-received_at').first()
        if nxt:
            return redirect('inbox_process', pk=nxt.pk)
        return redirect('inbox_list')

    agencies = User.objects.filter(role='agency', is_active=True).order_by('name')
    agencies_json = json.dumps([{
        'id': a.pk, 'name': a.name,
    } for a in agencies], ensure_ascii=False)

    teachers = (
        User.objects.filter(role='teacher', is_active=True)
        .select_related('agency', 'delivery_address')
        .order_by('agency__name', 'name')
    )
    teachers_json = json.dumps([{
        'id': t.pk,
        'name': t.name,
        'phone': t.phone or '',
        'agency_id': t.agency_id,
        'delivery_id': t.delivery_address_id,
        'delivery_name': t.delivery_address.name if t.delivery_address else '',
        'delivery_address': t.delivery_address.address if t.delivery_address else '',
        'delivery_phone': t.delivery_address.phone if t.delivery_address else '',
        'has_delivery': bool(t.delivery_address),
    } for t in teachers], ensure_ascii=False)

    books = Book.objects.filter(is_active=True).select_related('publisher')
    series_list = sorted(set(b.series for b in books if b.series))
    books_json = json.dumps([{
        'id': b.id,
        'series': b.series or '기타',
        'name': b.name,
        'publisher': b.publisher.name,
        'unit_price': math.floor(b.list_price * float(b.publisher.supply_rate) / 100),
    } for b in books], ensure_ascii=False)

    if request.method == 'POST' and 'skip' not in request.POST:
        agency_id = request.POST.get('agency_id', '').strip()
        teacher_id = request.POST.get('teacher_id', '').strip()
        new_teacher_name = request.POST.get('new_teacher_name', '').strip()
        new_teacher_phone = request.POST.get('new_teacher_phone', '').strip()
        delivery_school = request.POST.get('delivery_school', '').strip()
        delivery_address_val = request.POST.get('delivery_address', '').strip()
        delivery_phone = request.POST.get('delivery_phone', '').strip()

        try:
            agency = User.objects.get(pk=agency_id, role='agency', is_active=True)
        except (User.DoesNotExist, ValueError):
            messages.error(request, '업체를 선택해 주세요.')
            return redirect('inbox_process', pk=pk)

        if teacher_id:
            try:
                teacher = User.objects.select_related('delivery_address').get(
                    pk=teacher_id, role='teacher', is_active=True
                )
            except (User.DoesNotExist, ValueError):
                messages.error(request, '선생님을 선택해 주세요.')
                return redirect('inbox_process', pk=pk)
        elif new_teacher_name:
            login_id = f'a_{new_teacher_phone or "nophone"}_{agency.pk}'
            if User.objects.filter(login_id=login_id).exists():
                teacher = User.objects.get(login_id=login_id)
            else:
                teacher = User(
                    login_id=login_id, role='teacher',
                    name=new_teacher_name, phone=new_teacher_phone,
                    agency=agency, must_change_password=False,
                )
                teacher.set_unusable_password()
                teacher.save()
        else:
            messages.error(request, '선생님을 선택하거나 새로 입력해 주세요.')
            return redirect('inbox_process', pk=pk)

        if delivery_school:
            delivery, created = DeliveryAddress.objects.get_or_create(
                agency=agency, name=delivery_school,
                defaults={'address': delivery_address_val, 'phone': delivery_phone},
            )
            if not created and delivery_address_val:
                delivery.address = delivery_address_val
                delivery.phone = delivery_phone
                delivery.save(update_fields=['address', 'phone'])
            teacher.delivery_address = delivery
            teacher.save(update_fields=['delivery_address'])
        elif not teacher.delivery_address:
            messages.error(request, '배송지를 입력해 주세요.')
            return redirect('inbox_process', pk=pk)

        items = []
        i = 0
        while f'book_{i}' in request.POST or f'custom_name_{i}' in request.POST:
            book_id = request.POST.get(f'book_{i}', '').strip()
            custom_name = request.POST.get(f'custom_name_{i}', '').strip()
            custom_price = request.POST.get(f'custom_price_{i}', '').strip()
            qty_str = request.POST.get(f'qty_{i}', '').strip()
            if book_id and qty_str:
                try:
                    qty = int(qty_str)
                    if qty > 0:
                        items.append({'book_id': int(book_id), 'qty': qty})
                except (ValueError, TypeError):
                    pass
            elif custom_name and qty_str:
                try:
                    qty = int(qty_str)
                    price = int(custom_price) if custom_price else 0
                    if qty > 0:
                        items.append({'custom_name': custom_name, 'custom_price': price, 'qty': qty})
                except (ValueError, TypeError):
                    pass
            i += 1

        if not items:
            messages.error(request, '주문할 교재를 1권 이상 선택하세요.')
        else:
            order = Order.objects.create(
                order_no=Order.generate_order_no(),
                agency=agency,
                teacher=teacher,
                delivery=teacher.delivery_address,
                memo=request.POST.get('memo', ''),
                source=Order.Source.INBOX,
            )
            for item in items:
                if 'book_id' in item:
                    try:
                        book = Book.objects.get(id=item['book_id'], is_active=True)
                        OrderItem(order=order, book=book, quantity=item['qty']).save()
                    except Book.DoesNotExist:
                        pass
                else:
                    OrderItem(order=order, custom_book_name=item['custom_name'],
                              unit_price=item['custom_price'], quantity=item['qty']).save()

            OrderStatusLog.objects.create(
                order=order, old_status='', new_status='pending',
                changed_by=request.user, memo='수신함에서 주문 생성',
            )
            _audit(request, AuditLog.Action.ORDER_CREATE, order, f'[수신함] 주문 {order.order_no} 생성')

            inbox_msg.is_processed = True
            inbox_msg.order = order
            inbox_msg.save(update_fields=['is_processed', 'order'])

            # AJAX → JSON 응답
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'ok': True,
                    'order_pk': order.pk,
                    'order_no': order.order_no,
                    'inbox_pk': inbox_msg.pk,
                })
            messages.success(request, f'주문 등록 완료. 주문번호: {order.order_no}')
            nxt = InboxMessage.objects.filter(is_processed=False, source=inbox_msg.source).order_by('-received_at').first()
            if nxt:
                return redirect('inbox_process', pk=nxt.pk)
            return redirect('inbox_list')

    attachments = inbox_msg.attachments.all()

    # AI 이메일 파싱
    parsed_json = 'null'
    parse_error = ''
    if inbox_msg.content and not inbox_msg.is_processed:
        try:
            from orders.call_order import parse_order_from_email
            books_data = [{
                'id': b.id, 'series': b.series or '기타', 'name': b.name,
                'publisher': b.publisher.name,
                'unit_price': math.floor(b.list_price * float(b.publisher.supply_rate) / 100),
            } for b in books]
            agencies_data = [{'id': a.pk, 'name': a.name} for a in agencies]
            teachers_data = [{
                'id': t.pk, 'name': t.name, 'agency_id': t.agency_id,
                'agency_name': t.agency.name if t.agency else '',
                'delivery_name': t.delivery_address.name if t.delivery_address else '',
            } for t in teachers]
            parsed, err = parse_order_from_email(
                sender=inbox_msg.sender or '',
                subject=inbox_msg.subject or '',
                body=inbox_msg.content,
                book_list=books_data,
                agencies=agencies_data,
                teachers=teachers_data,
            )
            if parsed and not err:
                parsed_json = json.dumps(parsed, ensure_ascii=False)
            elif err:
                parse_error = err
                logger.warning('이메일 AI 파싱 실패: %s', err)
        except Exception as e:
            parse_error = f'AI 파싱 중 오류 발생: {e}'
            logger.exception('이메일 AI 파싱 오류')

    # SMS 답장용 전화번호 추출 (sender, subject, content 순서로 탐색)
    sms_reply_phone = ''
    if inbox_msg.source == 'sms':
        for text in [inbox_msg.sender, inbox_msg.subject, inbox_msg.content[:200]]:
            if not text:
                continue
            phone_match = _re.search(r'(\d[\d\-]{8,})', text)
            if phone_match:
                sms_reply_phone = phone_match.group(1).replace('-', '')
                break
        # 번호 못 찾으면 기본값 (테스트용)
        if not sms_reply_phone:
            sms_reply_phone = '01054296196'
        if len(sms_reply_phone) == 11:
            sms_reply_phone = f'{sms_reply_phone[:3]}-{sms_reply_phone[3:7]}-{sms_reply_phone[7:]}'
        elif len(sms_reply_phone) == 10:
            sms_reply_phone = f'{sms_reply_phone[:3]}-{sms_reply_phone[3:6]}-{sms_reply_phone[6:]}'

    # SMS 대화 내역 (같은 발신번호의 메시지들 - 수신+발신 포함)
    sms_conversation = []
    if inbox_msg.source == 'sms':
        q_filter = Q(sender__contains=inbox_msg.sender)
        # sender에서 전화번호 추출
        phone_in_sender = _re.search(r'(\d[\d\-]{8,})', inbox_msg.sender or '')
        if phone_in_sender:
            phone_digits = phone_in_sender.group(1).replace('-', '')
            q_filter = q_filter | Q(sender__contains=phone_digits)
        # sms_reply_phone으로도 매칭 (발신 문자 찾기)
        if sms_reply_phone:
            q_filter = q_filter | Q(sender__contains=sms_reply_phone)
            reply_digits = sms_reply_phone.replace('-', '')
            q_filter = q_filter | Q(sender__contains=reply_digits)
        sms_conversation = list(
            InboxMessage.objects.filter(source='sms')
            .filter(q_filter)
            .order_by('received_at')[:100]
        )

    return render(request, 'orders/inbox_process.html', {
        'inbox_msg': inbox_msg,
        'agencies': agencies,
        'agencies_json': agencies_json,
        'teachers_json': teachers_json,
        'series_list': series_list,
        'books_json': books_json,
        'attachments': attachments,
        'next_unprocessed': next_unprocessed,
        'parsed_json': parsed_json,
        'parse_error': parse_error,
        'sms_reply_phone': sms_reply_phone,
        'sms_conversation': sms_conversation,
    })


@role_required('admin')
def inbox_reply(request, pk):
    """수신 이메일에 답장 발송 (첨부파일 지원)"""
    inbox_msg = get_object_or_404(InboxMessage, pk=pk)
    if request.method != 'POST' or inbox_msg.source != 'email':
        return redirect('inbox_process', pk=pk)

    reply_body = request.POST.get('reply_body', '').strip()
    if not reply_body:
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            return JsonResponse({'ok': False, 'error': '답장 내용을 입력하세요.'})
        messages.warning(request, '답장 내용을 입력하세요.')
        return redirect('inbox_process', pk=pk)

    from django.conf import settings as conf

    # 커스텀 수신자/제목 지원
    custom_to = request.POST.get('to_email', '').strip()
    custom_subject = request.POST.get('subject', '').strip()

    sender = inbox_msg.sender or ''
    match = _re.search(r'[\w.\-+]+@[\w.\-]+\.\w+', sender)
    to_email = custom_to or (match.group(0) if match else sender)

    subj = inbox_msg.subject or ''
    if custom_subject:
        reply_subject = custom_subject
    else:
        reply_subject = subj if subj.lower().startswith('re:') else f'Re: {subj}'

    in_reply_to = inbox_msg.message_id or None
    references = inbox_msg.message_id or None

    account_map = {}
    if hasattr(conf, 'NAVER_EMAIL_1_ID'):
        account_map['007bm'] = (conf.NAVER_EMAIL_1_ID, conf.NAVER_EMAIL_1_PW)
    if hasattr(conf, 'NAVER_EMAIL_2_ID'):
        account_map['002bm'] = (conf.NAVER_EMAIL_2_ID, conf.NAVER_EMAIL_2_PW)

    creds = account_map.get(inbox_msg.account_label)
    if not creds:
        creds = (conf.NAVER_EMAIL_1_ID, conf.NAVER_EMAIL_1_PW)

    # 사업자 서류 첨부 처리
    doc_ids = request.POST.getlist('doc_ids')
    attachments = []
    if doc_ids:
        from orders.models import BusinessDocument
        for doc in BusinessDocument.objects.filter(pk__in=doc_ids):
            try:
                doc.file.open('rb')
                data = doc.file.read()
                doc.file.close()
                ext = doc.extension
                filename = f'{doc.name}.{ext}' if ext else doc.name
                attachments.append({
                    'filename': filename,
                    'data': data,
                    'content_type': 'application/octet-stream',
                })
            except Exception as e:
                logger.error('서류 파일 읽기 실패 (%s): %s', doc.name, e)

    if attachments:
        from orders.email_utils import send_email_with_attachments
        ok = send_email_with_attachments(
            account_id=creds[0],
            account_pw=creds[1],
            to_email=to_email,
            subject=reply_subject,
            body=reply_body,
            attachments=attachments,
        )
    else:
        from orders.email_utils import send_reply_email
        ok = send_reply_email(
            account_id=creds[0],
            account_pw=creds[1],
            to_email=to_email,
            subject=reply_subject,
            body=reply_body,
            in_reply_to=in_reply_to,
            references=references,
        )

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if ok:
        # 발신 이메일을 InboxMessage에 저장 (스레드 표시용)
        from django.utils import timezone
        InboxMessage.objects.create(
            source='email',
            account_label=inbox_msg.account_label,
            sender=creds[0],  # 발신 계정
            subject='[발신] ' + reply_subject,
            content=reply_body,
            received_at=timezone.now(),
            is_processed=True,
            is_read=True,
            phone=to_email,  # phone 필드를 이메일 그룹핑에 재사용
        )
        att_msg = f' (첨부 {len(attachments)}건)' if attachments else ''
        if is_ajax:
            return JsonResponse({'ok': True, 'message': f'{to_email}에 메일을 발송했습니다.{att_msg}'})
        messages.success(request, f'{to_email}에 메일을 발송했습니다.{att_msg}')
    else:
        if is_ajax:
            return JsonResponse({'ok': False, 'error': '메일 발송에 실패했습니다.'}, status=500)
        messages.error(request, '메일 발송에 실패했습니다. 로그를 확인하세요.')

    return redirect('inbox_process', pk=pk)


@role_required('admin')
def attachment_download(request, pk):
    """첨부파일 다운로드"""
    from urllib.parse import quote
    att = get_object_or_404(InboxAttachment, pk=pk)
    att.file.open('rb')
    data = att.file.read()
    att.file.close()
    resp = HttpResponse(data, content_type=att.content_type or 'application/octet-stream')
    encoded_filename = quote(att.filename)
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
    return resp


@role_required('admin')
@xframe_options_sameorigin
def attachment_preview(request, pk):
    """첨부파일 미리보기 (엑셀/이미지/PDF/HWP)"""
    att = get_object_or_404(InboxAttachment, pk=pk)

    # 이미지 미리보기
    if att.is_image:
        att.file.open('rb')
        data = att.file.read()
        att.file.close()
        content_type = att.content_type or f'image/{att.extension}'
        return HttpResponse(data, content_type=content_type)

    # PDF 미리보기
    if att.is_pdf:
        att.file.open('rb')
        data = att.file.read()
        att.file.close()
        resp = HttpResponse(data, content_type='application/pdf')
        resp['Content-Disposition'] = 'inline'
        return resp

    # HWP 텍스트 추출 미리보기
    if att.is_hwp:
        try:
            import olefile
            att.file.open('rb')
            raw = att.file.read()
            att.file.close()
            import io
            ole = olefile.OleFileIO(io.BytesIO(raw))
            if ole.exists('PrvText'):
                text = ole.openstream('PrvText').read().decode('utf-16-le', errors='replace')
            elif ole.exists('BodyText/Section0'):
                raw_body = ole.openstream('BodyText/Section0').read()
                import zlib
                try:
                    decompressed = zlib.decompress(raw_body, -15)
                except Exception:
                    decompressed = raw_body
                text = decompressed.decode('utf-16-le', errors='replace')
                text = _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)
            else:
                text = '(HWP 텍스트를 추출할 수 없습니다)'
            ole.close()
            from django.utils.html import escape
            html = f'''<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{escape(att.filename)}</title>
<style>
body {{ font-family:'Malgun Gothic',sans-serif; font-size:13px; padding:16px; background:#f5f5f5; line-height:1.8; }}
.hwp-content {{ background:#fff; padding:20px; border:1px solid #ddd; border-radius:6px; white-space:pre-wrap; word-break:break-all; }}
h2 {{ font-size:14px; margin-bottom:8px; color:#334155; }}
</style></head><body>
<h2>{escape(att.filename)}</h2>
<div class="hwp-content">{escape(text)}</div>
</body></html>'''
            return HttpResponse(html)
        except ImportError:
            return HttpResponse('HWP 미리보기를 사용하려면 olefile 패키지가 필요합니다.', status=400)
        except Exception as e:
            return HttpResponse(f'HWP 파일 처리 오류: {e}', status=400)

    # 엑셀 미리보기
    if att.is_excel:
        import openpyxl
        try:
            wb = openpyxl.load_workbook(att.file, read_only=True, data_only=True)
        except Exception:
            return HttpResponse('엑셀 파일을 열 수 없습니다.', status=400)

        sheets_html = []
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            from django.utils.html import escape
            table = f'<h3 style="margin:10px 0 4px;font-size:13px">{escape(ws.title)}</h3>'
            table += '<table class="data-table"><thead><tr>'
            for cell in rows[0]:
                table += f'<th>{escape(str(cell)) if cell is not None else ""}</th>'
            table += '</tr></thead><tbody>'
            for row in rows[1:200]:
                table += '<tr>'
                for cell in row:
                    val = f'{cell:,}' if isinstance(cell, (int, float)) and not isinstance(cell, bool) else (escape(str(cell)) if cell is not None else '')
                    table += f'<td>{val}</td>'
                table += '</tr>'
            table += '</tbody></table>'
            sheets_html.append(table)
        wb.close()

        html = f'''<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{att.filename}</title>
<style>
body {{ font-family:'Malgun Gothic',sans-serif; font-size:12px; padding:12px; background:#f5f5f5; }}
table {{ width:100%; border-collapse:collapse; margin-bottom:16px; background:#fff; }}
th {{ background:#666; color:#fff; padding:5px 8px; border:1px solid #555; text-align:center; white-space:nowrap; }}
td {{ padding:4px 8px; border:1px solid #ddd; }}
tr:nth-child(even) td {{ background:#fafafa; }}
h3 {{ color:#e8720c; }}
</style></head><body>
<h2 style="font-size:14px;margin-bottom:8px">{att.filename}</h2>
{"".join(sheets_html)}
</body></html>'''
        return HttpResponse(html)

    return HttpResponse('미리보기를 지원하지 않는 파일 형식입니다.', status=400)


@csrf_exempt
def sms_webhook(request):
    """
    갤럭시 SMS Forwarder 앱에서 POST로 오는 문자 수신.
    앱 설정: URL = https://도메인/webhook/sms/
    Body format (JSON): {"from":"010-xxxx", "text":"...", "sentStamp":1234567890}
    """
    if request.method != 'POST':
        return HttpResponse('OK')

    from django.utils import timezone
    import datetime

    try:
        body = request.body
        try:
            body_str = body.decode('utf-8')
        except UnicodeDecodeError:
            body_str = body.decode('euc-kr', errors='replace')
        data = json.loads(body_str)
        logger.info('SMS 웹훅 수신 데이터: %s', body_str[:500])
        from_name = data.get('from', '')
        from_number = data.get('from_number') or data.get('number', '')
        # formatted-msg에서 전화번호 추출 시도
        formatted = data.get('detail') or data.get('formatted-msg', '')
        if formatted:
            logger.info('SMS formatted-msg: %s', formatted[:300])
            fm = _re.search(r'(01\d[\d\-]{7,})', formatted)
            if fm and not from_number:
                from_number = fm.group(1).replace('-', '')
        # {number} 리터럴 제거 (Forward SMS 앱에서 변수 치환 실패 시)
        if from_number == '{number}':
            from_number = ''
        from_name = _re.sub(r'\(\{number\}\)', '', from_name).strip()
        # from_name에서 실제 전화번호 추출: "김은경선생님 (01032278210)" → 번호 분리
        name_phone_match = _re.search(r'\(?(01\d[\d\-]{7,})\)?', from_name)
        if name_phone_match and not from_number:
            from_number = name_phone_match.group(1).replace('-', '')
            # 이름에서 번호 부분 제거: "김은경선생님 (010...)" → "김은경선생님"
            from_name = from_name[:name_phone_match.start()].strip().rstrip('(').strip()
        # 이름과 번호 모두 있으면 "이름(번호)" 형식
        if from_name and from_number and from_number not in from_name:
            sender = f'{from_name}({from_number})'
        else:
            sender = from_name or from_number
        content = data.get('text') or data.get('message', '')
        ts = data.get('sentStamp') or data.get('timestamp')
        received_at = None
        if ts:
            ts_str = str(ts).strip()
            # ISO 8601 형식 (Forward SMS 앱의 {time})
            if 'T' in ts_str or '-' in ts_str[:10]:
                try:
                    from django.utils.dateparse import parse_datetime
                    parsed = parse_datetime(ts_str)
                    if parsed:
                        received_at = parsed if timezone.is_aware(parsed) else timezone.make_aware(parsed)
                except (ValueError, TypeError):
                    pass
            # 밀리초 timestamp
            if not received_at:
                try:
                    received_at = timezone.make_aware(
                        datetime.datetime.fromtimestamp(int(ts_str) / 1000)
                    )
                except (ValueError, TypeError, OSError):
                    pass
        if not received_at:
            received_at = timezone.now()

        # 전화번호 추출 (from_number 또는 sender에서)
        phone_digits = from_number.replace('-', '') if from_number else ''
        if not phone_digits:
            pm = _re.search(r'(\d[\d\-]{8,})', sender)
            if pm:
                phone_digits = pm.group(1).replace('-', '')

        if content:
            InboxMessage.objects.create(
                source=InboxMessage.Source.SMS,
                sender=sender,
                content=content,
                received_at=received_at,
                phone=phone_digits,
            )
    except Exception:
        logger.exception('SMS 웹훅 처리 오류')

    return HttpResponse('OK')


@role_required('admin')
def sms_import_xml(request):
    """SMS Backup & Restore 앱의 XML 파일에서 문자 일괄 가져오기"""
    if request.method != 'POST' or not request.FILES.get('xml_file'):
        return render(request, 'orders/sms_import.html', {'result': None})

    from django.utils import timezone
    import datetime
    import xml.etree.ElementTree as ET

    xml_file = request.FILES['xml_file']
    if not xml_file.name.endswith('.xml'):
        messages.error(request, '.xml 파일만 지원합니다.')
        return render(request, 'orders/sms_import.html', {'result': None})

    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
    except ET.ParseError as e:
        messages.error(request, f'XML 파싱 오류: {e}')
        return render(request, 'orders/sms_import.html', {'result': None})

    # 기존 SMS 중복 방지: (sender, received_at) 조합
    existing = set(
        InboxMessage.objects.filter(source='sms')
        .values_list('sender', 'received_at')
    )

    new_count = 0
    skip_count = 0
    total = 0

    for sms in root.iter('sms'):
        total += 1
        address = sms.get('address', '').strip()
        body = sms.get('body', '').strip()
        date_ms = sms.get('date', '')
        sms_type = sms.get('type', '1')  # 1=수신, 2=발신
        contact_name = sms.get('contact_name', '').strip()

        if not body:
            skip_count += 1
            continue

        # timestamp 파싱
        try:
            received_at = timezone.make_aware(
                datetime.datetime.fromtimestamp(int(date_ms) / 1000)
            )
        except (ValueError, TypeError, OSError):
            received_at = timezone.now()

        # 발신자 표시: 연락처명이 있으면 "이름(번호)", 없으면 번호만
        if contact_name and contact_name != '(Unknown)':
            sender = f'{contact_name}({address})'
        else:
            sender = address

        # 발신 문자는 subject에 표시
        subject = '[발신]' if sms_type == '2' else ''

        # 중복 체크
        if (sender, received_at) in existing:
            skip_count += 1
            continue

        InboxMessage.objects.create(
            source=InboxMessage.Source.SMS,
            sender=sender,
            subject=subject,
            content=body,
            received_at=received_at,
            is_processed=True,  # 과거 문자는 처리완료 상태로
            is_read=True,
        )
        existing.add((sender, received_at))
        new_count += 1

    result = {
        'total': total,
        'new_count': new_count,
        'skip_count': skip_count,
    }
    messages.success(request, f'총 {total}건 중 {new_count}건 가져옴 (중복/빈내용 {skip_count}건 건너뜀)')
    return render(request, 'orders/sms_import.html', {'result': result})


@role_required('admin')
def sms_desk(request):
    """Google Messages + 주문 입력 분할 화면"""
    agencies = User.objects.filter(role='agency', is_active=True).order_by('name')
    agencies_json = json.dumps([{
        'id': a.pk, 'name': a.name,
    } for a in agencies], ensure_ascii=False)

    teachers = (
        User.objects
        .filter(role='teacher', is_active=True)
        .select_related('agency', 'delivery_address')
        .order_by('agency__name', 'name')
    )
    teachers_json = json.dumps([{
        'id': t.pk,
        'name': t.name,
        'phone': t.phone or '',
        'agency_id': t.agency_id,
        'delivery_id': t.delivery_address_id,
        'delivery_name': t.delivery_address.name if t.delivery_address else '',
        'delivery_address': t.delivery_address.address if t.delivery_address else '',
        'delivery_phone': t.delivery_address.phone if t.delivery_address else '',
        'has_delivery': bool(t.delivery_address),
    } for t in teachers], ensure_ascii=False)

    books = Book.objects.filter(is_active=True).select_related('publisher')
    series_list = sorted(set(b.series for b in books if b.series))
    if any(not b.series for b in books):
        series_list.append('기타')
    books_json = json.dumps([{
        'id': b.id,
        'series': b.series or '기타',
        'name': b.name,
        'publisher': b.publisher.name,
        'unit_price': math.floor(b.list_price * float(b.publisher.supply_rate) / 100),
    } for b in books], ensure_ascii=False)

    if request.method == 'POST':
        agency_id = request.POST.get('agency_id', '').strip()
        teacher_id = request.POST.get('teacher_id', '').strip()
        new_teacher_name = request.POST.get('new_teacher_name', '').strip()
        new_teacher_phone = request.POST.get('new_teacher_phone', '').strip()
        delivery_school = request.POST.get('delivery_school', '').strip()
        delivery_address = request.POST.get('delivery_address', '').strip()
        delivery_phone = request.POST.get('delivery_phone', '').strip()

        try:
            agency = User.objects.get(pk=agency_id, role='agency', is_active=True)
        except (User.DoesNotExist, ValueError):
            messages.error(request, '업체를 선택해 주세요.')
            return redirect('sms_desk')

        if teacher_id:
            try:
                teacher = User.objects.select_related('delivery_address').get(
                    pk=teacher_id, role='teacher', is_active=True
                )
            except (User.DoesNotExist, ValueError):
                messages.error(request, '선생님을 선택해 주세요.')
                return redirect('sms_desk')
        elif new_teacher_name:
            login_id = f'a_{new_teacher_phone or "nophone"}_{agency.pk}'
            if User.objects.filter(login_id=login_id).exists():
                teacher = User.objects.get(login_id=login_id)
            else:
                teacher = User(
                    login_id=login_id, role='teacher',
                    name=new_teacher_name, phone=new_teacher_phone,
                    agency=agency, must_change_password=False,
                )
                teacher.set_unusable_password()
                teacher.save()
        else:
            messages.error(request, '선생님을 선택하거나 새로 입력해 주세요.')
            return redirect('sms_desk')

        if delivery_school:
            delivery, created = DeliveryAddress.objects.get_or_create(
                agency=agency, name=delivery_school,
                defaults={'address': delivery_address, 'phone': delivery_phone},
            )
            if not created and delivery_address:
                delivery.address = delivery_address
                delivery.phone = delivery_phone
                delivery.save(update_fields=['address', 'phone'])
            teacher.delivery_address = delivery
            teacher.save(update_fields=['delivery_address'])
        elif not teacher.delivery_address:
            messages.error(request, '배송지를 입력해 주세요.')
            return redirect('sms_desk')

        items = []
        i = 0
        while f'book_{i}' in request.POST or f'custom_name_{i}' in request.POST:
            book_id = request.POST.get(f'book_{i}', '').strip()
            custom_name = request.POST.get(f'custom_name_{i}', '').strip()
            custom_price = request.POST.get(f'custom_price_{i}', '').strip()
            qty_str = request.POST.get(f'qty_{i}', '').strip()
            if book_id and qty_str:
                try:
                    qty = int(qty_str)
                    if qty > 0:
                        items.append({'book_id': int(book_id), 'qty': qty})
                except (ValueError, TypeError):
                    pass
            elif custom_name and qty_str:
                try:
                    qty = int(qty_str)
                    price = int(custom_price) if custom_price else 0
                    if qty > 0:
                        items.append({'custom_name': custom_name, 'custom_price': price, 'qty': qty})
                except (ValueError, TypeError):
                    pass
            i += 1

        if not items:
            messages.error(request, '주문할 교재를 1권 이상 선택하세요.')
        else:
            order = Order.objects.create(
                order_no=Order.generate_order_no(),
                agency=agency,
                teacher=teacher,
                delivery=teacher.delivery_address,
                memo=request.POST.get('memo', ''),
                source=Order.Source.ADMIN,
            )
            for item in items:
                if 'book_id' in item:
                    try:
                        book = Book.objects.get(id=item['book_id'], is_active=True)
                        OrderItem(order=order, book=book, quantity=item['qty']).save()
                    except Book.DoesNotExist:
                        pass
                else:
                    OrderItem(order=order, custom_book_name=item['custom_name'],
                              unit_price=item['custom_price'], quantity=item['qty']).save()
            messages.success(request, f'[{teacher.name}] 주문 완료. 주문번호: {order.order_no}')
            return redirect('sms_desk')

    return render(request, 'orders/sms_desk.html', {
        'agencies_json': agencies_json,
        'teachers_json': teachers_json,
        'series_list': series_list,
        'books_json': books_json,
    })


@role_required('admin')
def send_sms_ajax(request):
    """알리고를 통한 SMS 발송 (AJAX) — 발송 내역도 InboxMessage에 저장"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    receiver = request.POST.get('receiver', '').strip()
    message = request.POST.get('message', '').strip()
    if not receiver or not message:
        return JsonResponse({'error': '수신번호와 메시지를 입력하세요.'})
    ok = send_sms(receiver, message)
    if ok:
        from django.utils import timezone
        InboxMessage.objects.create(
            source=InboxMessage.Source.SMS,
            sender=receiver,
            subject='[발신]',
            content=message,
            received_at=timezone.now(),
            is_processed=True,
            is_read=True,
            phone=receiver.replace('-', ''),
        )
        return JsonResponse({'success': True, 'message': '발송 완료'})
    detail = getattr(send_sms, '_last_error', '알리고 설정을 확인하세요.')
    return JsonResponse({'error': f'발송 실패: {detail}'})


@role_required('admin')
def parse_order_excel(request):
    """엑셀 파일을 파싱하여 교재명·수량을 매칭한 JSON 반환"""
    from openpyxl import load_workbook

    if request.method != 'POST' or not request.FILES.get('file'):
        return HttpResponse(json.dumps({'error': '파일이 없습니다.'}),
                            content_type='application/json', status=400)

    file = request.FILES['file']
    if not file.name.endswith(('.xlsx', '.xls')):
        return HttpResponse(json.dumps({'error': '.xlsx 파일만 지원합니다.'}),
                            content_type='application/json', status=400)

    try:
        import re
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        books = Book.objects.filter(is_active=True).select_related('publisher')
        book_map = {}
        for b in books:
            book_map[b.name.strip()] = {
                'id': b.id,
                'series': b.series or '기타',
                'name': b.name,
                'publisher': b.publisher.name,
                'unit_price': math.floor(b.list_price * float(b.publisher.supply_rate) / 100),
            }

        def clean_book_name(name):
            name = re.sub(r'^[^)]+\)\s*', '', name)
            return name.strip()

        def normalize(text):
            t = text.strip()
            t = re.sub(r'[\s\-_·•:：/\\&+<>()（）【】\[\]「」『』]', '', t)
            return t.lower()

        def try_match(name):
            if name in book_map:
                return book_map[name]
            cleaned = clean_book_name(name)
            if cleaned != name and cleaned in book_map:
                return book_map[cleaned]
            norm_name = normalize(cleaned)
            for bname, binfo in book_map.items():
                if normalize(bname) == norm_name:
                    return binfo
            best = None
            best_ratio = 0
            for bname, binfo in book_map.items():
                norm_b = normalize(bname)
                shorter = min(len(norm_name), len(norm_b))
                longer = max(len(norm_name), len(norm_b))
                if shorter < 3 or longer == 0:
                    continue
                ratio = shorter / longer
                if ratio < 0.5:
                    continue
                if norm_name in norm_b or norm_b in norm_name:
                    if ratio > best_ratio:
                        best = binfo
                        best_ratio = ratio
            return best

        def is_skip_row(text):
            t = text.strip()
            if not t:
                return True
            if re.match(r'^[●•◆■□▶▷※★☆\-\*]?\s*(출판사|샘플|합계|소계|총합계)', t):
                return True
            if re.match(r'^\d{2,3}[-.\s]?\d{3,4}[-.\s]?\d{4}$', t):
                return True
            if re.match(r'^[가-힣]+\s*(시|도|군|구|읍|면|로|길|동|번지)', t):
                return True
            return False

        def parse_qty(val):
            if val is None:
                return 0
            if isinstance(val, (int, float)):
                return max(0, int(val))
            s = str(val).strip().replace(',', '')
            try:
                return max(0, int(float(s)))
            except (ValueError, TypeError):
                return 0

        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return HttpResponse(json.dumps({'error': '빈 파일입니다.'}),
                                content_type='application/json', status=400)

        # ── 헤더 자동 탐지: 처음 20행 스캔 ──
        NAME_KW = ['교재', '도서', '상품', '품명', '제목', 'book', 'title', 'name']
        QTY_KW = ['수량', '부수', '권수', 'qty', 'quantity', '주문수량', '사용수량']
        HEADER_KW = NAME_KW + QTY_KW + ['출판사', '정가', '단가', '시리즈', '과정', 'no']

        header_row_idx = None
        name_col = qty_col = None

        for ri, row in enumerate(rows[:20]):
            if not row:
                continue
            cells = [str(c).strip().lower() if c is not None else '' for c in row]
            # 이 행에 헤더 키워드가 몇 개 있는지 카운트
            hit = sum(1 for c in cells if c and any(k in c for k in HEADER_KW))
            if hit >= 2:
                header_row_idx = ri
                # 교재명 열 찾기
                for idx, c in enumerate(cells):
                    if not c:
                        continue
                    if name_col is None and any(k in c for k in NAME_KW):
                        name_col = idx
                    if qty_col is None and any(k in c for k in QTY_KW):
                        qty_col = idx
                break

        # 헤더를 못 찾으면 첫 행이 바로 데이터인 경우 (견적서 등)
        # → 가장 긴 텍스트가 있는 열을 교재명으로, 숫자 열을 수량으로 추정
        if name_col is None:
            header_row_idx = None  # 헤더 없음, row 0부터 데이터
            # 데이터 행 샘플링 (처음 20행)
            sample_rows = [r for r in rows[:20] if r and any(c is not None for c in r)]
            if sample_rows:
                col_count = max(len(r) for r in sample_rows)
                # 각 열별로: 텍스트 길이 평균 vs 숫자 비율
                col_text_len = [0] * col_count
                col_num_ratio = [0] * col_count
                col_samples = [0] * col_count
                for r in sample_rows:
                    for ci in range(min(len(r), col_count)):
                        v = r[ci]
                        if v is None:
                            continue
                        s = str(v).strip()
                        if not s:
                            continue
                        col_samples[ci] += 1
                        col_text_len[ci] += len(s)
                        try:
                            float(s.replace(',', ''))
                            col_num_ratio[ci] += 1
                        except (ValueError, TypeError):
                            pass
                # 교재명 = 텍스트가 길고 숫자 비율이 낮은 열
                best_name_score = -1
                for ci in range(col_count):
                    if col_samples[ci] < 2:
                        continue
                    avg_len = col_text_len[ci] / col_samples[ci]
                    num_pct = col_num_ratio[ci] / col_samples[ci]
                    if num_pct > 0.7:  # 숫자 위주 열은 skip
                        continue
                    if avg_len > best_name_score:
                        best_name_score = avg_len
                        name_col = ci
                # 수량 = 숫자 위주이고 값이 작은 열 (정가/금액 열 제외)
                for ci in range(col_count):
                    if ci == name_col or col_samples[ci] < 2:
                        continue
                    num_pct = col_num_ratio[ci] / col_samples[ci]
                    if num_pct < 0.5:
                        continue
                    # 값 크기 체크: 수량은 보통 1~999
                    vals = []
                    for r in sample_rows:
                        if ci < len(r) and r[ci] is not None:
                            try:
                                vals.append(abs(float(str(r[ci]).replace(',', ''))))
                            except (ValueError, TypeError):
                                pass
                    if vals:
                        avg_val = sum(vals) / len(vals)
                        if avg_val < 1000:  # 수량 범위
                            qty_col = ci
                            break

        if name_col is None:
            return HttpResponse(json.dumps({'error': '교재명 열을 찾을 수 없습니다.'}),
                                content_type='application/json', status=400)

        data_start = (header_row_idx + 1) if header_row_idx is not None else 0

        # ── 데이터 행 필터링 강화 ──
        def is_data_skip(name_val, row):
            """교재 데이터가 아닌 행 skip"""
            t = str(name_val).strip()
            if not t:
                return True
            # 숫자만 (행번호, 날짜 숫자 등)
            try:
                float(t.replace(',', ''))
                return True
            except (ValueError, TypeError):
                pass
            # 날짜 형식
            if re.match(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}', t):
                return True
            # datetime 객체
            import datetime
            if isinstance(name_val, (datetime.datetime, datetime.date)):
                return True
            # 너무 짧은 텍스트 (1~2글자이면서 한글이 아닌 경우)
            if len(t) <= 2 and not re.search(r'[가-힣]', t):
                return True
            # 메타/합계 행
            if re.match(r'^[●•◆■□▶▷※★☆\-\*]?\s*(출판사|샘플|합계|소계|총합계|거래처|전일잔액|기간누계|NO\.?)', t, re.IGNORECASE):
                return True
            # 전화번호
            if re.match(r'^\d{2,3}[-.\s]?\d{3,4}[-.\s]?\d{4}$', t):
                return True
            # 주소
            if re.match(r'^[가-힣]+\s*(시|도|군|구|읍|면|로|길|동|번지)', t):
                return True
            return False

        results = []
        for row in rows[data_start:]:
            if name_col >= len(row):
                continue
            raw_name = row[name_col]
            if raw_name is None:
                continue
            if is_data_skip(raw_name, row):
                continue
            name = str(raw_name).strip()
            qty = parse_qty(row[qty_col]) if qty_col is not None and qty_col < len(row) else 1
            if qty <= 0:
                qty = 1

            matched = try_match(name)
            if matched:
                results.append({
                    'book_id': matched['id'],
                    'series': matched['series'],
                    'name': matched['name'],
                    'publisher': matched['publisher'],
                    'unit_price': matched['unit_price'],
                    'qty': qty,
                    'matched': True,
                    'original_name': name,
                })
            else:
                results.append({
                    'book_id': None,
                    'name': name,
                    'qty': qty,
                    'matched': False,
                    'original_name': name,
                })

        matched = [r for r in results if r['matched']]
        unmatched = [r for r in results if not r['matched']]
        return HttpResponse(
            json.dumps({'matched': matched, 'unmatched': unmatched}, ensure_ascii=False),
            content_type='application/json'
        )
    except Exception as e:
        return HttpResponse(
            json.dumps({'error': f'파싱 오류: {str(e)}'}),
            content_type='application/json', status=400,
        )


# ── 슬라이드패널 API ──────────────────────────────────────────────────────────

@role_required('admin')
def inbox_detail_api(request, pk):
    """슬라이드패널용 메시지 상세 JSON"""
    msg = get_object_or_404(InboxMessage, pk=pk)

    # Mark as read
    if not msg.is_read:
        msg.is_read = True
        msg.save(update_fields=['is_read'])

    data = {
        'pk': msg.pk,
        'source': msg.source,
        'sender': msg.sender,
        'subject': msg.subject or '',
        'content': msg.content or '',
        'received_at': msg.received_at.strftime('%m/%d %H:%M') if msg.received_at else '',
        'is_processed': msg.is_processed,
        'order_pk': msg.order_id,
        'order_no': msg.order.order_no if msg.order else None,
        'attachments': [
            {'pk': att.pk, 'filename': att.filename, 'is_excel': att.is_excel,
             'is_image': att.is_image, 'is_pdf': att.is_pdf, 'is_hwp': att.is_hwp,
             'extension': att.extension,
             'download_url': f'/inbox/attachment/{att.pk}/download/',
             'preview_url': f'/inbox/attachment/{att.pk}/preview/' if (att.is_excel or att.is_image or att.is_pdf or att.is_hwp) else None}
            for att in msg.attachments.all()
        ],
    }

    # SMS: include conversation thread
    if msg.source == 'sms':
        phone = msg.phone or _extract_phone_digits(msg.sender)
        thread_msgs = []
        if phone:
            sms_qs = InboxMessage.objects.filter(source='sms', phone=phone).order_by('received_at')[:100]
            for m in sms_qs:
                thread_msgs.append({
                    'pk': m.pk,
                    'content': m.content or '',
                    'time': m.received_at.strftime('%H:%M') if m.received_at else '',
                    'date': m.received_at.strftime('%Y-%m-%d') if m.received_at else '',
                    'is_sent': m.subject == '[발신]',
                    'is_current': m.pk == msg.pk,
                })
        data['sms_thread'] = thread_msgs
        data['sms_reply_phone'] = _extract_reply_phone(msg)

    # Email: include conversation thread + business docs
    if msg.source == 'email':
        # 사업자 서류 목록 (메일발송 폼에서 사용)
        from orders.models import BusinessDocument
        data['business_docs'] = [
            {'id': d.pk, 'name': d.name, 'auto_attach': d.auto_attach}
            for d in BusinessDocument.objects.all()
        ]

        sender_str = msg.sender or ''
        email_match = _re.search(r'[\w.\-+]+@[\w.\-]+\.\w+', sender_str)
        sender_email = email_match.group(0) if email_match else ''
        data['sender_email'] = sender_email
        email_thread = []
        if sender_email:
            # 수신: sender에 해당 이메일 포함 / 발신: phone 필드에 해당 이메일 저장
            thread_qs = InboxMessage.objects.filter(
                source='email',
            ).filter(
                Q(sender__icontains=sender_email) | Q(phone=sender_email)
            ).order_by('received_at')[:50]
            for m in thread_qs:
                is_sent = (m.subject or '').startswith('[발신]')
                email_thread.append({
                    'pk': m.pk,
                    'sender': m.sender or '',
                    'subject': m.subject or '',
                    'content': (m.content or '')[:500],
                    'time': m.received_at.strftime('%H:%M') if m.received_at else '',
                    'date': m.received_at.strftime('%Y-%m-%d') if m.received_at else '',
                    'is_sent': is_sent,
                    'is_current': m.pk == msg.pk,
                })
        data['email_thread'] = email_thread

    return JsonResponse(data)


@role_required('admin')
def inbox_call_detail_api(request, pk):
    """슬라이드패널용 통화 상세 JSON"""
    rec = get_object_or_404(CallRecording, pk=pk)
    data = {
        'pk': rec.pk,
        'caller_phone': rec.caller_phone or '',
        'file_name': rec.file_name or '',
        'status': rec.status,
        'status_display': rec.get_status_display(),
        'is_order': rec.is_order,
        'summary': rec.summary or '',
        'transcript': rec.transcript or '',
        'error_msg': rec.error_msg or '',
        'recorded_at': rec.recorded_at.strftime('%Y-%m-%d %H:%M') if rec.recorded_at else '',
        'created_at': rec.created_at.strftime('%m/%d %H:%M') if rec.created_at else '',
        'audio_url': rec.audio_file.url if rec.audio_file else '',
        'order_pk': rec.order_id,
        'order_no': rec.order.order_no if rec.order else None,
        'parsed_data': rec.parsed_data or {},
    }
    return JsonResponse(data)


@role_required('admin')
def inbox_order_form_partial(request, pk):
    """슬라이드패널용 주문 폼 부분 템플릿"""
    inbox_msg = get_object_or_404(InboxMessage, pk=pk)

    books = Book.objects.filter(is_active=True).select_related('publisher')
    series_list = sorted(set(b.series for b in books if b.series))
    books_json = json.dumps([{
        'id': b.id,
        'series': b.series or '기타',
        'name': b.name,
        'publisher': b.publisher.name,
        'unit_price': math.floor(b.list_price * float(b.publisher.supply_rate) / 100),
    } for b in books], ensure_ascii=False)

    _, agencies_json = get_agencies_json()
    _, teachers_json = get_teachers_json()

    return render(request, 'orders/_inbox_order_form.html', {
        'inbox_msg': inbox_msg,
        'agencies_json': agencies_json,
        'teachers_json': teachers_json,
        'series_list': series_list,
        'books_json': books_json,
    })



@role_required('admin')
def call_order_form_partial(request, pk):
    """슬라이드패널용 통화 주문 폼 부분 템플릿"""
    rec = get_object_or_404(CallRecording, pk=pk)

    books = Book.objects.filter(is_active=True).select_related('publisher')
    series_list = sorted(set(b.series for b in books if b.series))
    books_json = json.dumps([{
        'id': b.id,
        'series': b.series or '기타',
        'name': b.name,
        'publisher': b.publisher.name,
        'unit_price': math.floor(b.list_price * float(b.publisher.supply_rate) / 100),
    } for b in books], ensure_ascii=False)

    _, agencies_json = get_agencies_json()
    _, teachers_json = get_teachers_json()

    # Pre-fill data from parsed_data
    prefill_json = 'null'
    if rec.parsed_data:
        pd = rec.parsed_data
        prefill_json = json.dumps({
            'teacher_name': pd.get('teacher_name', ''),
            'school_name': pd.get('school_name', ''),
            'phone': pd.get('phone', ''),
            'items': pd.get('items', []),
        }, ensure_ascii=False)

    return render(request, 'orders/_call_order_form.html', {
        'call_rec': rec,
        'agencies_json': agencies_json,
        'teachers_json': teachers_json,
        'series_list': series_list,
        'books_json': books_json,
        'prefill_json': prefill_json,
    })


@role_required('admin')
def inbox_next_api(request):
    """다음 미처리 메시지 API"""
    tab = request.GET.get('tab', 'email')
    exclude_pk = request.GET.get('exclude', '')

    if tab == 'call':
        qs = CallRecording.objects.filter(status__in=['pending', 'parsed'])
        if exclude_pk:
            qs = qs.exclude(pk=exclude_pk)
        rec = qs.order_by('-created_at').first()
        if rec:
            return JsonResponse({
                'pk': rec.pk,
                'type': 'call',
                'caller_phone': rec.caller_phone or rec.file_name or '녹음',
                'status': rec.status,
            })
    else:
        source = 'sms' if tab == 'sms' else 'email'
        qs = InboxMessage.objects.filter(is_processed=False, source=source)
        if exclude_pk:
            qs = qs.exclude(pk=exclude_pk)
        msg = qs.order_by('-received_at').first()
        if msg:
            return JsonResponse({
                'pk': msg.pk,
                'type': source,
                'sender': msg.sender,
                'subject': msg.subject[:50] if msg.subject else '',
            })

    return JsonResponse({'pk': None})
