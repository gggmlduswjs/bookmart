"""엑셀 → 교재 매칭 파싱 (공통 모듈)"""
import re
import math
from books.models import Book


def _build_book_map(include_unit_price=False):
    """DB에서 교재 목록을 로드하여 매칭용 딕셔너리 생성"""
    books = Book.objects.filter(is_active=True).select_related('publisher')
    book_map = {}
    for b in books:
        info = {
            'id': b.id,
            'series': b.series or '기타',
            'name': b.name,
            'publisher': b.publisher.name,
            'list_price': b.list_price,
        }
        if include_unit_price:
            info['unit_price'] = math.floor(b.list_price * float(b.publisher.supply_rate) / 100)
        book_map[b.name.strip()] = info
    return book_map


def clean_book_name(name):
    """출판사 접두사 제거: '마린) 교재명' → '교재명'"""
    name = re.sub(r'^[^)]+\)\s*', '', name)
    return name.strip()


def normalize(text):
    """매칭용 문자열 정규화: 특수문자/공백 제거"""
    t = text.strip()
    t = re.sub(r'[\s\-_·•:：/\\&+<>()（）【】\[\]「」『』]', '', t)
    return t.lower()


def try_match(name, book_map):
    """교재명 매칭: 정확→접두사제거→정규화→부분매칭 순서"""
    if name in book_map:
        return book_map[name]
    cleaned = clean_book_name(name)
    if cleaned != name and cleaned in book_map:
        return book_map[cleaned]
    norm_name = normalize(cleaned)
    for bname, binfo in book_map.items():
        if normalize(bname) == norm_name:
            return binfo
    best = None
    best_ratio = 0
    for bname, binfo in book_map.items():
        norm_b = normalize(bname)
        shorter = min(len(norm_name), len(norm_b))
        longer = max(len(norm_name), len(norm_b))
        if shorter < 3 or longer == 0:
            continue
        if shorter / longer < 0.5:
            continue
        if norm_name in norm_b or norm_b in norm_name:
            ratio = shorter / longer
            if ratio > best_ratio:
                best = binfo
                best_ratio = ratio
    return best


def is_skip_row(text):
    """교재가 아닌 메타데이터 행인지 판별"""
    t = text.strip()
    if not t:
        return True
    if re.match(r'^[●•◆■□▶▷※★☆\-\*]?\s*(출판사|샘플|합계|소계|총합계)', t):
        return True
    if re.match(r'출판사\s*[:：]', t):
        return True
    if re.match(r'(주소|연락처|전화|핸드폰|휴대폰|팩스|이메일|메일|E-?mail)\s*[:：]', t, re.IGNORECASE):
        return True
    if re.match(r'^[\d\-\s\(\)]+$', t):
        return True
    if re.search(r'01[016789][\-\s]?\d{3,4}[\-\s]?\d{4}', t):
        return True
    if re.search(r'(서울|부산|대구|인천|광주|대전|울산|세종|경기|강원|충북|충남|전북|전남|경북|경남|제주)', t):
        if re.search(r'(시|구|군|읍|면|동|로|길)\s', t):
            return True
    if t in ('NO.', 'NO', '비고', '합계', '총합계', '소계'):
        return True
    return False


def parse_qty(val):
    """수량 파싱: 13, '1권', '13부' 등 처리"""
    if val is None:
        return None
    s = str(val).strip()
    m = re.match(r'^(\d+)\s*(권|부|개|세트)?$', s)
    if m:
        v = int(m.group(1))
        return v if 0 < v < 10000 else None
    try:
        v = int(float(s))
        return v if 0 < v < 10000 else None
    except (ValueError, TypeError):
        return None


def extract_metadata(all_rows):
    """엑셀에서 업체/학교/선생님/주소/전화 정보 추출"""
    meta = {}
    for row in all_rows:
        cells = [str(c or '').strip() for c in row]
        line = ' '.join(c for c in cells if c).strip()
        if not line:
            continue
        m = re.match(r'(주소)\s*[:：]\s*(.+)', line)
        if m and not meta.get('address'):
            meta['address'] = m.group(2).strip()
            continue
        m = re.match(r'(연락처|전화|핸드폰|휴대폰)\s*[:：]\s*(.+)', line)
        if m and not meta.get('phone'):
            phone = re.sub(r'[^\d\-]', '', m.group(2))
            if phone:
                meta['phone'] = phone
            continue
        if '/' in line and not meta.get('teacher'):
            parts = [p.strip() for p in line.split('/')]
            if 2 <= len(parts) <= 4:
                all_short = all(len(p) <= 20 for p in parts)
                no_keywords = not any(k in line for k in ['주소', '출판사', '연락처', '전화'])
                if all_short and no_keywords:
                    if len(parts) >= 3:
                        meta['agency'] = parts[0]
                        meta['school'] = parts[1]
                        meta['teacher'] = parts[2]
                    elif len(parts) == 2:
                        meta['school'] = parts[0]
                        meta['teacher'] = parts[1]
        if not meta.get('school'):
            m2 = re.search(r'(\S*(?:초등학교|초|중학교|중|고등학교|고)\S*)', line)
            if m2 and '출판사' not in line:
                meta['school'] = m2.group(1)
        if not meta.get('phone'):
            m3 = re.search(r'(01[016789][\-\s]?\d{3,4}[\-\s]?\d{4})', line)
            if m3 and '출판사' not in line:
                meta['phone'] = re.sub(r'[^\d\-]', '', m3.group(1))
                if not meta.get('teacher'):
                    name_part = line[:m3.start()].strip()
                    if name_part and len(name_part) <= 10 and not any(k in name_part for k in ['주소', '출판사', '연락처']):
                        meta['teacher'] = name_part
        if not meta.get('address'):
            m4 = re.match(r'(서울|부산|대구|인천|광주|대전|울산|세종|경기|강원|충북|충남|전북|전남|경북|경남|제주)', line)
            if m4 and re.search(r'(구|군|시|읍|면|동|로|길)', line):
                addr = line.split('\n')[0].strip()
                meta['address'] = addr
                if not meta.get('school'):
                    s = re.search(r'(\S*(?:초등학교|중학교|고등학교)\S*)', addr)
                    if s:
                        meta['school'] = s.group(1)
    return meta


def parse_excel(file, include_unit_price=False, include_metadata=False, include_sample=False):
    """
    엑셀 파일을 파싱하여 교재 매칭 결과를 반환.

    Args:
        file: 업로드된 엑셀 파일
        include_unit_price: book_map에 unit_price 포함 여부 (admin용)
        include_metadata: 메타데이터(업체/학교/선생님) 추출 여부 (admin용)
        include_sample: 강사용(샘플) 컬럼 처리 여부 (admin용)

    Returns:
        dict: {'matched': [...], 'unmatched': [...], 'meta': {...}}
    """
    from openpyxl import load_workbook

    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active

    book_map = _build_book_map(include_unit_price=include_unit_price)

    all_rows = list(ws.iter_rows(values_only=True))

    meta = extract_metadata(all_rows) if include_metadata else {}

    # 헤더 탐색
    header_row = -1
    col_name = -1
    col_qty = -1
    col_price = -1
    col_sample = -1
    name_keywords = ('교재명', '도서명', '교재')
    qty_keywords = ('수량', '부수', '권수', '주문수량', '신청수량')
    price_keywords = ('단가', '정가', '가격')
    sample_keywords = ('강사용', '샘플', '강사용(샘플)')

    for i, row in enumerate(all_rows):
        cells = [str(c or '').strip().replace(' ', '') for c in row]
        for j, cell in enumerate(cells):
            cell_clean = cell.split('\n')[0].strip()
            if col_name < 0 and cell_clean in name_keywords:
                col_name = j
            if col_qty < 0:
                if cell_clean in qty_keywords or any(cell_clean.startswith(k) for k in qty_keywords):
                    col_qty = j
            if col_price < 0 and cell_clean in price_keywords:
                col_price = j
            if include_sample and col_sample < 0:
                if cell_clean in sample_keywords or any(cell_clean.startswith(k) for k in sample_keywords):
                    col_sample = j
        if col_name >= 0:
            header_row = i
            break

    matched = []
    unmatched = []

    if header_row >= 0 and col_name >= 0:
        for row in all_rows[header_row + 1:]:
            cells = list(row)
            if col_name >= len(cells):
                continue
            name = str(cells[col_name] or '').strip()
            if not name or is_skip_row(name):
                continue
            qty = 1
            if col_qty >= 0 and col_qty < len(cells):
                q = parse_qty(cells[col_qty])
                if q:
                    qty = q
            if qty <= 0:
                continue
            excel_price = 0
            if col_price >= 0 and col_price < len(cells):
                try:
                    excel_price = int(float(str(cells[col_price] or '0').replace(',', '')))
                except (ValueError, TypeError):
                    pass
            sample_qty = 0
            if include_sample and col_sample >= 0 and col_sample < len(cells):
                sq = parse_qty(cells[col_sample])
                if sq:
                    sample_qty = sq

            info = try_match(name, book_map)
            if info:
                matched.append({**info, 'qty': qty})
                if sample_qty > 0:
                    matched.append({**info, 'qty': sample_qty, 'is_sample': True})
            else:
                unmatched.append({'name': name, 'qty': qty, 'excel_price': excel_price})
                if sample_qty > 0:
                    unmatched.append({'name': name, 'qty': sample_qty, 'excel_price': excel_price, 'is_sample': True})
    else:
        in_sample_section = False
        for row in all_rows:
            cells = [str(c or '').strip() for c in row if c is not None]
            line = ' '.join(c for c in cells if c)

            if include_sample and re.search(r'샘플\s*(교재|신청)', line):
                in_sample_section = True
                continue

            name_candidate = None
            qty_candidate = 1
            for cell in cells:
                q = parse_qty(cell)
                if q is not None and cell != name_candidate:
                    qty_candidate = q
                elif len(cell) >= 2 and cell not in ('', 'NO.', 'NO') and not is_skip_row(cell):
                    name_candidate = cell
            if name_candidate:
                info = try_match(name_candidate, book_map)
                if info:
                    matched.append({**info, 'qty': qty_candidate, 'is_sample': in_sample_section if include_sample else False})
                else:
                    unmatched.append({'name': name_candidate, 'qty': qty_candidate, 'is_sample': in_sample_section if include_sample else False})

    wb.close()

    result = {'matched': matched, 'unmatched': unmatched}
    if meta:
        result['meta'] = meta
    return result
