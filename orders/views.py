import io
import json
import math
from datetime import date, datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q, Count
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from accounts.decorators import role_required
from accounts.models import User
from books.models import Book
from .models import Order, OrderItem, Return, ReturnItem, Payment, DeliveryAddress, InboxMessage, InboxAttachment
from .services import get_order_queryset, get_return_queryset, get_delivery_queryset
from .sms import send_ship_notification, send_delivery_notification


# ── 대시보드 ──────────────────────────────────────────────────────────────────

@role_required('admin')
def dashboard(request):
    today = timezone.localtime().date()
    now = timezone.localtime()

    # 상단 4개 요약 카드
    unprocessed_inbox = InboxMessage.objects.filter(is_processed=False).count()
    today_orders = Order.objects.filter(ordered_at__date=today).count()
    pending_orders = Order.objects.filter(status='pending').count()
    shipping_orders = Order.objects.filter(status='shipping').count()

    # 오늘 할 일: 미처리 수신함 최근 5건
    recent_inbox = InboxMessage.objects.filter(is_processed=False)[:5]
    # 발송 대기 주문 5건
    pending_order_list = Order.objects.filter(status='pending').order_by('-ordered_at')[:5]
    # 미확인 반품 3건
    pending_returns = Return.objects.filter(status='requested').order_by('-requested_at')[:3]

    # 마감 시간
    deadline_city = now.replace(hour=11, minute=20, second=0, microsecond=0)
    deadline_region = now.replace(hour=13, minute=50, second=0, microsecond=0)

    # 최근 발송 완료 5건
    recent_delivered = Order.objects.filter(status='delivered').order_by('-ordered_at')[:5]

    return render(request, 'orders/dashboard.html', {
        'unprocessed_inbox': unprocessed_inbox,
        'today_orders': today_orders,
        'pending_orders': pending_orders,
        'shipping_orders': shipping_orders,
        'recent_inbox': recent_inbox,
        'pending_order_list': pending_order_list,
        'pending_returns': pending_returns,
        'deadline_city': deadline_city,
        'deadline_region': deadline_region,
        'now': now,
        'recent_delivered': recent_delivered,
    })


# ── 업체 대시보드 ─────────────────────────────────────────────────────────────

@role_required('agency')
def agency_dashboard(request):
    user = request.user
    today = timezone.localtime().date()
    now = timezone.localtime()

    # 소속 학교
    deliveries = DeliveryAddress.objects.filter(agency=user, is_active=True)
    teachers = User.objects.filter(role='teacher', agency=user, is_active=True)

    # 주문 통계
    my_orders = Order.objects.filter(agency=user)
    today_orders = my_orders.filter(ordered_at__date=today).count()
    pending_orders = my_orders.filter(status='pending').count()
    shipping_orders = my_orders.filter(status='shipping').count()
    total_schools = deliveries.count()

    # 이번 달 주문금액
    month_start = today.replace(day=1)
    month_items = OrderItem.objects.filter(
        order__agency=user,
        order__status__in=['pending', 'shipping', 'delivered'],
        order__ordered_at__date__gte=month_start,
    )
    month_amount = month_items.aggregate(total=Sum('amount'))['total'] or 0

    # 최근 주문 10건
    recent_orders = my_orders.select_related(
        'teacher', 'delivery'
    ).order_by('-ordered_at')[:10]

    # 학교별 최근 주문 현황
    school_stats = []
    for d in deliveries:
        last_order = my_orders.filter(delivery=d).order_by('-ordered_at').first()
        school_teachers = teachers.filter(delivery_address=d).count()
        school_stats.append({
            'school': d,
            'teacher_count': school_teachers,
            'last_order': last_order,
            'pending': my_orders.filter(delivery=d, status='pending').count(),
            'shipping': my_orders.filter(delivery=d, status='shipping').count(),
        })

    # 미확인 반품
    pending_returns = Return.objects.filter(
        agency=user, status='requested'
    ).select_related('teacher', 'delivery').order_by('-requested_at')[:5]

    return render(request, 'orders/agency_dashboard.html', {
        'today_orders': today_orders,
        'pending_orders': pending_orders,
        'shipping_orders': shipping_orders,
        'total_schools': total_schools,
        'month_amount': month_amount,
        'recent_orders': recent_orders,
        'school_stats': school_stats,
        'pending_returns': pending_returns,
        'now': now,
    })


# ── 주문 목록 ──────────────────────────────────────────────────────────────────

@login_required
def order_list(request):
    qs = get_order_queryset(request.user)

    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    delivery_id = request.GET.get('delivery', '')

    if status:
        qs = qs.filter(status=status)
    if date_from:
        qs = qs.filter(ordered_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(ordered_at__date__lte=date_to)
    if delivery_id:
        qs = qs.filter(delivery_id=delivery_id)

    # 마감 시간 기준: 11:20 시내, 13:50 지방
    now = timezone.localtime()
    deadline_city = now.replace(hour=11, minute=20, second=0, microsecond=0)
    deadline_region = now.replace(hour=13, minute=50, second=0, microsecond=0)
    past_city = now > deadline_city
    past_region = now > deadline_region

    deliveries = get_delivery_queryset(request.user)

    return render(request, 'orders/order_list.html', {
        'orders': qs.order_by('-ordered_at')[:200],
        'deliveries': deliveries,
        'status_choices': Order.Status.choices,
        'filters': {
            'status': status, 'date_from': date_from,
            'date_to': date_to, 'delivery': delivery_id,
        },
        'past_city': past_city,
        'past_region': past_region,
    })


# ── 주문 입력 ──────────────────────────────────────────────────────────────────

@role_required('teacher')
def order_create(request):
    user = request.user
    delivery = user.delivery_address
    if not delivery:
        messages.error(request, '담당 학교가 지정되지 않았습니다. 업체에 문의하세요.')
        return redirect('home')

    # 마감 시간 경고
    now = timezone.localtime()
    deadline_city = now.replace(hour=11, minute=20, second=0, microsecond=0)
    deadline_region = now.replace(hour=13, minute=50, second=0, microsecond=0)
    past_city = now > deadline_city
    past_region = now > deadline_region

    books = Book.objects.filter(is_active=True).select_related('publisher')
    series_list = sorted(set(b.series for b in books if b.series))
    books_json = json.dumps([{
        'id': b.id,
        'series': b.series or '기타',
        'name': b.name,
        'publisher': b.publisher.name,
        'unit_price': math.floor(b.list_price * float(b.publisher.supply_rate) / 100),
    } for b in books], ensure_ascii=False)

    if request.method == 'POST':
        items = []
        i = 0
        while f'book_{i}' in request.POST:
            book_id = request.POST.get(f'book_{i}', '').strip()
            qty_str = request.POST.get(f'qty_{i}', '').strip()
            if book_id and qty_str:
                try:
                    qty = int(qty_str)
                    if qty > 0:
                        items.append((int(book_id), qty))
                except (ValueError, TypeError):
                    pass
            i += 1

        if not items:
            messages.error(request, '주문할 교재를 1권 이상 선택하세요.')
        else:
            order = Order.objects.create(
                order_no=Order.generate_order_no(),
                agency=user.agency,
                teacher=user,
                delivery=delivery,
                memo=request.POST.get('memo', ''),
            )
            for book_id, qty in items:
                try:
                    book = Book.objects.get(id=book_id, is_active=True)
                    OrderItem(order=order, book=book, quantity=qty).save()
                except Book.DoesNotExist:
                    pass
            messages.success(request, f'주문 완료. 주문번호: {order.order_no}')
            return redirect('order_detail', pk=order.pk)

    return render(request, 'orders/order_create.html', {
        'delivery': delivery,
        'series_list': series_list,
        'books_json': books_json,
        'past_city': past_city,
        'past_region': past_region,
        'now_str': now.strftime('%H:%M'),
    })


# ── 주문 상세 ──────────────────────────────────────────────────────────────────

@login_required
def order_detail(request, pk):
    order = get_object_or_404(get_order_queryset(request.user), pk=pk)
    items = order.items.select_related('book', 'book__publisher')
    return render(request, 'orders/order_detail.html', {
        'order': order,
        'items': items,
        'can_cancel': (
            order.status == Order.Status.PENDING
            and request.user.role == 'teacher'
            and order.teacher == request.user
        ),
        'can_ship': (
            request.user.role == 'admin'
            and order.status == Order.Status.PENDING
        ),
        'can_deliver': (
            request.user.role == 'admin'
            and order.status == Order.Status.SHIPPING
        ),
    })


# ── 총판 대리 주문 ─────────────────────────────────────────────────────────────

@role_required('admin')
def order_create_admin(request):
    """총판이 전화 주문을 받아 대신 입력하는 뷰"""
    agencies = User.objects.filter(role='agency', is_active=True).order_by('name')
    agencies_json = json.dumps([{
        'id': a.pk, 'name': a.name,
    } for a in agencies], ensure_ascii=False)

    teachers = (
        User.objects
        .filter(role='teacher', is_active=True)
        .select_related('agency', 'delivery_address')
        .order_by('agency__name', 'name')
    )
    teachers_json = json.dumps([{
        'id': t.pk,
        'name': t.name,
        'phone': t.phone or '',
        'agency_id': t.agency_id,
        'delivery_id': t.delivery_address_id,
        'delivery_name': t.delivery_address.name if t.delivery_address else '',
        'delivery_address': t.delivery_address.address if t.delivery_address else '',
        'delivery_phone': t.delivery_address.phone if t.delivery_address else '',
        'has_delivery': bool(t.delivery_address),
    } for t in teachers], ensure_ascii=False)

    # 마감 시간 경고
    now = timezone.localtime()
    deadline_city = now.replace(hour=11, minute=20, second=0, microsecond=0)
    deadline_region = now.replace(hour=13, minute=50, second=0, microsecond=0)
    past_city = now > deadline_city
    past_region = now > deadline_region

    books = Book.objects.filter(is_active=True).select_related('publisher')
    series_list = sorted(set(b.series for b in books if b.series))
    books_json = json.dumps([{
        'id': b.id,
        'series': b.series or '기타',
        'name': b.name,
        'publisher': b.publisher.name,
        'unit_price': math.floor(b.list_price * float(b.publisher.supply_rate) / 100),
    } for b in books], ensure_ascii=False)

    if request.method == 'POST':
        agency_id = request.POST.get('agency_id', '').strip()
        teacher_id = request.POST.get('teacher_id', '').strip()
        new_teacher_name = request.POST.get('new_teacher_name', '').strip()
        new_teacher_phone = request.POST.get('new_teacher_phone', '').strip()
        delivery_school = request.POST.get('delivery_school', '').strip()
        delivery_address = request.POST.get('delivery_address', '').strip()
        delivery_phone = request.POST.get('delivery_phone', '').strip()

        # 업체 확인
        try:
            agency = User.objects.get(pk=agency_id, role='agency', is_active=True)
        except (User.DoesNotExist, ValueError):
            messages.error(request, '업체를 선택해 주세요.')
            return redirect('order_create_admin')

        # 선생님: 기존 선택 또는 신규 생성
        if teacher_id:
            try:
                teacher = User.objects.select_related('delivery_address').get(
                    pk=teacher_id, role='teacher', is_active=True
                )
            except (User.DoesNotExist, ValueError):
                messages.error(request, '선생님을 선택해 주세요.')
                return redirect('order_create_admin')
        elif new_teacher_name:
            login_id = f'a_{new_teacher_phone or "nophone"}_{agency.pk}'
            if User.objects.filter(login_id=login_id).exists():
                teacher = User.objects.get(login_id=login_id)
            else:
                teacher = User(
                    login_id=login_id, role='teacher',
                    name=new_teacher_name, phone=new_teacher_phone,
                    agency=agency, must_change_password=False,
                )
                teacher.set_unusable_password()
                teacher.save()
        else:
            messages.error(request, '선생님을 선택하거나 새로 입력해 주세요.')
            return redirect('order_create_admin')

        # 배송지: 입력된 학교명이 있으면 생성/매칭
        if delivery_school:
            delivery, created = DeliveryAddress.objects.get_or_create(
                agency=agency, name=delivery_school,
                defaults={'address': delivery_address, 'phone': delivery_phone},
            )
            if not created and delivery_address:
                delivery.address = delivery_address
                delivery.phone = delivery_phone
                delivery.save(update_fields=['address', 'phone'])
            teacher.delivery_address = delivery
            teacher.save(update_fields=['delivery_address'])
        elif not teacher.delivery_address:
            messages.error(request, '배송지를 입력해 주세요.')
            return redirect('order_create_admin')

        items = []
        i = 0
        while f'book_{i}' in request.POST:
            book_id = request.POST.get(f'book_{i}', '').strip()
            qty_str = request.POST.get(f'qty_{i}', '').strip()
            if book_id and qty_str:
                try:
                    qty = int(qty_str)
                    if qty > 0:
                        items.append((int(book_id), qty))
                except (ValueError, TypeError):
                    pass
            i += 1

        if not items:
            messages.error(request, '주문할 교재를 1권 이상 선택하세요.')
        else:
            order = Order.objects.create(
                order_no=Order.generate_order_no(),
                agency=agency,
                teacher=teacher,
                delivery=teacher.delivery_address,
                memo=request.POST.get('memo', ''),
            )
            for book_id, qty in items:
                try:
                    book = Book.objects.get(id=book_id, is_active=True)
                    OrderItem(order=order, book=book, quantity=qty).save()
                except Book.DoesNotExist:
                    pass
            messages.success(request, f'[{teacher.name}] 대리 주문 완료. 주문번호: {order.order_no}')
            return redirect('order_detail', pk=order.pk)

    return render(request, 'orders/order_create_admin.html', {
        'agencies': agencies,
        'agencies_json': agencies_json,
        'teachers_json': teachers_json,
        'series_list': series_list,
        'books_json': books_json,
        'past_city': past_city,
        'past_region': past_region,
        'now_str': now.strftime('%H:%M'),
    })


# ── 주문 취소 (선생님) ─────────────────────────────────────────────────────────

@role_required('teacher')
def order_cancel(request, pk):
    order = get_object_or_404(
        get_order_queryset(request.user),
        pk=pk, teacher=request.user, status=Order.Status.PENDING
    )
    if request.method == 'POST':
        order.status = Order.Status.CANCELLED
        order.save(update_fields=['status'])
        messages.success(request, f'주문 {order.order_no}이 취소되었습니다.')
        return redirect('order_list')
    return render(request, 'orders/order_cancel_confirm.html', {'order': order})


# ── 발송 처리 (총판) ───────────────────────────────────────────────────────────

@role_required('admin')
def order_ship(request, pk):
    order = get_object_or_404(Order, pk=pk, status=Order.Status.PENDING)
    if request.method == 'POST':
        order.tracking_no = request.POST.get('tracking_no', '').strip()
        order.status = Order.Status.SHIPPING
        order.save(update_fields=['status', 'tracking_no'])

        # 문자 발송
        sms_ok = send_ship_notification(order)
        if sms_ok:
            messages.success(request, f'주문 {order.order_no} → 발송중 처리 + 선생님께 문자 발송 완료.')
        else:
            messages.success(request, f'주문 {order.order_no} → 발송중 처리 완료.')
            if not (order.teacher.phone or order.teacher.mobile):
                messages.warning(request, '선생님 전화번호가 등록되어 있지 않아 문자를 보내지 못했습니다.')
    return redirect('order_detail', pk=pk)


@role_required('admin')
def order_deliver(request, pk):
    order = get_object_or_404(Order, pk=pk, status=Order.Status.SHIPPING)
    if request.method == 'POST':
        order.status = Order.Status.DELIVERED
        order.save(update_fields=['status'])
        send_delivery_notification(order)
        messages.success(request, f'주문 {order.order_no} → 발송완료 처리되었습니다.')
    return redirect('order_detail', pk=pk)


# ── 거래명세서 인쇄 ────────────────────────────────────────────────────────────

@role_required('admin', 'agency')
def order_invoice(request, pk):
    order = get_object_or_404(Order, pk=pk)
    items = order.items.select_related('book', 'book__publisher')
    total_amount = sum(item.amount for item in items)
    total_qty = sum(item.quantity for item in items)
    # 빈 행 채우기 (최소 10행)
    empty_rows = range(max(0, 10 - items.count()))
    return render(request, 'orders/order_invoice.html', {
        'order': order,
        'items': items,
        'total_amount': total_amount,
        'total_qty': total_qty,
        'empty_rows': empty_rows,
    })


@role_required('admin', 'agency')
def order_invoice_bulk(request):
    ids_str = request.GET.get('ids', '')
    try:
        ids = [int(x) for x in ids_str.split(',') if x.strip()]
    except ValueError:
        ids = []
    if not ids:
        return HttpResponse('인쇄할 주문이 없습니다.', status=400)

    orders_data = []
    qs = Order.objects.filter(pk__in=ids).select_related(
        'teacher', 'delivery', 'agency'
    ).order_by('ordered_at')
    if request.user.role == 'agency':
        qs = qs.filter(agency=request.user)

    for order in qs:
        items = order.items.select_related('book', 'book__publisher')
        total_amount = sum(item.amount for item in items)
        total_qty = sum(item.quantity for item in items)
        empty_rows = range(max(0, 8 - items.count()))
        orders_data.append({
            'order': order,
            'items': items,
            'total_amount': total_amount,
            'total_qty': total_qty,
            'empty_rows': empty_rows,
        })

    return render(request, 'orders/order_invoice_bulk.html', {
        'orders_data': orders_data,
    })


# ── 반품 목록 ──────────────────────────────────────────────────────────────────

@login_required
def return_list(request):
    qs = get_return_queryset(request.user)

    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    delivery_id = request.GET.get('delivery', '')

    if status:
        qs = qs.filter(status=status)
    if date_from:
        qs = qs.filter(requested_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(requested_at__date__lte=date_to)
    if delivery_id:
        qs = qs.filter(delivery_id=delivery_id)

    deliveries = get_delivery_queryset(request.user)

    return render(request, 'orders/return_list.html', {
        'returns': qs.order_by('-requested_at')[:200],
        'deliveries': deliveries,
        'status_choices': Return.Status.choices,
        'filters': {
            'status': status, 'date_from': date_from,
            'date_to': date_to, 'delivery': delivery_id,
        },
    })


# ── 반품 신청 (선생님) ─────────────────────────────────────────────────────────

@role_required('teacher')
def return_create(request):
    user = request.user
    delivery = user.delivery_address
    if not delivery:
        messages.error(request, '담당 학교가 지정되지 않았습니다.')
        return redirect('home')

    # 판본도 제외한 교재만
    books = Book.objects.filter(is_active=True, is_returnable=True).select_related('publisher')
    series_list = sorted(set(b.series for b in books if b.series))
    books_json = json.dumps([{
        'id': b.id,
        'series': b.series or '기타',
        'name': b.name,
        'publisher': b.publisher.name,
        'unit_price': math.floor(b.list_price * float(b.publisher.supply_rate) / 100),
    } for b in books], ensure_ascii=False)

    if request.method == 'POST':
        items = []
        i = 0
        while f'book_{i}' in request.POST:
            book_id = request.POST.get(f'book_{i}', '').strip()
            qty_str = request.POST.get(f'qty_{i}', '').strip()
            if book_id and qty_str:
                try:
                    qty = int(qty_str)
                    if qty > 0:
                        items.append((int(book_id), qty))
                except (ValueError, TypeError):
                    pass
            i += 1

        if not items:
            messages.error(request, '반품할 교재를 1권 이상 선택하세요.')
        else:
            ret = Return.objects.create(
                return_no=Return.generate_return_no(),
                agency=user.agency,
                teacher=user,
                delivery=delivery,
                memo=request.POST.get('memo', ''),
            )
            for book_id, qty in items:
                try:
                    book = Book.objects.get(id=book_id, is_active=True, is_returnable=True)
                    ReturnItem(ret=ret, book=book, requested_qty=qty).save()
                except Book.DoesNotExist:
                    pass
            messages.success(request, f'반품 신청 완료. 반품번호: {ret.return_no}')
            return redirect('return_detail', pk=ret.pk)

    return render(request, 'orders/return_create.html', {
        'delivery': delivery,
        'series_list': series_list,
        'books_json': books_json,
    })


# ── 반품 상세 ──────────────────────────────────────────────────────────────────

@login_required
def return_detail(request, pk):
    ret = get_object_or_404(get_return_queryset(request.user), pk=pk)
    items = ret.items.select_related('book', 'book__publisher')
    return render(request, 'orders/return_detail.html', {
        'ret': ret,
        'items': items,
        'can_confirm': request.user.role == 'admin' and ret.status == Return.Status.REQUESTED,
        'can_reject': request.user.role == 'admin' and ret.status == Return.Status.REQUESTED,
    })


# ── 반품 확정 (총판) ───────────────────────────────────────────────────────────

@role_required('admin')
def return_confirm(request, pk):
    ret = get_object_or_404(Return, pk=pk, status=Return.Status.REQUESTED)
    items = ret.items.select_related('book')

    if request.method == 'POST':
        for item in items:
            confirmed_qty_str = request.POST.get(f'confirmed_qty_{item.pk}', '0')
            adjusted_str = request.POST.get(f'adjusted_{item.pk}', '0')
            try:
                item.confirmed_qty = max(0, int(confirmed_qty_str))
                item.adjusted_amount = int(adjusted_str)
            except (ValueError, TypeError):
                item.confirmed_qty = 0
                item.adjusted_amount = 0
            item.save()

        ret.status = Return.Status.CONFIRMED
        ret.confirmed_at = timezone.now()
        ret.save(update_fields=['status', 'confirmed_at'])
        messages.success(request, f'반품 {ret.return_no} 확정 처리되었습니다.')
        return redirect('return_detail', pk=pk)

    return render(request, 'orders/return_confirm.html', {'ret': ret, 'items': items})


# ── 반품 거절 (총판) ───────────────────────────────────────────────────────────

@role_required('admin')
def return_reject(request, pk):
    ret = get_object_or_404(Return, pk=pk, status=Return.Status.REQUESTED)
    if request.method == 'POST':
        ret.status = Return.Status.REJECTED
        ret.memo = request.POST.get('memo', ret.memo)
        ret.save(update_fields=['status', 'memo'])
        messages.success(request, f'반품 {ret.return_no} 거절 처리되었습니다.')
        return redirect('return_detail', pk=pk)
    return render(request, 'orders/return_reject.html', {'ret': ret})


# ── 거래내역 (V03) ─────────────────────────────────────────────────────────────

@login_required
def ledger(request):
    today = date.today()
    # 총판은 업체 선택 가능, 업체는 본인만
    if request.user.role == 'admin':
        agencies = User.objects.filter(role='agency', is_active=True).order_by('name')
        agency_id = request.GET.get('agency', '')
        if agency_id:
            selected_agency = get_object_or_404(User, pk=agency_id, role='agency')
        else:
            selected_agency = agencies.first()
    else:
        agencies = None
        selected_agency = request.user
        agency_id = str(request.user.pk)

    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))

    rows = []
    total_sales = total_returns = total_paid = 0

    if selected_agency:
        # 주문(매출)
        order_items = (
            OrderItem.objects
            .filter(
                order__agency=selected_agency,
                order__status__in=[Order.Status.SHIPPING, Order.Status.DELIVERED],
                order__ordered_at__year=year,
                order__ordered_at__month=month,
            )
            .select_related('order', 'order__delivery', 'book', 'book__publisher')
            .order_by('order__ordered_at')
        )
        for oi in order_items:
            rows.append({
                'date': oi.order.ordered_at.date(),
                'type': '매출',
                'delivery': oi.order.delivery.name,
                'publisher': oi.book.publisher.name,
                'book': oi.book.name,
                'qty': oi.quantity,
                'list_price': oi.list_price,
                'supply_rate': oi.supply_rate,
                'amount': oi.amount,
            })
            total_sales += oi.amount

        # 반품
        return_items = (
            ReturnItem.objects
            .filter(
                ret__agency=selected_agency,
                ret__status=Return.Status.CONFIRMED,
                ret__confirmed_at__year=year,
                ret__confirmed_at__month=month,
            )
            .select_related('ret', 'ret__delivery', 'book', 'book__publisher')
            .order_by('ret__confirmed_at')
        )
        for ri in return_items:
            confirmed_amount = ri.confirmed_amount or 0
            rows.append({
                'date': ri.ret.confirmed_at.date(),
                'type': '반품',
                'delivery': ri.ret.delivery.name,
                'publisher': ri.book.publisher.name,
                'book': ri.book.name,
                'qty': -(ri.confirmed_qty or 0),
                'list_price': ri.list_price,
                'supply_rate': ri.supply_rate,
                'amount': -confirmed_amount,
            })
            total_returns += confirmed_amount

        # 입금
        payments = Payment.objects.filter(
            agency=selected_agency,
            paid_at__year=year,
            paid_at__month=month,
        ).order_by('paid_at')
        for p in payments:
            total_paid += p.amount

        rows.sort(key=lambda r: r['date'])

    balance = total_sales - total_returns - total_paid  # 채권잔액

    return render(request, 'orders/ledger.html', {
        'agencies': agencies,
        'selected_agency': selected_agency,
        'agency_id': agency_id,
        'year': year,
        'month': month,
        'years': range(today.year - 2, today.year + 1),
        'months': range(1, 13),
        'rows': rows,
        'total_sales': total_sales,
        'total_returns': total_returns,
        'total_paid': total_paid,
        'balance': balance,
        'payments': payments if selected_agency else [],
    })


# ── 판매현황 (V04) ─────────────────────────────────────────────────────────────

@login_required
def sales_report(request):
    today = date.today()
    date_from = request.GET.get('date_from', today.strftime('%Y-%m-01'))
    date_to = request.GET.get('date_to', today.strftime('%Y-%m-%d'))

    qs = OrderItem.objects.filter(
        order__status__in=[Order.Status.SHIPPING, Order.Status.DELIVERED],
        order__ordered_at__date__gte=date_from,
        order__ordered_at__date__lte=date_to,
    ).select_related('order', 'order__teacher', 'order__delivery', 'order__agency',
                     'book', 'book__publisher').order_by('order__ordered_at')

    if request.user.role == 'agency':
        qs = qs.filter(order__agency=request.user)
    elif request.user.role == 'teacher':
        qs = qs.filter(order__teacher=request.user)

    total_amount = sum(i.amount for i in qs)

    return render(request, 'orders/sales_report.html', {
        'items': qs,
        'date_from': date_from,
        'date_to': date_to,
        'total_amount': total_amount,
    })


# ── 발주 집계 (V05, 총판 전용) ─────────────────────────────────────────────────

@role_required('admin')
def purchase_order(request):
    today = date.today()
    date_from = request.GET.get('date_from', today.strftime('%Y-%m-01'))
    date_to = request.GET.get('date_to', today.strftime('%Y-%m-%d'))

    items = (
        OrderItem.objects
        .filter(
            order__status__in=[Order.Status.PENDING, Order.Status.SHIPPING, Order.Status.DELIVERED],
            order__ordered_at__date__gte=date_from,
            order__ordered_at__date__lte=date_to,
        )
        .select_related('book', 'book__publisher')
        .order_by('book__publisher__name', 'book__series', 'book__name')
    )

    # 출판사 → 교재별 집계
    from collections import defaultdict
    publishers = defaultdict(lambda: {'books': defaultdict(lambda: {'name': '', 'series': '', 'qty': 0, 'amount': 0})})
    for item in items:
        pub = item.book.publisher.name
        book_id = item.book.pk
        publishers[pub]['books'][book_id]['name'] = item.book.name
        publishers[pub]['books'][book_id]['series'] = item.book.series
        publishers[pub]['books'][book_id]['qty'] += item.quantity
        publishers[pub]['books'][book_id]['amount'] += item.amount

    # 정렬된 구조로 변환
    pub_list = []
    for pub_name, data in sorted(publishers.items()):
        book_rows = sorted(data['books'].values(), key=lambda b: (b['series'], b['name']))
        pub_total_qty = sum(b['qty'] for b in book_rows)
        pub_total_amount = sum(b['amount'] for b in book_rows)
        pub_list.append({
            'name': pub_name,
            'books': book_rows,
            'total_qty': pub_total_qty,
            'total_amount': pub_total_amount,
        })

    return render(request, 'orders/purchase_order.html', {
        'publishers': pub_list,
        'date_from': date_from,
        'date_to': date_to,
        'grand_total': sum(p['total_amount'] for p in pub_list),
    })


# ── 입금 등록 (총판 전용) ──────────────────────────────────────────────────────

@role_required('admin')
def payment_create(request):
    agencies = User.objects.filter(role='agency', is_active=True).order_by('name')
    if request.method == 'POST':
        agency_id = request.POST.get('agency')
        amount_str = request.POST.get('amount', '0').replace(',', '')
        paid_at = request.POST.get('paid_at', '')
        memo = request.POST.get('memo', '')
        try:
            agency = User.objects.get(pk=agency_id, role='agency')
            amount = int(amount_str)
            Payment.objects.create(agency=agency, amount=amount, paid_at=paid_at, memo=memo)
            messages.success(request, f'{agency.name} 입금 {amount:,}원 등록 완료.')
            return redirect('ledger')
        except (User.DoesNotExist, ValueError) as e:
            messages.error(request, '입력값을 확인해주세요.')
    return render(request, 'orders/payment_form.html', {'agencies': agencies})


# ── 엑셀 Export ────────────────────────────────────────────────────────────────

def _make_workbook():
    try:
        import openpyxl
        return openpyxl, openpyxl.Workbook()
    except ImportError:
        return None, None


@login_required
def export_ledger(request):
    openpyxl, wb = _make_workbook()
    if not wb:
        messages.error(request, 'openpyxl 패키지가 필요합니다: uv add openpyxl')
        return redirect('ledger')

    today = date.today()
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))
    agency_id = request.GET.get('agency', '')

    if request.user.role == 'admin' and agency_id:
        selected_agency = get_object_or_404(User, pk=agency_id, role='agency')
    elif request.user.role == 'agency':
        selected_agency = request.user
    else:
        return redirect('ledger')

    ws = wb.active
    ws.title = '거래내역'
    ws.append(['날짜', '구분', '배송지', '출판사', '교재명', '수량', '정가', '공급률', '금액'])

    order_items = OrderItem.objects.filter(
        order__agency=selected_agency,
        order__status__in=[Order.Status.SHIPPING, Order.Status.DELIVERED],
        order__ordered_at__year=year, order__ordered_at__month=month,
    ).select_related('order', 'order__delivery', 'book', 'book__publisher').order_by('order__ordered_at')

    for oi in order_items:
        ws.append([
            oi.order.ordered_at.strftime('%Y-%m-%d'), '매출',
            oi.order.delivery.name, oi.book.publisher.name, oi.book.name,
            oi.quantity, oi.list_price, float(oi.supply_rate), oi.amount,
        ])

    return_items = ReturnItem.objects.filter(
        ret__agency=selected_agency, ret__status=Return.Status.CONFIRMED,
        ret__confirmed_at__year=year, ret__confirmed_at__month=month,
    ).select_related('ret', 'ret__delivery', 'book', 'book__publisher').order_by('ret__confirmed_at')

    for ri in return_items:
        ws.append([
            ri.ret.confirmed_at.strftime('%Y-%m-%d'), '반품',
            ri.ret.delivery.name, ri.book.publisher.name, ri.book.name,
            -(ri.confirmed_qty or 0), ri.list_price, float(ri.supply_rate),
            -(ri.confirmed_amount or 0),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'거래내역_{selected_agency.name}_{year}년{month}월.xlsx'
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
    return resp


@login_required
def export_sales(request):
    openpyxl, wb = _make_workbook()
    if not wb:
        messages.error(request, 'openpyxl 패키지가 필요합니다: uv add openpyxl')
        return redirect('sales_report')

    today = date.today()
    date_from = request.GET.get('date_from', today.strftime('%Y-%m-01'))
    date_to = request.GET.get('date_to', today.strftime('%Y-%m-%d'))

    qs = OrderItem.objects.filter(
        order__status__in=[Order.Status.SHIPPING, Order.Status.DELIVERED],
        order__ordered_at__date__gte=date_from,
        order__ordered_at__date__lte=date_to,
    ).select_related('order', 'order__teacher', 'order__delivery', 'order__agency',
                     'book', 'book__publisher').order_by('order__ordered_at')

    if request.user.role == 'agency':
        qs = qs.filter(order__agency=request.user)

    ws = wb.active
    ws.title = '판매현황'
    ws.append(['출고일', '업체', '배송지', '선생님', '출판사', '시리즈', '교재명', '수량', '정가', '공급률', '금액'])
    for oi in qs:
        ws.append([
            oi.order.ordered_at.strftime('%Y-%m-%d'),
            oi.order.agency.name, oi.order.delivery.name, oi.order.teacher.name,
            oi.book.publisher.name, oi.book.series, oi.book.name,
            oi.quantity, oi.list_price, float(oi.supply_rate), oi.amount,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'판매현황_{date_from}_{date_to}.xlsx'
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
    return resp


# ── 통합 수신함 ────────────────────────────────────────────────────────────────

@role_required('admin')
def inbox_list(request):
    show_done = request.GET.get('done', '')
    tab = request.GET.get('tab', 'email')  # email or sms
    qs = InboxMessage.objects.annotate(
        attachment_count=Count('attachments')
    ).select_related('order')
    if not show_done:
        qs = qs.filter(is_processed=False)
    email_qs = qs.filter(source='email')
    sms_qs = qs.filter(source='sms')
    unread_email = InboxMessage.objects.filter(is_read=False, source='email').count()
    unread_sms = InboxMessage.objects.filter(is_read=False, source='sms').count()
    return render(request, 'orders/inbox_list.html', {
        'email_messages': email_qs[:200],
        'sms_messages': sms_qs[:200],
        'tab': tab,
        'show_done': show_done,
        'unread_email': unread_email,
        'unread_sms': unread_sms,
        'unread_count': unread_email + unread_sms,
    })


@role_required('admin')
def inbox_bulk_skip(request):
    """선택한 수신 메시지를 일괄 건너뛰기 (is_processed=True)"""
    if request.method != 'POST':
        return redirect('inbox_list')

    action = request.POST.get('action', '')
    tab = request.POST.get('tab', 'email')
    if action == 'skip_all':
        updated = InboxMessage.objects.filter(is_processed=False, source='email').update(is_processed=True)
        messages.success(request, f'미처리 메일 {updated}건을 모두 건너뛰었습니다.')
    elif action == 'skip_all_sms':
        updated = InboxMessage.objects.filter(is_processed=False, source='sms').update(is_processed=True)
        messages.success(request, f'미처리 문자 {updated}건을 모두 건너뛰었습니다.')
    else:
        msg_ids = request.POST.getlist('msg_ids')
        if msg_ids:
            updated = InboxMessage.objects.filter(
                pk__in=msg_ids, is_processed=False
            ).update(is_processed=True)
            messages.success(request, f'{updated}건을 건너뛰었습니다.')
        else:
            messages.warning(request, '선택된 메시지가 없습니다.')

    return redirect(f'/inbox/?tab={tab}')


@role_required('admin')
def fetch_emails(request):
    """네이버 IMAP 메일 가져오기"""
    if request.method != 'POST':
        return redirect('inbox_list')

    from django.conf import settings as conf
    from .email_utils import fetch_naver_emails, is_order_related

    accounts = [
        (conf.NAVER_EMAIL_1_ID, conf.NAVER_EMAIL_1_PW, '007bm'),
        (conf.NAVER_EMAIL_2_ID, conf.NAVER_EMAIL_2_PW, '002bm'),
    ]

    from django.core.files.base import ContentFile

    # DB에 있는 imap_key를 미리 전부 조회 (IMAP 다운로드 전 필터링용)
    all_existing_keys = set(
        InboxMessage.objects.values_list('imap_key', flat=True)
    )

    new_count = 0
    sync_count = 0
    for acc_id, acc_pw, label in accounts:
        if not acc_id or not acc_pw:
            continue
        emails, read_sync = fetch_naver_emails(acc_id, acc_pw, label,
                                               existing_keys=all_existing_keys)

        # 기존 메일 읽음 상태 동기화
        if read_sync:
            for imap_key, is_seen in read_sync.items():
                updated = InboxMessage.objects.filter(
                    imap_key=imap_key, is_read=not is_seen
                ).update(is_read=is_seen)
                sync_count += updated

        if not emails:
            continue

        for e in emails:
            # 스팸 판별: 주문 관련이 아니면 자동으로 처리완료 표시
            auto_skip = not is_order_related(e['sender'], e['subject'], e['content'])
            msg_obj = InboxMessage.objects.create(
                source=InboxMessage.Source.EMAIL,
                account_label=e['account_label'],
                sender=e['sender'],
                subject=e['subject'],
                content=e['content'],
                received_at=e['received_at'],
                imap_key=e['imap_key'],
                is_processed=auto_skip,
                is_read=e.get('is_seen', False),
                message_id=e.get('message_id', ''),
            )
            # 첨부파일 저장
            for att in e.get('attachments', []):
                file_obj = ContentFile(att['data'], name=att['filename'])
                InboxAttachment.objects.create(
                    message=msg_obj,
                    file=file_obj,
                    filename=att['filename'],
                    content_type=att['content_type'],
                    size=len(att['data']),
                )
            new_count += 1

    sync_msg = f' (읽음 상태 {sync_count}건 동기화)' if sync_count else ''
    messages.success(request, f'새 메일 {new_count}건을 가져왔습니다.{sync_msg}')
    return redirect('inbox_list')


@role_required('admin')
def inbox_process(request, pk):
    """수신 메시지를 보면서 주문 등록"""
    inbox_msg = get_object_or_404(InboxMessage, pk=pk)

    # 열 때 읽음 처리
    if not inbox_msg.is_read:
        inbox_msg.is_read = True
        inbox_msg.save(update_fields=['is_read'])
        # IMAP에서도 읽음 표시
        if inbox_msg.source == 'email' and inbox_msg.imap_key:
            from django.conf import settings as conf
            from .email_utils import mark_as_read_imap
            # imap_key = "account_label:uid"
            parts = inbox_msg.imap_key.split(':', 1)
            if len(parts) == 2:
                label, uid_str = parts
                # 계정 매핑
                account_map = {}
                if hasattr(conf, 'NAVER_EMAIL_1_ID'):
                    account_map['007bm'] = (conf.NAVER_EMAIL_1_ID, conf.NAVER_EMAIL_1_PW)
                if hasattr(conf, 'NAVER_EMAIL_2_ID'):
                    account_map['002bm'] = (conf.NAVER_EMAIL_2_ID, conf.NAVER_EMAIL_2_PW)
                creds = account_map.get(label)
                if creds:
                    mark_as_read_imap(creds[0], creds[1], uid_str)

    # 처리 완료 (주문 없이)
    if request.method == 'POST' and 'skip' in request.POST:
        inbox_msg.is_processed = True
        inbox_msg.save(update_fields=['is_processed'])
        messages.info(request, '처리 완료되었습니다.')
        return redirect('inbox_list')

    agencies = User.objects.filter(role='agency', is_active=True).order_by('name')
    agencies_json = json.dumps([{
        'id': a.pk, 'name': a.name,
    } for a in agencies], ensure_ascii=False)

    teachers = (
        User.objects.filter(role='teacher', is_active=True)
        .select_related('agency', 'delivery_address')
        .order_by('agency__name', 'name')
    )
    teachers_json = json.dumps([{
        'id': t.pk,
        'name': t.name,
        'phone': t.phone or '',
        'agency_id': t.agency_id,
        'delivery_id': t.delivery_address_id,
        'delivery_name': t.delivery_address.name if t.delivery_address else '',
        'delivery_address': t.delivery_address.address if t.delivery_address else '',
        'delivery_phone': t.delivery_address.phone if t.delivery_address else '',
        'has_delivery': bool(t.delivery_address),
    } for t in teachers], ensure_ascii=False)

    books = Book.objects.filter(is_active=True).select_related('publisher')
    series_list = sorted(set(b.series for b in books if b.series))
    books_json = json.dumps([{
        'id': b.id,
        'series': b.series or '기타',
        'name': b.name,
        'publisher': b.publisher.name,
        'unit_price': math.floor(b.list_price * float(b.publisher.supply_rate) / 100),
    } for b in books], ensure_ascii=False)

    if request.method == 'POST' and 'skip' not in request.POST:
        agency_id = request.POST.get('agency_id', '').strip()
        teacher_id = request.POST.get('teacher_id', '').strip()
        new_teacher_name = request.POST.get('new_teacher_name', '').strip()
        new_teacher_phone = request.POST.get('new_teacher_phone', '').strip()
        delivery_school = request.POST.get('delivery_school', '').strip()
        delivery_address_val = request.POST.get('delivery_address', '').strip()
        delivery_phone = request.POST.get('delivery_phone', '').strip()

        # 업체 확인
        try:
            agency = User.objects.get(pk=agency_id, role='agency', is_active=True)
        except (User.DoesNotExist, ValueError):
            messages.error(request, '업체를 선택해 주세요.')
            return redirect('inbox_process', pk=pk)

        # 선생님: 기존 선택 또는 신규 생성
        if teacher_id:
            try:
                teacher = User.objects.select_related('delivery_address').get(
                    pk=teacher_id, role='teacher', is_active=True
                )
            except (User.DoesNotExist, ValueError):
                messages.error(request, '선생님을 선택해 주세요.')
                return redirect('inbox_process', pk=pk)
        elif new_teacher_name:
            login_id = f'a_{new_teacher_phone or "nophone"}_{agency.pk}'
            if User.objects.filter(login_id=login_id).exists():
                teacher = User.objects.get(login_id=login_id)
            else:
                teacher = User(
                    login_id=login_id, role='teacher',
                    name=new_teacher_name, phone=new_teacher_phone,
                    agency=agency, must_change_password=False,
                )
                teacher.set_unusable_password()
                teacher.save()
        else:
            messages.error(request, '선생님을 선택하거나 새로 입력해 주세요.')
            return redirect('inbox_process', pk=pk)

        # 배송지
        if delivery_school:
            delivery, created = DeliveryAddress.objects.get_or_create(
                agency=agency, name=delivery_school,
                defaults={'address': delivery_address_val, 'phone': delivery_phone},
            )
            if not created and delivery_address_val:
                delivery.address = delivery_address_val
                delivery.phone = delivery_phone
                delivery.save(update_fields=['address', 'phone'])
            teacher.delivery_address = delivery
            teacher.save(update_fields=['delivery_address'])
        elif not teacher.delivery_address:
            messages.error(request, '배송지를 입력해 주세요.')
            return redirect('inbox_process', pk=pk)

        items = []
        i = 0
        while f'book_{i}' in request.POST:
            book_id = request.POST.get(f'book_{i}', '').strip()
            qty_str = request.POST.get(f'qty_{i}', '').strip()
            if book_id and qty_str:
                try:
                    qty = int(qty_str)
                    if qty > 0:
                        items.append((int(book_id), qty))
                except (ValueError, TypeError):
                    pass
            i += 1

        if not items:
            messages.error(request, '주문할 교재를 1권 이상 선택하세요.')
        else:
            order = Order.objects.create(
                order_no=Order.generate_order_no(),
                agency=agency,
                teacher=teacher,
                delivery=teacher.delivery_address,
                memo=request.POST.get('memo', ''),
            )
            for book_id, qty in items:
                try:
                    book = Book.objects.get(id=book_id, is_active=True)
                    OrderItem(order=order, book=book, quantity=qty).save()
                except Book.DoesNotExist:
                    pass

            inbox_msg.is_processed = True
            inbox_msg.order = order
            inbox_msg.save(update_fields=['is_processed', 'order'])

            messages.success(request, f'주문 등록 완료. 주문번호: {order.order_no}')
            return redirect('inbox_list')

    attachments = inbox_msg.attachments.all()

    return render(request, 'orders/inbox_process.html', {
        'inbox_msg': inbox_msg,
        'agencies': agencies,
        'agencies_json': agencies_json,
        'teachers_json': teachers_json,
        'series_list': series_list,
        'books_json': books_json,
        'attachments': attachments,
    })


@role_required('admin')
def inbox_reply(request, pk):
    """수신 이메일에 답장 발송"""
    inbox_msg = get_object_or_404(InboxMessage, pk=pk)
    if request.method != 'POST' or inbox_msg.source != 'email':
        return redirect('inbox_process', pk=pk)

    reply_body = request.POST.get('reply_body', '').strip()
    if not reply_body:
        messages.warning(request, '답장 내용을 입력하세요.')
        return redirect('inbox_process', pk=pk)

    from django.conf import settings as conf
    from .email_utils import send_reply_email

    # 발신자에서 이메일 주소 추출
    import re as _re
    sender = inbox_msg.sender or ''
    match = _re.search(r'[\w.\-+]+@[\w.\-]+\.\w+', sender)
    to_email = match.group(0) if match else sender

    # 답장 제목
    subj = inbox_msg.subject or ''
    reply_subject = subj if subj.lower().startswith('re:') else f'Re: {subj}'

    # In-Reply-To / References 헤더
    in_reply_to = inbox_msg.message_id or None
    references = inbox_msg.message_id or None

    # 계정 매핑: account_label → (id, pw)
    account_map = {}
    if hasattr(conf, 'NAVER_EMAIL_1_ID'):
        account_map['007bm'] = (conf.NAVER_EMAIL_1_ID, conf.NAVER_EMAIL_1_PW)
    if hasattr(conf, 'NAVER_EMAIL_2_ID'):
        account_map['002bm'] = (conf.NAVER_EMAIL_2_ID, conf.NAVER_EMAIL_2_PW)

    creds = account_map.get(inbox_msg.account_label)
    if not creds:
        # 기본 계정 사용
        creds = (conf.NAVER_EMAIL_1_ID, conf.NAVER_EMAIL_1_PW)

    ok = send_reply_email(
        account_id=creds[0],
        account_pw=creds[1],
        to_email=to_email,
        subject=reply_subject,
        body=reply_body,
        in_reply_to=in_reply_to,
        references=references,
    )

    if ok:
        messages.success(request, f'{to_email}에 답장을 발송했습니다.')
    else:
        messages.error(request, '답장 발송에 실패했습니다. 로그를 확인하세요.')

    return redirect('inbox_process', pk=pk)


@role_required('admin')
def attachment_download(request, pk):
    """첨부파일 다운로드"""
    att = get_object_or_404(InboxAttachment, pk=pk)
    resp = HttpResponse(att.file.read(), content_type=att.content_type or 'application/octet-stream')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{att.filename}"
    return resp


@role_required('admin')
def attachment_preview(request, pk):
    """엑셀 첨부파일 미리보기 (HTML 테이블)"""
    att = get_object_or_404(InboxAttachment, pk=pk)
    if not att.is_excel:
        return HttpResponse('미리보기를 지원하지 않는 파일입니다.', status=400)

    import openpyxl

    try:
        wb = openpyxl.load_workbook(att.file, read_only=True, data_only=True)
    except Exception:
        return HttpResponse('엑셀 파일을 열 수 없습니다.', status=400)

    sheets_html = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        table = f'<h3 style="margin:10px 0 4px;font-size:13px">{ws.title}</h3>'
        table += '<table class="data-table"><thead><tr>'
        for cell in rows[0]:
            table += f'<th>{cell if cell is not None else ""}</th>'
        table += '</tr></thead><tbody>'
        for row in rows[1:200]:  # 최대 200행
            table += '<tr>'
            for cell in row:
                val = f'{cell:,}' if isinstance(cell, (int, float)) and not isinstance(cell, bool) else (cell if cell is not None else '')
                table += f'<td>{val}</td>'
            table += '</tr>'
        table += '</tbody></table>'
        sheets_html.append(table)
    wb.close()

    html = f'''<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{att.filename}</title>
<style>
body {{ font-family:'Malgun Gothic',sans-serif; font-size:12px; padding:12px; background:#f5f5f5; }}
table {{ width:100%; border-collapse:collapse; margin-bottom:16px; background:#fff; }}
th {{ background:#666; color:#fff; padding:5px 8px; border:1px solid #555; text-align:center; white-space:nowrap; }}
td {{ padding:4px 8px; border:1px solid #ddd; }}
tr:nth-child(even) td {{ background:#fafafa; }}
h3 {{ color:#e8720c; }}
</style></head><body>
<h2 style="font-size:14px;margin-bottom:8px">{att.filename}</h2>
{"".join(sheets_html)}
</body></html>'''
    return HttpResponse(html)


from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def sms_webhook(request):
    """
    갤럭시 SMS Forwarder 앱에서 POST로 오는 문자 수신.
    앱 설정: URL = https://도메인/webhook/sms/
    Body format (JSON): {"from":"010-xxxx", "text":"...", "sentStamp":1234567890}
    """
    if request.method != 'POST':
        return HttpResponse('OK')

    from django.utils import timezone
    import datetime

    try:
        data = json.loads(request.body)
        sender  = data.get('from') or data.get('from_number', '')
        content = data.get('text') or data.get('message', '')
        ts      = data.get('sentStamp') or data.get('timestamp')
        if ts:
            received_at = timezone.make_aware(
                datetime.datetime.fromtimestamp(int(ts) / 1000)
            )
        else:
            received_at = timezone.now()

        if content:
            InboxMessage.objects.create(
                source=InboxMessage.Source.SMS,
                sender=sender,
                content=content,
                received_at=received_at,
            )
    except Exception:
        pass

    return HttpResponse('OK')


@role_required('admin')
def export_purchase(request):
    openpyxl, wb = _make_workbook()
    if not wb:
        messages.error(request, 'openpyxl 패키지가 필요합니다: uv add openpyxl')
        return redirect('purchase_order')

    today = date.today()
    date_from = request.GET.get('date_from', today.strftime('%Y-%m-01'))
    date_to = request.GET.get('date_to', today.strftime('%Y-%m-%d'))

    items = OrderItem.objects.filter(
        order__status__in=[Order.Status.PENDING, Order.Status.SHIPPING, Order.Status.DELIVERED],
        order__ordered_at__date__gte=date_from,
        order__ordered_at__date__lte=date_to,
    ).select_related('book', 'book__publisher').order_by('book__publisher__name', 'book__series', 'book__name')

    from collections import defaultdict
    publishers = defaultdict(lambda: defaultdict(lambda: {'name': '', 'series': '', 'qty': 0, 'amount': 0}))
    for item in items:
        pub = item.book.publisher.name
        bid = item.book.pk
        publishers[pub][bid]['name'] = item.book.name
        publishers[pub][bid]['series'] = item.book.series
        publishers[pub][bid]['qty'] += item.quantity
        publishers[pub][bid]['amount'] += item.amount

    for pub_name in sorted(publishers.keys()):
        ws = wb.create_sheet(title=pub_name[:31])
        ws.append(['시리즈', '교재명', '수량', '금액'])
        pub_data = publishers[pub_name]
        for book_data in sorted(pub_data.values(), key=lambda b: (b['series'], b['name'])):
            ws.append([book_data['series'], book_data['name'], book_data['qty'], book_data['amount']])

    if not wb.sheetnames:
        wb.create_sheet('발주집계')
    else:
        # 기본 시트 제거
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'발주집계_{date_from}_{date_to}.xlsx'
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
    return resp
