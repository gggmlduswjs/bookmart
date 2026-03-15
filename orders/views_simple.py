import json
import math
from functools import wraps

from django.contrib import messages
from django.contrib.auth import login as auth_login
from django.shortcuts import render, redirect, get_object_or_404
from django.http import Http404, HttpResponse

from accounts.models import User
from books.models import Book
from .models import Order, OrderItem, DeliveryAddress, LinkAccessLog
from .sms import send_order_confirmation


def _get_agency_or_404(slug):
    """slug 또는 code로 업체 조회, 없거나 비활성이면 404"""
    try:
        # 짧은 코드 먼저 시도
        agency = User.objects.get(agency_code=slug, role='agency', is_active=True)
    except User.DoesNotExist:
        try:
            # 기존 UUID 호환
            agency = User.objects.get(agency_slug=slug, role='agency', is_active=True)
        except (User.DoesNotExist, ValueError):
            raise Http404
    return agency


def _get_session_teacher(request, agency):
    """세션에서 teacher 조회. 없거나 불일치 시 None"""
    teacher_id = request.session.get('simple_teacher_id')
    session_code = request.session.get('simple_agency_code') or request.session.get('simple_agency_slug')
    if not teacher_id or (str(session_code) != str(agency.agency_code) and str(session_code) != str(agency.agency_slug)):
        return None
    try:
        return User.objects.get(pk=teacher_id, role='teacher', agency=agency, is_active=True)
    except User.DoesNotExist:
        return None


def simple_session_required(view_func):
    """세션에 teacher 정보가 없으면 landing으로 리다이렉트"""
    @wraps(view_func)
    def wrapper(request, slug, *args, **kwargs):
        agency = _get_agency_or_404(slug)
        teacher = _get_session_teacher(request, agency)
        if not teacher:
            return redirect('simple_landing', slug=slug)
        request.simple_agency = agency
        request.simple_teacher = teacher
        return view_func(request, slug, *args, **kwargs)
    return wrapper


# ── Landing: 가입/인증 ──────────────────────────────────────────────────────────

def simple_landing(request, slug):
    agency = _get_agency_or_404(slug)

    # 접속 이력 기록 (GET only)
    if request.method == 'GET':
        LinkAccessLog.objects.create(
            agency=agency,
            ip_address=request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', '')).split(',')[0].strip(),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
            action='visit',
        )

    # 이미 세션 있으면 홈(대시보드)으로
    teacher = _get_session_teacher(request, agency)
    if teacher:
        return redirect('simple_home', slug=slug)

    error = None
    active_tab = 'new'

    if request.method == 'POST':
        mode = request.POST.get('mode', 'new')

        if mode == 'check':
            # ── AJAX: 이름+전화번호로 기존 회원 확인 ──
            check_name = request.POST.get('name', '').strip()
            check_phone = request.POST.get('phone', '').strip()
            if check_name and check_phone:
                existing = User.objects.filter(
                    phone=check_phone, name=check_name,
                    role='teacher', agency=agency, is_active=True
                ).first()
                if existing:
                    # 기존 회원 → 바로 홈(대시보드)으로
                    request.session['simple_teacher_id'] = existing.pk
                    request.session['simple_agency_code'] = str(agency.agency_code)
                    auth_login(request, existing, backend='django.contrib.auth.backends.ModelBackend')
                    from django.urls import reverse
                    return HttpResponse(
                        json.dumps({'exists': True, 'redirect': reverse('simple_home', args=[slug])}),
                        content_type='application/json'
                    )
            return HttpResponse(
                json.dumps({'exists': False}),
                content_type='application/json'
            )

        elif mode == 'lookup':
            # ── 주문내역 확인 모드: 이름+전화번호만으로 세션 복구 ──
            active_tab = 'lookup'
            lookup_name = request.POST.get('lookup_name', '').strip()
            lookup_phone = request.POST.get('lookup_phone', '').strip()

            if not lookup_name or not lookup_phone:
                error = '이름과 전화번호를 모두 입력해주세요.'
            else:
                teacher = User.objects.filter(
                    phone=lookup_phone, name=lookup_name,
                    role='teacher', agency=agency, is_active=True
                ).first()

                if not teacher:
                    error = '일치하는 주문 정보를 찾을 수 없습니다. 이름과 전화번호를 확인해주세요.'
                else:
                    request.session['simple_teacher_id'] = teacher.pk
                    request.session['simple_agency_code'] = str(agency.agency_code)
                    auth_login(request, teacher, backend='django.contrib.auth.backends.ModelBackend')
                    return redirect('simple_home', slug=slug)

        else:
            # ── 신규 주문 모드 ──
            name = request.POST.get('name', '').strip()
            phone = request.POST.get('phone', '').strip()
            school = request.POST.get('school', '').strip()
            address_main = request.POST.get('address', '').strip()
            address_detail = request.POST.get('address_detail', '').strip()
            address = f'{address_main} {address_detail}'.strip() if address_main else ''

            if not phone:
                error = '전화번호를 입력해주세요.'
            elif not name:
                error = '이름을 입력해주세요.'
            elif not school:
                error = '학교명을 입력해주세요.'
            elif not address:
                error = '배송주소를 입력해주세요.'
            else:
                # 전화번호로 기존 유저 검색 (같은 업체 소속)
                existing = User.objects.filter(
                    phone=phone, role='teacher', agency=agency, is_active=True
                ).first()

                if existing:
                    teacher = existing
                else:
                    # 신규 유저 생성
                    login_id = f's_{phone}_{agency.pk}'
                    # login_id 중복 방지
                    if User.objects.filter(login_id=login_id).exists():
                        # 이미 존재하면 해당 유저 사용
                        teacher = User.objects.get(login_id=login_id)
                    else:
                        teacher = User(
                            login_id=login_id,
                            role='teacher',
                            name=name,
                            phone=phone,
                            agency=agency,
                            must_change_password=False,
                        )
                        teacher.set_unusable_password()
                        teacher.save()
                        LinkAccessLog.objects.create(
                            agency=agency, teacher=teacher,
                            ip_address=request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', '')).split(',')[0].strip(),
                            user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
                            action='register',
                        )

                # 배송지(학교) 생성 또는 매칭 (기존/신규 모두)
                delivery, _ = DeliveryAddress.objects.get_or_create(
                    agency=agency,
                    name=school,
                    defaults={'address': address, 'phone': phone},
                )
                teacher.delivery_address = delivery
                teacher.save(update_fields=['delivery_address'])

                # 세션에 저장 + Django 로그인
                request.session['simple_teacher_id'] = teacher.pk
                request.session['simple_agency_code'] = str(agency.agency_code)
                auth_login(request, teacher, backend='django.contrib.auth.backends.ModelBackend')
                return redirect('simple_order', slug=slug)

    return render(request, 'simple/landing.html', {
        'agency': agency,
        'error': error,
        'slug': slug,
        'active_tab': active_tab,
    })


# ── 홈 (주문현황 대시보드) ─────────────────────────────────────────────────────

@simple_session_required
def simple_home(request, slug):
    """로그인 후 바로 보이는 대시보드 — 주문현황 한눈에"""
    teacher = request.simple_teacher
    agency = request.simple_agency

    orders = (
        Order.objects.filter(teacher=teacher)
        .select_related('delivery')
        .prefetch_related('items', 'items__book')
        .order_by('-ordered_at')[:20]
    )

    # 상태별 카운트
    counts = {'pending': 0, 'shipping': 0, 'delivered': 0}
    for order in orders:
        if order.status in counts:
            counts[order.status] += 1
        # 항목 정보
        all_items = list(order.items.all())
        for item in all_items:
            item.list_price_amount = item.list_price * item.quantity
        order.list_price_total = sum(item.list_price_amount for item in all_items)
        order.items_list = all_items

    return render(request, 'simple/home.html', {
        'agency': agency,
        'teacher': teacher,
        'orders': orders,
        'counts': counts,
        'slug': slug,
    })


# ── 주문 ────────────────────────────────────────────────────────────────────────

@simple_session_required
def simple_order(request, slug):
    agency = request.simple_agency
    teacher = request.simple_teacher
    delivery = teacher.delivery_address

    # 업체 취급 교재만 필터링 (지정 없으면 전체)
    agency_books = agency.available_books.filter(is_active=True)
    if agency_books.exists():
        books = agency_books.select_related('publisher')
    else:
        books = Book.objects.filter(is_active=True).select_related('publisher')
    series_list = sorted(set(b.series for b in books if b.series))
    books_json = json.dumps([{
        'id': b.id,
        'series': b.series or '기타',
        'name': b.name,
        'publisher': b.publisher.name,
        'list_price': b.list_price,
    } for b in books], ensure_ascii=False)

    error = None

    if request.method == 'POST':
        # 배송지 업데이트
        new_school = request.POST.get('delivery_school', '').strip()
        new_addr = request.POST.get('delivery_address', '').strip()
        new_detail = request.POST.get('delivery_detail', '').strip()
        new_phone = request.POST.get('delivery_phone', '').strip()
        # 상세정보가 있으면 주소에 합침
        if new_detail:
            new_addr = f'{new_addr} ({new_detail})' if new_addr else new_detail
        if not delivery or (new_school and new_school != delivery.name):
            if new_school:
                delivery, _ = DeliveryAddress.objects.get_or_create(
                    agency=agency, name=new_school,
                    defaults={'address': new_addr, 'phone': new_phone},
                )
                if new_addr:
                    delivery.address = new_addr
                if new_phone:
                    delivery.phone = new_phone
                delivery.save()
                teacher.delivery_address = delivery
                teacher.save(update_fields=['delivery_address'])
        else:
            if new_addr and new_addr != delivery.address:
                delivery.address = new_addr
                delivery.save(update_fields=['address'])
            if new_phone and new_phone != delivery.phone:
                delivery.phone = new_phone
                delivery.save(update_fields=['phone'])

        # Save location_detail separately
        if delivery:
            detail_val = request.POST.get('delivery_detail', '').strip()
            if detail_val and delivery.location_detail != detail_val:
                delivery.location_detail = detail_val
                delivery.save(update_fields=['location_detail'])

        items = []
        custom_items = []
        # 일반 교재
        i = 0
        while f'book_{i}' in request.POST or f'custom_name_{i}' in request.POST:
            book_id = request.POST.get(f'book_{i}', '').strip()
            custom_name = request.POST.get(f'custom_name_{i}', '').strip()
            qty_str = request.POST.get(f'qty_{i}', '').strip()
            custom_price_str = request.POST.get(f'custom_price_{i}', '').strip()

            if book_id and qty_str:
                try:
                    qty = int(qty_str)
                    if qty > 0:
                        items.append((int(book_id), qty))
                except (ValueError, TypeError):
                    pass
            elif custom_name and qty_str:
                try:
                    qty = int(qty_str)
                    price = int(custom_price_str) if custom_price_str else 0
                    if qty > 0:
                        custom_items.append((custom_name, qty, price))
                except (ValueError, TypeError):
                    pass
            i += 1

        # 교사용 책 (tb_ 접두사)
        tc_items = []
        tc_custom_items = []
        i = 0
        while f'tb_book_{i}' in request.POST or f'tb_custom_name_{i}' in request.POST:
            book_id = request.POST.get(f'tb_book_{i}', '').strip()
            custom_name = request.POST.get(f'tb_custom_name_{i}', '').strip()
            qty_str = request.POST.get(f'tb_qty_{i}', '').strip()
            custom_price_str = request.POST.get(f'tb_custom_price_{i}', '').strip()

            if book_id and qty_str:
                try:
                    qty = int(qty_str)
                    if qty > 0:
                        tc_items.append((int(book_id), qty))
                except (ValueError, TypeError):
                    pass
            elif custom_name and qty_str:
                try:
                    qty = int(qty_str)
                    price = int(custom_price_str) if custom_price_str else 0
                    if qty > 0:
                        tc_custom_items.append((custom_name, qty, price))
                except (ValueError, TypeError):
                    pass
            i += 1

        if not items and not custom_items and not tc_items and not tc_custom_items:
            error = '주문할 교재를 1권 이상 선택하세요.'
        else:
            requested_delivery_date = request.POST.get('requested_delivery_date', '').strip() or None
            order = Order.objects.create(
                order_no=Order.generate_order_no(),
                agency=agency,
                teacher=teacher,
                delivery=delivery,
                memo=request.POST.get('memo', ''),
                source=Order.Source.SIMPLE,
                requested_delivery_date=requested_delivery_date,
            )
            for book_id, qty in items:
                try:
                    book = Book.objects.get(id=book_id, is_active=True)
                    OrderItem(order=order, book=book, quantity=qty).save()
                except Book.DoesNotExist:
                    pass
            for cname, qty, price in custom_items:
                OrderItem(
                    order=order, book=None,
                    custom_book_name=cname, quantity=qty, unit_price=price,
                ).save()
            # 교사용
            for book_id, qty in tc_items:
                try:
                    book = Book.objects.get(id=book_id, is_active=True)
                    OrderItem(order=order, book=book, quantity=qty, is_teacher_copy=True).save()
                except Book.DoesNotExist:
                    pass
            for cname, qty, price in tc_custom_items:
                OrderItem(
                    order=order, book=None,
                    custom_book_name=cname, quantity=qty, unit_price=price,
                    is_teacher_copy=True,
                ).save()

            LinkAccessLog.objects.create(
                agency=agency, teacher=teacher,
                ip_address=request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', '')).split(',')[0].strip(),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
                action='order',
            )
            send_order_confirmation(order)
            return redirect('simple_confirm', slug=slug, order_id=order.pk)

    # 재주문 복사
    copy_rows_json = '[]'
    copy_id = request.GET.get('copy', '')
    if copy_id:
        try:
            src_order = Order.objects.get(pk=int(copy_id), teacher=teacher)
            copy_rows = []
            for item in src_order.items.select_related('book', 'book__publisher'):
                if item.book:
                    copy_rows.append({
                        'series': item.book.series or '기타',
                        'book_id': str(item.book_id),
                        'qty': item.quantity,
                        'is_custom': False,
                    })
                else:
                    copy_rows.append({
                        'is_custom': True,
                        'custom_name': item.custom_book_name,
                        'qty': item.quantity,
                        'unit_price': item.unit_price,
                    })
            copy_rows_json = json.dumps(copy_rows, ensure_ascii=False)
        except (Order.DoesNotExist, ValueError):
            pass

    # 최근 주문 교재
    recent_items = OrderItem.objects.filter(
        order__teacher=teacher, book__isnull=False
    ).select_related('book').order_by('-order__ordered_at').values_list('book_id', flat=True)
    seen = set()
    recent_book_ids = []
    for bid in recent_items:
        if bid not in seen:
            seen.add(bid)
            recent_book_ids.append(bid)
        if len(recent_book_ids) >= 10:
            break
    recent_books_qs = Book.objects.filter(id__in=recent_book_ids, is_active=True).select_related('publisher')
    recent_books_json = json.dumps([{
        'id': b.id,
        'series': b.series or '기타',
        'name': b.name,
        'publisher': b.publisher.name,
        'list_price': b.list_price,
    } for b in recent_books_qs], ensure_ascii=False)

    return render(request, 'simple/order.html', {
        'agency': agency,
        'teacher': teacher,
        'delivery': delivery,
        'series_list': series_list,
        'books_json': books_json,
        'slug': slug,
        'error': error,
        'copy_rows_json': copy_rows_json,
        'recent_books_json': recent_books_json,
    })


# ── 주문 수정 (배송 전만) ─────────────────────────────────────────────────────

@simple_session_required
def simple_order_edit(request, slug, order_id):
    agency = request.simple_agency
    teacher = request.simple_teacher
    order = get_object_or_404(Order, pk=order_id, teacher=teacher, agency=agency)

    if order.status != Order.Status.PENDING:
        messages.error(request, '배송 시작된 주문은 수정할 수 없습니다.')
        return redirect('simple_home', slug=slug)

    delivery = order.delivery

    # 업체 취급 교재만 필터링
    agency_books = agency.available_books.filter(is_active=True)
    if agency_books.exists():
        books = agency_books.select_related('publisher')
    else:
        books = Book.objects.filter(is_active=True).select_related('publisher')
    series_list = sorted(set(b.series for b in books if b.series))
    books_json = json.dumps([{
        'id': b.id,
        'series': b.series or '기타',
        'name': b.name,
        'publisher': b.publisher.name,
        'list_price': b.list_price,
    } for b in books], ensure_ascii=False)

    error = None

    if request.method == 'POST':
        # 배송지 업데이트
        new_school = request.POST.get('delivery_school', '').strip()
        new_addr = request.POST.get('delivery_address', '').strip()
        new_detail = request.POST.get('delivery_detail', '').strip()
        new_phone = request.POST.get('delivery_phone', '').strip()
        if new_detail:
            new_addr = f'{new_addr} ({new_detail})' if new_addr else new_detail
        if new_school and new_school != delivery.name:
            delivery, _ = DeliveryAddress.objects.get_or_create(
                agency=agency, name=new_school,
                defaults={'address': new_addr, 'phone': new_phone},
            )
            if new_addr:
                delivery.address = new_addr
            if new_phone:
                delivery.phone = new_phone
            delivery.save()
            order.delivery = delivery
        else:
            if new_addr and new_addr != delivery.address:
                delivery.address = new_addr
                delivery.save(update_fields=['address'])
            if new_phone and new_phone != delivery.phone:
                delivery.phone = new_phone
                delivery.save(update_fields=['phone'])
        if delivery:
            detail_val = request.POST.get('delivery_detail', '').strip()
            if detail_val and delivery.location_detail != detail_val:
                delivery.location_detail = detail_val
                delivery.save(update_fields=['location_detail'])

        # 일반 교재
        items = []
        custom_items = []
        i = 0
        while f'book_{i}' in request.POST or f'custom_name_{i}' in request.POST:
            book_id = request.POST.get(f'book_{i}', '').strip()
            custom_name = request.POST.get(f'custom_name_{i}', '').strip()
            qty_str = request.POST.get(f'qty_{i}', '').strip()
            custom_price_str = request.POST.get(f'custom_price_{i}', '').strip()
            if book_id and qty_str:
                try:
                    qty = int(qty_str)
                    if qty > 0:
                        items.append((int(book_id), qty))
                except (ValueError, TypeError):
                    pass
            elif custom_name and qty_str:
                try:
                    qty = int(qty_str)
                    price = int(custom_price_str) if custom_price_str else 0
                    if qty > 0:
                        custom_items.append((custom_name, qty, price))
                except (ValueError, TypeError):
                    pass
            i += 1

        # 교사용
        tc_items = []
        tc_custom_items = []
        i = 0
        while f'tb_book_{i}' in request.POST or f'tb_custom_name_{i}' in request.POST:
            book_id = request.POST.get(f'tb_book_{i}', '').strip()
            custom_name = request.POST.get(f'tb_custom_name_{i}', '').strip()
            qty_str = request.POST.get(f'tb_qty_{i}', '').strip()
            custom_price_str = request.POST.get(f'tb_custom_price_{i}', '').strip()
            if book_id and qty_str:
                try:
                    qty = int(qty_str)
                    if qty > 0:
                        tc_items.append((int(book_id), qty))
                except (ValueError, TypeError):
                    pass
            elif custom_name and qty_str:
                try:
                    qty = int(qty_str)
                    price = int(custom_price_str) if custom_price_str else 0
                    if qty > 0:
                        tc_custom_items.append((custom_name, qty, price))
                except (ValueError, TypeError):
                    pass
            i += 1

        if not items and not custom_items and not tc_items and not tc_custom_items:
            error = '주문할 교재를 1권 이상 선택하세요.'
        else:
            # 기존 항목 삭제 후 재생성
            order.items.all().delete()
            order.memo = request.POST.get('memo', '')
            requested_delivery_date = request.POST.get('requested_delivery_date', '').strip() or None
            order.requested_delivery_date = requested_delivery_date
            order.is_edited = True
            order.save(update_fields=['delivery', 'memo', 'requested_delivery_date', 'is_edited', 'updated_at'])

            for book_id, qty in items:
                try:
                    book = Book.objects.get(id=book_id, is_active=True)
                    OrderItem(order=order, book=book, quantity=qty).save()
                except Book.DoesNotExist:
                    pass
            for cname, qty, price in custom_items:
                OrderItem(order=order, book=None, custom_book_name=cname, quantity=qty, unit_price=price).save()
            for book_id, qty in tc_items:
                try:
                    book = Book.objects.get(id=book_id, is_active=True)
                    OrderItem(order=order, book=book, quantity=qty, is_teacher_copy=True).save()
                except Book.DoesNotExist:
                    pass
            for cname, qty, price in tc_custom_items:
                OrderItem(order=order, book=None, custom_book_name=cname, quantity=qty, unit_price=price, is_teacher_copy=True).save()

            # 관리자에게 수정 알림
            _notify_order_edited(order)
            return redirect('simple_confirm', slug=slug, order_id=order.pk)

    # 기존 항목을 edit_rows / edit_tb_rows 로 분리
    edit_rows = []
    edit_tb_rows = []
    for item in order.items.select_related('book', 'book__publisher'):
        row = {}
        if item.book:
            row = {
                'series': item.book.series or '기타',
                'book_id': str(item.book_id),
                'qty': item.quantity,
                'is_custom': False,
            }
        else:
            row = {
                'is_custom': True,
                'custom_name': item.custom_book_name,
                'qty': item.quantity,
                'unit_price': item.unit_price,
            }
        if item.is_teacher_copy:
            edit_tb_rows.append(row)
        else:
            edit_rows.append(row)

    edit_rows_json = json.dumps(edit_rows, ensure_ascii=False) if edit_rows else '[]'
    edit_tb_rows_json = json.dumps(edit_tb_rows, ensure_ascii=False) if edit_tb_rows else '[]'

    return render(request, 'simple/order.html', {
        'agency': agency,
        'teacher': teacher,
        'delivery': delivery,
        'series_list': series_list,
        'books_json': books_json,
        'slug': slug,
        'error': error,
        'copy_rows_json': edit_rows_json,
        'recent_books_json': '[]',
        'edit_mode': True,
        'edit_order': order,
        'edit_tb_rows_json': edit_tb_rows_json,
    })


def _notify_order_edited(order):
    """주문 수정 시 관리자에게 이메일 알림 (비동기)"""
    import threading

    def _send(order_pk):
        try:
            from django.conf import settings as conf
            from .email_utils import send_reply_email

            o = (Order.objects.select_related('agency', 'teacher', 'delivery')
                 .prefetch_related('items', 'items__book', 'items__book__publisher')
                 .get(pk=order_pk))

            account_id = getattr(conf, 'NAVER_EMAIL_2_ID', '')
            account_pw = getattr(conf, 'NAVER_EMAIL_2_PW', '')
            if not account_id or not account_pw:
                return

            admin_email = f'{account_id}@naver.com'
            items = o.items.all()
            item_lines = []
            total = 0
            for item in items:
                name = item.display_name
                price = item.list_price
                amount = price * item.quantity
                total += amount
                item_lines.append(f'  - {name} x {item.quantity}권 = {amount:,}원')

            subject = f'[북마트 주문수정] {o.order_no} - {o.delivery.name}'
            body = (
                f'*** 주문이 수정되었습니다 ***\n\n'
                f'주문번호: {o.order_no}\n'
                f'업체: {o.agency.name}\n'
                f'선생님: {o.teacher.name} ({o.teacher.phone})\n'
                f'배송지: {o.delivery.name}\n'
                f'주소: {o.delivery.address}\n'
                f'\n── 수정된 교재 목록 ──\n'
                + '\n'.join(item_lines) +
                f'\n──────────\n'
                f'합계: {total:,}원 (정가 기준)\n'
            )
            if o.memo:
                body += f'메모: {o.memo}\n'

            send_reply_email(account_id, account_pw, admin_email, subject, body)
        except Exception:
            pass

    threading.Thread(target=_send, args=(order.pk,), daemon=True).start()


# ── 배송현황 ──────────────────────────────────────────────────────────────────

@simple_session_required
def simple_delivery_status(request, slug):
    teacher = request.simple_teacher
    orders = (
        Order.objects.filter(teacher=teacher)
        .select_related('delivery')
        .prefetch_related('items', 'items__book')
        .order_by('-ordered_at')
    )
    for order in orders:
        all_items = order.items.all()
        order.list_price_total = sum(item.list_price * item.quantity for item in all_items)
        count = all_items.count()
        first = all_items.first()
        if first:
            name = first.display_name
            order.items_summary = f'{name} 외 {count - 1}건' if count > 1 else name
        else:
            order.items_summary = '-'

    return render(request, 'simple/delivery_status.html', {
        'agency': request.simple_agency,
        'orders': orders,
        'slug': slug,
    })


# ── 엑셀 파싱 (간편주문용, 세션 인증) ────────────────────────────────────────────

def simple_parse_excel(request, slug):
    """간편주문에서 엑셀 업로드 시 교재 매칭 JSON 반환 (세션 인증)"""
    agency = _get_agency_or_404(slug)
    teacher = _get_session_teacher(request, agency)
    if not teacher:
        return HttpResponse(json.dumps({'error': '세션이 만료되었습니다.'}),
                            content_type='application/json', status=403)

    return _do_parse_excel(request)


def _do_parse_excel(request):
    """엑셀 파싱 핵심 로직 (인증 없이)"""
    from openpyxl import load_workbook
    import re

    if request.method != 'POST' or not request.FILES.get('file'):
        return HttpResponse(json.dumps({'error': '파일이 없습니다.'}),
                            content_type='application/json', status=400)

    file = request.FILES['file']
    if not file.name.endswith(('.xlsx', '.xls')):
        return HttpResponse(json.dumps({'error': '.xlsx 파일만 지원합니다.'}),
                            content_type='application/json', status=400)

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        books = Book.objects.filter(is_active=True).select_related('publisher')
        book_map = {}
        for b in books:
            book_map[b.name.strip()] = {
                'id': b.id,
                'series': b.series or '기타',
                'name': b.name,
                'publisher': b.publisher.name,
                'list_price': b.list_price,
            }

        def clean_book_name(name):
            name = re.sub(r'^[^)]+\)\s*', '', name)
            return name.strip()

        def normalize(text):
            t = text.strip()
            t = re.sub(r'[\s\-_·•:：/\\&+<>()（）【】\[\]「」『』]', '', t)
            return t.lower()

        def try_match(name):
            if name in book_map:
                return book_map[name]
            cleaned = clean_book_name(name)
            if cleaned != name and cleaned in book_map:
                return book_map[cleaned]
            norm_name = normalize(cleaned)
            for bname, binfo in book_map.items():
                if normalize(bname) == norm_name:
                    return binfo
            best = None
            best_ratio = 0
            for bname, binfo in book_map.items():
                norm_b = normalize(bname)
                shorter = min(len(norm_name), len(norm_b))
                longer = max(len(norm_name), len(norm_b))
                if shorter < 3 or longer == 0:
                    continue
                if shorter / longer < 0.5:
                    continue
                if norm_name in norm_b or norm_b in norm_name:
                    ratio = shorter / longer
                    if ratio > best_ratio:
                        best = binfo
                        best_ratio = ratio
            return best

        def is_skip_row(text):
            t = text.strip()
            if not t:
                return True
            if re.match(r'^[●•◆■□▶▷※★☆\-\*]?\s*(출판사|샘플|합계|소계|총합계)', t):
                return True
            if re.match(r'출판사\s*[:：]', t):
                return True
            if re.match(r'(주소|연락처|전화|핸드폰|휴대폰|팩스|이메일|메일|E-?mail)\s*[:：]', t, re.IGNORECASE):
                return True
            if re.match(r'^[\d\-\s\(\)]+$', t):
                return True
            if re.search(r'01[016789][\-\s]?\d{3,4}[\-\s]?\d{4}', t):
                return True
            if re.search(r'(서울|부산|대구|인천|광주|대전|울산|세종|경기|강원|충북|충남|전북|전남|경북|경남|제주)', t):
                if re.search(r'(시|구|군|읍|면|동|로|길)\s', t):
                    return True
            if t in ('NO.', 'NO', '비고', '합계', '총합계', '소계'):
                return True
            return False

        def parse_qty(val):
            if val is None:
                return None
            s = str(val).strip()
            m = re.match(r'^(\d+)\s*(권|부|개|세트)?$', s)
            if m:
                v = int(m.group(1))
                return v if 0 < v < 10000 else None
            try:
                v = int(float(s))
                return v if 0 < v < 10000 else None
            except (ValueError, TypeError):
                return None

        all_rows = list(ws.iter_rows(values_only=True))

        # 헤더 탐색
        header_row = -1
        col_name = -1
        col_qty = -1
        col_price = -1
        name_keywords = ('교재명', '도서명', '교재')
        qty_keywords = ('수량', '부수', '권수', '주문수량', '신청수량')
        price_keywords = ('단가', '정가', '가격')

        for i, row in enumerate(all_rows):
            cells = [str(c or '').strip().replace(' ', '') for c in row]
            for j, cell in enumerate(cells):
                cell_clean = cell.split('\n')[0].strip()
                if col_name < 0 and cell_clean in name_keywords:
                    col_name = j
                if col_qty < 0:
                    if cell_clean in qty_keywords or any(cell_clean.startswith(k) for k in qty_keywords):
                        col_qty = j
                if col_price < 0 and cell_clean in price_keywords:
                    col_price = j
            if col_name >= 0:
                header_row = i
                break

        matched = []
        unmatched = []

        if header_row >= 0 and col_name >= 0:
            for row in all_rows[header_row + 1:]:
                cells = list(row)
                if col_name >= len(cells):
                    continue
                name = str(cells[col_name] or '').strip()
                if not name or is_skip_row(name):
                    continue
                qty = 1
                if col_qty >= 0 and col_qty < len(cells):
                    q = parse_qty(cells[col_qty])
                    if q:
                        qty = q
                if qty <= 0:
                    continue
                excel_price = 0
                if col_price >= 0 and col_price < len(cells):
                    try:
                        excel_price = int(float(str(cells[col_price] or '0').replace(',', '')))
                    except (ValueError, TypeError):
                        pass
                info = try_match(name)
                if info:
                    matched.append({**info, 'qty': qty})
                else:
                    unmatched.append({'name': name, 'qty': qty, 'excel_price': excel_price})
        else:
            for row in all_rows:
                cells = [str(c or '').strip() for c in row if c is not None]
                name_candidate = None
                qty_candidate = 1
                for cell in cells:
                    q = parse_qty(cell)
                    if q is not None and cell != name_candidate:
                        qty_candidate = q
                    elif len(cell) >= 2 and cell not in ('', 'NO.', 'NO') and not is_skip_row(cell):
                        name_candidate = cell
                if name_candidate:
                    info = try_match(name_candidate)
                    if info:
                        matched.append({**info, 'qty': qty_candidate})
                    else:
                        unmatched.append({'name': name_candidate, 'qty': qty_candidate})

        wb.close()
        return HttpResponse(
            json.dumps({'matched': matched, 'unmatched': unmatched}, ensure_ascii=False),
            content_type='application/json',
        )
    except Exception as e:
        return HttpResponse(
            json.dumps({'error': f'파일 처리 오류: {str(e)}'}, ensure_ascii=False),
            content_type='application/json', status=400,
        )


# ── 주문 확인 ───────────────────────────────────────────────────────────────────

@simple_session_required
def simple_confirm(request, slug, order_id):
    teacher = request.simple_teacher
    order = get_object_or_404(Order, pk=order_id, teacher=teacher)
    items = order.items.select_related('book', 'book__publisher')
    for item in items:
        item.list_price_amount = item.list_price * item.quantity
    list_price_total = sum(item.list_price_amount for item in items)

    return render(request, 'simple/confirm.html', {
        'agency': request.simple_agency,
        'order': order,
        'items': items,
        'list_price_total': f'{list_price_total:,}',
        'slug': slug,
    })


# ── 주문 내역 ───────────────────────────────────────────────────────────────────

@simple_session_required
def simple_order_list(request, slug):
    teacher = request.simple_teacher
    orders = Order.objects.filter(teacher=teacher).prefetch_related('items').order_by('-ordered_at')
    for order in orders:
        order.list_price_total = sum(item.list_price * item.quantity for item in order.items.all())

    return render(request, 'simple/order_list.html', {
        'agency': request.simple_agency,
        'orders': orders,
        'slug': slug,
    })


# ── 비밀번호 설정 ─────────────────────────────────────────────────────────────

@simple_session_required
def simple_set_password(request, slug):
    """선생님이 비밀번호를 설정하면 정식 로그인 가능"""
    agency = request.simple_agency
    teacher = request.simple_teacher

    if request.method == 'POST':
        pw1 = request.POST.get('password1', '')
        pw2 = request.POST.get('password2', '')

        if not pw1:
            messages.error(request, '비밀번호를 입력해주세요.')
        elif len(pw1) < 4:
            messages.error(request, '비밀번호는 4자 이상이어야 합니다.')
        elif pw1 != pw2:
            messages.error(request, '비밀번호가 일치하지 않습니다.')
        else:
            teacher.set_password(pw1)
            teacher.must_change_password = False
            teacher.save(update_fields=['password', 'must_change_password'])
            # 비밀번호 변경 후 재로그인
            auth_login(request, teacher, backend='django.contrib.auth.backends.ModelBackend')
            messages.success(request, '비밀번호가 설정되었습니다. 이제 정식 로그인 페이지에서도 로그인할 수 있습니다.')
            return redirect('simple_order', slug=slug)

    has_password = teacher.has_usable_password()
    return render(request, 'simple/set_password.html', {
        'agency': agency,
        'slug': slug,
        'teacher': teacher,
        'has_password': has_password,
    })
