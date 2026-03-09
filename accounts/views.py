import uuid

from django.contrib import messages
from django.contrib.auth import login as auth_login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView, PasswordChangeView
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook

from .decorators import role_required
from .forms import (LoginForm, CustomPasswordChangeForm, AgencyForm,
                    TeacherForm, DeliveryAddressForm, generate_temp_password)
from .models import User, AgencyInfo, InviteToken
from orders.models import DeliveryAddress
from orders.sms import send_sms


class CustomLoginView(LoginView):
    form_class = LoginForm
    template_name = 'auth/login.html'

    def get_success_url(self):
        return reverse_lazy('home')


class CustomPasswordChangeView(PasswordChangeView):
    form_class = CustomPasswordChangeForm
    template_name = 'auth/password_change.html'
    success_url = reverse_lazy('home')

    def form_valid(self, form):
        response = super().form_valid(form)
        self.request.user.must_change_password = False
        self.request.user.save(update_fields=['must_change_password'])
        messages.success(self.request, '비밀번호가 변경되었습니다.')
        return response


# ── 총판 전용 ─────────────────────────────────────────────────────────────────

@role_required('admin')
def agency_list(request):
    show_inactive = request.GET.get('inactive') == '1'
    if show_inactive:
        agencies = User.objects.filter(role='agency').order_by('name')
    else:
        agencies = User.objects.filter(role='agency', is_active=True).order_by('name')
    site_url = request.build_absolute_uri('/').rstrip('/')
    return render(request, 'accounts/agency_list.html', {
        'agencies': agencies,
        'site_url': site_url,
        'show_inactive': show_inactive,
    })


@role_required('admin')
def agency_create(request):
    if request.method == 'POST':
        form = AgencyForm(request.POST)
        if form.is_valid():
            temp_pw = generate_temp_password()
            agency = form.save(commit=False)
            agency.role = 'agency'
            agency.set_password(temp_pw)
            agency.must_change_password = True
            agency.save()
            simple_link = request.build_absolute_uri(
                reverse('simple_landing', args=[agency.agency_code])
            )
            site_url = request.build_absolute_uri('/').rstrip('/')
            return render(request, 'accounts/credential.html', {
                'title': '업체 계정 생성 완료',
                'login_id': agency.login_id,
                'password': temp_pw,
                'simple_link': simple_link,
                'site_url': site_url,
                'back_url': reverse('agency_list'),
            })
    else:
        form = AgencyForm()
    return render(request, 'accounts/agency_form.html', {'form': form, 'title': '업체 계정 추가'})


@role_required('admin')
def agency_edit(request, pk):
    """업체 상세정보 수정"""
    from .models import AgencyInfo
    agency_user = get_object_or_404(User, pk=pk, role='agency')
    info, created = AgencyInfo.objects.get_or_create(user=agency_user)

    if request.method == 'POST':
        agency_user.name = request.POST.get('name', agency_user.name).strip()
        agency_user.phone = request.POST.get('phone', agency_user.phone).strip()
        new_pw = request.POST.get('new_password', '').strip()
        update_fields = ['name', 'phone']
        if new_pw:
            agency_user.set_password(new_pw)
            agency_user.must_change_password = True
            update_fields.extend(['password', 'must_change_password'])
        agency_user.save(update_fields=update_fields)

        info.rep_name = request.POST.get('rep_name', '').strip()
        info.biz_no = request.POST.get('biz_no', '').strip()
        info.fax = request.POST.get('fax', '').strip()
        info.postal_code = request.POST.get('postal_code', '').strip()
        info.address = request.POST.get('address', '').strip()
        info.save()

        messages.success(request, f'{agency_user.name} 업체 정보를 수정했습니다.')
        return redirect('agency_list')

    return render(request, 'accounts/agency_edit.html', {
        'agency_user': agency_user,
        'info': info,
    })


@role_required('admin')
def agency_reset_password(request, pk):
    agency = get_object_or_404(User, pk=pk, role='agency')
    temp_pw = generate_temp_password()
    agency.set_password(temp_pw)
    agency.must_change_password = True
    agency.save(update_fields=['password', 'must_change_password'])
    return JsonResponse({'password': temp_pw, 'name': agency.name})


@role_required('admin')
def agency_toggle(request, pk):
    agency = get_object_or_404(User, pk=pk, role='agency')
    agency.is_active = not agency.is_active
    agency.save(update_fields=['is_active'])
    if not agency.is_active:
        agency.teachers.update(is_active=False)
    action = '활성화' if agency.is_active else '비활성화'
    messages.success(request, f'{agency.name} 계정이 {action}되었습니다.')
    return redirect('agency_list')


# ── 업체 전용 — 배송지(학교) ─────────────────────────────────────────────────

@role_required('agency')
def delivery_list(request):
    deliveries = DeliveryAddress.objects.filter(agency=request.user).order_by('name')
    return render(request, 'accounts/delivery_list.html', {'deliveries': deliveries})


@role_required('agency')
def delivery_create(request):
    if request.method == 'POST':
        form = DeliveryAddressForm(request.POST)
        if form.is_valid():
            DeliveryAddress.objects.create(
                agency=request.user,
                name=form.cleaned_data['name'],
                address=form.cleaned_data['address'],
                phone=form.cleaned_data['phone'],
            )
            messages.success(request, f"학교 '{form.cleaned_data['name']}'이 등록되었습니다.")
            return redirect('delivery_list')
    else:
        form = DeliveryAddressForm()
    return render(request, 'accounts/delivery_form.html', {'form': form, 'title': '학교 등록'})


# ── 업체 전용 — 선생님 ────────────────────────────────────────────────────────

@role_required('agency')
def teacher_list(request):
    teachers = User.objects.filter(
        role='teacher', agency=request.user
    ).select_related('delivery_address')
    return render(request, 'accounts/teacher_list.html', {'teachers': teachers})


@role_required('agency')
def teacher_create(request):
    if request.method == 'POST':
        form = TeacherForm(request.user, request.POST)
        if form.is_valid():
            temp_pw = generate_temp_password()
            teacher = form.save(commit=False)
            teacher.role = 'teacher'
            teacher.agency = request.user
            teacher.set_password(temp_pw)
            teacher.must_change_password = True
            teacher.save()

            # 초대 링크 생성 후 바로 보여주기
            invite = InviteToken.create_for_user(teacher)
            invite_url = request.build_absolute_uri(
                reverse('invite_setup', args=[invite.token])
            )
            return render(request, 'accounts/invite_link.html', {
                'teacher': teacher, 'invite_url': invite_url,
                'is_new': True,
            })
    else:
        form = TeacherForm(request.user)
    return render(request, 'accounts/teacher_form.html', {'form': form, 'title': '선생님 계정 등록'})


@role_required('agency')
def teacher_reset_password(request, pk):
    teacher = get_object_or_404(User, pk=pk, role='teacher', agency=request.user)
    temp_pw = generate_temp_password()
    teacher.set_password(temp_pw)
    teacher.must_change_password = True
    teacher.save()
    return render(request, 'accounts/credential.html', {
        'title': f'{teacher.name} 비밀번호 초기화',
        'login_id': teacher.login_id,
        'password': temp_pw,
        'back_url': reverse('teacher_list'),
    })


@role_required('agency')
def teacher_toggle(request, pk):
    teacher = get_object_or_404(User, pk=pk, role='teacher', agency=request.user)
    teacher.is_active = not teacher.is_active
    teacher.save(update_fields=['is_active'])
    action = '활성화' if teacher.is_active else '비활성화'
    messages.success(request, f'{teacher.name} 계정이 {action}되었습니다.')
    return redirect('teacher_list')


# ── 업체 엑셀 일괄 등록 ───────────────────────────────────────────────────────

@role_required('admin')
def agency_import(request):
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file or not file.name.endswith('.xlsx'):
            messages.error(request, '.xlsx 파일만 업로드할 수 있습니다.')
            return redirect('agency_import')

        try:
            wb = load_workbook(file, read_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(min_row=2, values_only=True))
            created_list = []
            skipped = 0

            for row in rows:
                if not row or len(row) < 7:
                    continue

                agency_name = str(row[0] or '').strip()
                login_id = str(row[1] or '').strip()
                password = str(row[2] or '').strip()
                rep_name = str(row[3] or '').strip()
                biz_no = str(row[4] or '').strip()
                phone = str(row[5] or '').strip()
                address = str(row[6] or '').strip()

                if not agency_name or not login_id:
                    continue

                if User.objects.filter(login_id=login_id).exists():
                    skipped += 1
                    continue

                if not password:
                    password = generate_temp_password()

                user = User(
                    login_id=login_id,
                    role='agency',
                    name=agency_name,
                    phone=phone,
                    must_change_password=True,
                )
                user.set_password(password)
                user.save()

                AgencyInfo.objects.create(
                    user=user,
                    rep_name=rep_name,
                    biz_no=biz_no,
                    address=address,
                )

                created_list.append({
                    'name': agency_name,
                    'login_id': login_id,
                    'password': password,
                })

            wb.close()

            if created_list:
                messages.success(
                    request,
                    f'업체 {len(created_list)}건 등록 완료'
                    + (f' (중복 {skipped}건 건너뜀)' if skipped else ''),
                )
            else:
                messages.warning(request, '등록된 업체가 없습니다. 데이터를 확인해주세요.')

            return render(request, 'accounts/agency_import.html', {
                'created_list': created_list,
            })

        except Exception as e:
            messages.error(request, f'파일 처리 중 오류가 발생했습니다: {e}')
            return redirect('agency_import')

    return render(request, 'accounts/agency_import.html')


# ── 초대 링크 ─────────────────────────────────────────────────────────────────

@require_POST
@role_required('agency')
def invite_send(request, pk):
    """선생님에게 초대 링크를 문자 또는 이메일로 발송"""
    teacher = get_object_or_404(User, pk=pk, role='teacher', agency=request.user)
    method = request.POST.get('method', 'sms')  # sms or email

    # 기존 미사용 토큰 무효화
    teacher.invite_tokens.filter(used_at__isnull=True).update(
        expires_at=timezone.now()
    )

    # 새 토큰 생성
    invite = InviteToken.create_for_user(teacher)
    invite_url = request.build_absolute_uri(
        reverse('invite_setup', args=[invite.token])
    )

    if method == 'sms':
        if not teacher.phone:
            messages.error(request, f'{teacher.name} 선생님의 연락처가 등록되지 않았습니다.')
            return redirect('teacher_list')
        msg = (
            f'[북마트] {teacher.name} 선생님\n'
            f'주문 계정이 생성되었습니다.\n'
            f'아래 링크에서 비밀번호를 설정해 주세요.\n'
            f'{invite_url}'
        )
        ok = send_sms(teacher.phone, msg)
        if ok:
            messages.success(request, f'{teacher.name} 선생님에게 초대 문자를 발송했습니다.')
        else:
            messages.warning(
                request,
                f'문자 발송에 실패했습니다. 링크를 직접 전달해 주세요.'
            )
            return render(request, 'accounts/invite_link.html', {
                'teacher': teacher, 'invite_url': invite_url,
            })
    elif method == 'email':
        # TODO: 이메일 발송 구현 (Django send_mail)
        messages.info(request, '이메일 발송은 준비 중입니다. 아래 링크를 직접 전달해 주세요.')
        return render(request, 'accounts/invite_link.html', {
            'teacher': teacher, 'invite_url': invite_url,
        })
    else:
        # 링크만 생성 (직접 전달용)
        return render(request, 'accounts/invite_link.html', {
            'teacher': teacher, 'invite_url': invite_url,
        })

    return redirect('teacher_list')


def invite_setup(request, token):
    """초대 링크를 통한 비밀번호 설정 (로그인 불필요)"""
    invite = get_object_or_404(InviteToken, token=token)

    if not invite.is_valid:
        return render(request, 'accounts/invite_expired.html')

    user = invite.user

    if request.method == 'POST':
        pw1 = request.POST.get('password1', '')
        pw2 = request.POST.get('password2', '')

        if len(pw1) < 4:
            error = '비밀번호는 4자 이상이어야 합니다.'
        elif pw1 != pw2:
            error = '비밀번호가 일치하지 않습니다.'
        else:
            user.set_password(pw1)
            user.must_change_password = False
            user.save(update_fields=['password', 'must_change_password'])

            invite.used_at = timezone.now()
            invite.save(update_fields=['used_at'])

            # 자동 로그인
            auth_login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            messages.success(request, f'{user.name} 선생님, 환영합니다! 비밀번호가 설정되었습니다.')
            return redirect('home')

        return render(request, 'accounts/invite_setup.html', {
            'token': token, 'user': user, 'error': error,
        })

    return render(request, 'accounts/invite_setup.html', {
        'token': token, 'user': user,
    })


@role_required('agency')
def invite_link(request, pk):
    """초대 링크만 생성해서 보여주기 (직접 카톡 등으로 전달)"""
    teacher = get_object_or_404(User, pk=pk, role='teacher', agency=request.user)

    # 기존 미사용 토큰 무효화
    teacher.invite_tokens.filter(used_at__isnull=True).update(
        expires_at=timezone.now()
    )

    invite = InviteToken.create_for_user(teacher)
    invite_url = request.build_absolute_uri(
        reverse('invite_setup', args=[invite.token])
    )

    return render(request, 'accounts/invite_link.html', {
        'teacher': teacher, 'invite_url': invite_url,
    })


# ── 업체 간편주문 링크 ────────────────────────────────────────────────────────

@role_required('agency')
def agency_link(request):
    user = request.user
    link_url = request.build_absolute_uri(
        reverse('simple_landing', args=[user.agency_code])
    )
    return render(request, 'accounts/agency_link.html', {
        'link_url': link_url,
    })


@require_POST
@role_required('agency')
def agency_regenerate_slug(request):
    user = request.user
    user.agency_slug = uuid.uuid4()
    user.save(update_fields=['agency_slug'])
    messages.success(request, '간편주문 링크가 재생성되었습니다. 기존 링크는 더 이상 사용할 수 없습니다.')
    return redirect('agency_link')


@role_required('admin')
def agency_import_sample(request):
    wb = Workbook()
    ws = wb.active
    ws.title = '업체 일괄등록'

    headers = ['업체명', '아이디', '비밀번호', '대표자명', '사업자번호', '연락처', '주소']
    ws.append(headers)

    ws.append(['한빛서점', 'hanbit01', 'pass1234', '김대표', '123-45-67890', '010-1234-5678', '서울시 강남구 역삼동 123'])
    ws.append(['새롬북스', 'saerom01', 'pass5678', '이대표', '987-65-43210', '010-9876-5432', '부산시 해운대구 우동 456'])

    for col in range(1, 8):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="agency_import_sample.xlsx"'
    wb.save(response)
    return response
